from __future__ import annotations

from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
import patsy
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 08_blockwise_khb_detailed_explanation.py
#
# 목적
# 1) Model 1 -> Model 2에서 공공부문 계수가 왜 변하는지,
#    동일한 완전사례 표본에서 블록별 중첩 로지스틱 회귀로 점검
#
# 2) KHB(Karlson-Holm-Breen) 잔차화 방식으로,
#    비선형 로짓모형의 scale/rescaling 문제를 분리한
#    "추가 통제변수 기여"를 진단
#
# 매우 중요한 해석 원칙
# - 이 스크립트는 '무엇이 공공부문 효과를 인과적으로 매개했는가'를
#   검증하지 않습니다.
# - 연령, 교육, 직업, 조직규모, 디지털 역량은 공공부문 고용의
#   원인·결과·구성요인이 혼재할 수 있습니다.
# - 따라서 KHB 결과는 "동일 표본에서 Model 1->2 계수 변화가
#   어떤 통제변수 블록과 결부되는가"를 보여주는
#   rescaling-adjusted diagnostic으로만 사용합니다.
#
# 대상 결과변수:
# - detailed_explanation
#
# 메인 Model 2와의 일치:
# - country FE
# - age, gender, education_age
# - occupation_code, workplace_size, digital_skill_job
# - country-clustered robust standard errors
# ============================================================


# ------------------------------------------------------------
# 0. 경로
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = RESULTS_DIR = (
    PROJECT_ROOT / "results" / "08_blockwise_khb_detailed_explanation"
)
OUTPUT_XLSX = OUTPUT_DIR / "08_blockwise_khb_detailed_explanation.xlsx"


# ------------------------------------------------------------
# 1. 분석 정의
# ------------------------------------------------------------

OUTCOME = "detailed_explanation"
OUTCOME_LABEL = "상세 설명 제공"

BASE_TERMS = "public_sector + C(country_fe)"

# 이 순서는 인구학적 특성 -> 교육 -> 직업 위치 -> 조직 환경 -> 디지털 역량의
# 이론적 순서를 따릅니다. 순차 KHB 결과는 이 순서에 영향을 받습니다.
CONTROL_BLOCKS = [
    {
        "model": "M1",
        "block_ko": "연령 및 성별",
        "block_en": "Age and gender",
        "terms": "age + C(gender)",
    },
    {
        "model": "M2",
        "block_ko": "교육수준",
        "block_en": "Education",
        "terms": "education_age",
    },
    {
        "model": "M3",
        "block_ko": "직업 유형",
        "block_en": "Occupation",
        "terms": "C(occupation_code)",
    },
    {
        "model": "M4",
        "block_ko": "조직 규모",
        "block_en": "Workplace size",
        "terms": "C(workplace_size)",
    },
    {
        "model": "M5",
        "block_ko": "직무 디지털 역량",
        "block_en": "Digital skill",
        "terms": "digital_skill_job",
    },
]

REQUIRED_COLUMNS = [
    OUTCOME,
    "public_sector",
    "country_fe",
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]


# ------------------------------------------------------------
# 2. 공통 함수
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def format_or_ci(or_value: float, ci_low: float, ci_high: float) -> str:
    return f"{or_value:.3f} [{ci_low:.3f}, {ci_high:.3f}]"


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 60))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

        for row_number in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 24

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. 동일 완전사례 표본 생성
# ------------------------------------------------------------

def prepare_complete_case_sample(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Model 1부터 Model 2(M5)까지 N이 완전히 동일하도록,
    최종 Model 2에 필요한 모든 변수가 유효한 응답자만 유지합니다.

    이 처리 없이는 Model 1 -> Model 2 변화가 통제변수 효과인지,
    표본 변화인지 분리할 수 없습니다.
    """
    model_df = dataframe[REQUIRED_COLUMNS].copy()

    for column in [
        OUTCOME,
        "public_sector",
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    model_df = model_df[
        model_df[OUTCOME].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
    ].copy()

    model_df = model_df.dropna().copy()

    model_df[OUTCOME] = to_int64(model_df[OUTCOME])
    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in ["age", "education_age", "digital_skill_job"]:
        model_df[column] = to_float64(model_df[column])

    for column in ["gender", "occupation_code", "workplace_size"]:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


# ------------------------------------------------------------
# 4. 블록별 중첩 로지스틱 회귀
# ------------------------------------------------------------

def fit_formula_logit(
    model_df: pd.DataFrame,
    rhs_terms: str,
):
    """
    국가 단위 cluster-robust standard errors를 적용한 logit.

    반환값:
    - result
    - converged
    - warning text
    """
    formula = f"{OUTCOME} ~ {rhs_terms}"
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.logit(
            formula=formula,
            data=model_df,
            missing="raise",
        )

        result = model.fit(
            disp=False,
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    converged = getattr(result, "mle_retvals", {}).get(
        "converged",
        np.nan,
    )

    return result, converged, " | ".join(warning_messages), formula


def extract_blockwise_summary(
    result,
    model_label: str,
    added_block: str,
    rhs_terms: str,
    model_df: pd.DataFrame,
    converged,
    warning_text: str,
    previous_beta: float | None,
    baseline_beta: float | None,
) -> dict:
    parameter = "public_sector"

    beta = float(result.params[parameter])
    standard_error = float(result.bse[parameter])
    p_value = float(result.pvalues[parameter])

    ci = result.conf_int().loc[parameter]
    ci_low = float(ci.iloc[0])
    ci_high = float(ci.iloc[1])

    public_0 = model_df.copy()
    public_1 = model_df.copy()
    public_0["public_sector"] = 0
    public_1["public_sector"] = 1

    adjusted_probability_nonpublic = float(result.predict(public_0).mean())
    adjusted_probability_public = float(result.predict(public_1).mean())

    return {
        "Model": model_label,
        "Added block": added_block,
        "RHS terms": rhs_terms,
        "N": int(result.nobs),
        "Countries": int(model_df["country_fe"].nunique()),
        "Converged": converged,
        "Warning": warning_text,
        "Public-sector beta (logit)": round(beta, 4),
        "SE": round(standard_error, 4),
        "OR": round(math.exp(beta), 3),
        "OR 95% CI lower": round(math.exp(ci_low), 3),
        "OR 95% CI upper": round(math.exp(ci_high), 3),
        "OR [95% CI]": format_or_ci(
            math.exp(beta),
            math.exp(ci_low),
            math.exp(ci_high),
        ),
        "p-value": p_value,
        "p-value label": p_value_label(p_value),
        "Significance": significance_mark(p_value),
        "Adjusted Pr(non-public, %)": round(
            adjusted_probability_nonpublic * 100,
            2,
        ),
        "Adjusted Pr(public, %)": round(
            adjusted_probability_public * 100,
            2,
        ),
        "AAPD (public - non-public, pp)": round(
            (adjusted_probability_public - adjusted_probability_nonpublic)
            * 100,
            2,
        ),
        "Naive delta beta vs prior": (
            np.nan
            if previous_beta is None
            else round(beta - previous_beta, 4)
        ),
        "Naive delta beta vs M0": (
            np.nan
            if baseline_beta is None
            else round(beta - baseline_beta, 4)
        ),
    }


def extract_all_coefficients(
    result,
    model_label: str,
    added_block: str,
) -> pd.DataFrame:
    confidence_intervals = result.conf_int()

    records = []

    for parameter in result.params.index:
        beta = float(result.params[parameter])
        standard_error = float(result.bse[parameter])
        p_value = float(result.pvalues[parameter])

        ci_low = float(confidence_intervals.loc[parameter].iloc[0])
        ci_high = float(confidence_intervals.loc[parameter].iloc[1])

        records.append(
            {
                "Model": model_label,
                "Added block": added_block,
                "Parameter": parameter,
                "Beta": beta,
                "SE": standard_error,
                "OR": math.exp(beta),
                "OR 95% CI lower": math.exp(ci_low),
                "OR 95% CI upper": math.exp(ci_high),
                "p-value": p_value,
                "p-value label": p_value_label(p_value),
                "Significance": significance_mark(p_value),
            }
        )

    return pd.DataFrame(records)


# ------------------------------------------------------------
# 5. KHB residualization decomposition
# ------------------------------------------------------------

def design_matrix(rhs_terms: str, model_df: pd.DataFrame) -> pd.DataFrame:
    """
    patsy로 설계행렬을 만듭니다.
    RHS에 자동으로 절편을 추가합니다.
    """
    matrix = patsy.dmatrix(
        f"1 + {rhs_terms}",
        data=model_df,
        return_type="dataframe",
    )

    return matrix.astype(float)


def added_control_matrix(
    added_terms: str,
    model_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    KHB에서 residualize할 새 통제변수 설계행렬.
    절편은 제거하고 treatment-coded dummy만 남깁니다.
    """
    matrix = patsy.dmatrix(
        f"1 + {added_terms}",
        data=model_df,
        return_type="dataframe",
    )

    matrix = matrix.drop(columns="Intercept", errors="ignore")

    return matrix.astype(float)


def residualize_added_controls(
    base_matrix: pd.DataFrame,
    added_matrix: pd.DataFrame,
) -> pd.DataFrame:
    """
    각 추가 통제변수를 축소모형의 모든 변수에 OLS로 회귀한 뒤
    잔차를 반환합니다.

    KHB의 핵심 절차입니다.
    """
    residuals = pd.DataFrame(index=added_matrix.index)

    for column in added_matrix.columns:
        ols_result = sm.OLS(
            added_matrix[column].astype(float),
            base_matrix.astype(float),
        ).fit()

        residuals[f"resid__{column}"] = ols_result.resid.astype(float)

    return residuals


def fit_matrix_logit(
    y: pd.Series,
    exog: pd.DataFrame,
    country_groups: pd.Series,
):
    """
    행렬 기반 logit.
    모든 KHB 모형에서 국가 단위 cluster-robust SE를 사용합니다.
    """
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = sm.Logit(
            endog=y.astype(float),
            exog=exog.astype(float),
        )

        result = model.fit(
            disp=False,
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": country_groups.to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    converged = getattr(result, "mle_retvals", {}).get(
        "converged",
        np.nan,
    )

    return result, converged, " | ".join(warning_messages)


def khb_decompose(
    model_df: pd.DataFrame,
    reduced_terms: str,
    added_terms: str,
    stage: str,
    stage_type: str,
) -> dict:
    """
    KHB rescaling decomposition:

    Raw reduced:
      Y ~ X

    Rescaled reduced:
      Y ~ X + residualized(W | X)

    Full:
      Y ~ X + W

    여기서:
    - Raw reduced -> rescaled reduced 차이 = scaling/rescaling component
    - Rescaled reduced -> full 차이 = added-covariate component
    """
    y = model_df[OUTCOME]
    country_groups = model_df["country_fe"]

    x_reduced = design_matrix(
        rhs_terms=reduced_terms,
        model_df=model_df,
    )

    w_added = added_control_matrix(
        added_terms=added_terms,
        model_df=model_df,
    )

    w_residualized = residualize_added_controls(
        base_matrix=x_reduced,
        added_matrix=w_added,
    )

    x_rescaled = pd.concat(
        [x_reduced, w_residualized],
        axis=1,
    )

    x_full = pd.concat(
        [x_reduced, w_added],
        axis=1,
    )

    raw_result, raw_converged, raw_warning = fit_matrix_logit(
        y=y,
        exog=x_reduced,
        country_groups=country_groups,
    )

    rescaled_result, rescaled_converged, rescaled_warning = (
        fit_matrix_logit(
            y=y,
            exog=x_rescaled,
            country_groups=country_groups,
        )
    )

    full_result, full_converged, full_warning = fit_matrix_logit(
        y=y,
        exog=x_full,
        country_groups=country_groups,
    )

    parameter = "public_sector"

    beta_raw = float(raw_result.params[parameter])
    beta_rescaled = float(rescaled_result.params[parameter])
    beta_full = float(full_result.params[parameter])

    # KHB point components on logit scale
    scaling_component = beta_raw - beta_rescaled
    added_covariate_component = beta_rescaled - beta_full
    naive_change = beta_raw - beta_full

    percent_of_rescaled_total = np.nan

    if abs(beta_rescaled) > 1e-10:
        percent_of_rescaled_total = (
            added_covariate_component / beta_rescaled
        ) * 100

    return {
        "Stage": stage,
        "Stage type": stage_type,
        "Reduced specification": reduced_terms,
        "Added covariate block": added_terms,
        "N": len(model_df),
        "Countries": int(model_df["country_fe"].nunique()),
        "Raw reduced converged": raw_converged,
        "Rescaled reduced converged": rescaled_converged,
        "Full converged": full_converged,
        "Warnings": " | ".join(
            item
            for item in [raw_warning, rescaled_warning, full_warning]
            if item
        ),
        "Raw reduced beta": round(beta_raw, 6),
        "Rescaled reduced beta": round(beta_rescaled, 6),
        "Full beta": round(beta_full, 6),
        "Naive raw-to-full change": round(naive_change, 6),
        "Scaling / rescaling component": round(scaling_component, 6),
        "Added-covariate component": round(
            added_covariate_component,
            6,
        ),
        "Added-covariate component (% of rescaled beta)": round(
            percent_of_rescaled_total,
            2,
        ),
        "Raw reduced OR": round(math.exp(beta_raw), 3),
        "Rescaled reduced OR": round(math.exp(beta_rescaled), 3),
        "Full OR": round(math.exp(beta_full), 3),
        "Direction note": (
            "Adding this block makes the public-sector coefficient more negative"
            if added_covariate_component > 0
            else (
                "Adding this block makes the public-sector coefficient less negative"
                if added_covariate_component < 0
                else "No coefficient change"
            )
        ),
    }


# ------------------------------------------------------------
# 6. 실행
# ------------------------------------------------------------

def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            "전처리 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {INPUT_PATH}"
        )

    raw_df = pd.read_parquet(INPUT_PATH)

    missing_columns = sorted(
        set(REQUIRED_COLUMNS) - set(raw_df.columns)
    )

    if missing_columns:
        raise KeyError(
            "clean_data.parquet에 필요한 변수가 없습니다:\n"
            + "\n".join(f"- {column}" for column in missing_columns)
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    model_df = prepare_complete_case_sample(raw_df)

    print("=" * 76)
    print("상세 설명 제공: 블록별 회귀 + KHB 분해 분석 시작")
    print(f"동일 완전사례 표본 N: {len(model_df):,}")
    print(f"국가 수: {model_df['country_fe'].nunique()}")
    print("=" * 76)

    # --------------------------------------------------------
    # A. Blockwise nested models
    # --------------------------------------------------------

    blockwise_rows = []
    coefficient_tables = []

    current_terms = BASE_TERMS
    previous_beta = None
    baseline_beta = None

    model_specs = [
        {
            "model": "M0",
            "block_ko": "국가 고정효과만",
            "terms": BASE_TERMS,
        }
    ]

    for block in CONTROL_BLOCKS:
        current_terms = f"{current_terms} + {block['terms']}"

        model_specs.append(
            {
                "model": block["model"],
                "block_ko": block["block_ko"],
                "terms": current_terms,
            }
        )

    for spec in model_specs:
        print(f"블록별 로지스틱 회귀 추정: {spec['model']}")

        result, converged, warning_text, formula = fit_formula_logit(
            model_df=model_df,
            rhs_terms=spec["terms"],
        )

        beta = float(result.params["public_sector"])

        if spec["model"] == "M0":
            baseline_beta = beta

        blockwise_rows.append(
            extract_blockwise_summary(
                result=result,
                model_label=spec["model"],
                added_block=spec["block_ko"],
                rhs_terms=spec["terms"],
                model_df=model_df,
                converged=converged,
                warning_text=warning_text,
                previous_beta=previous_beta,
                baseline_beta=baseline_beta,
            )
        )

        coefficient_tables.append(
            extract_all_coefficients(
                result=result,
                model_label=spec["model"],
                added_block=spec["block_ko"],
            )
        )

        previous_beta = beta

    blockwise_table = pd.DataFrame(blockwise_rows)
    full_coefficients_table = pd.concat(
        coefficient_tables,
        ignore_index=True,
    )

    # --------------------------------------------------------
    # B. KHB: all controls jointly
    # --------------------------------------------------------

    print("KHB 전체 통제변수 분해 추정")

    all_control_terms = " + ".join(
        block["terms"] for block in CONTROL_BLOCKS
    )

    khb_all_controls = pd.DataFrame(
        [
            khb_decompose(
                model_df=model_df,
                reduced_terms=BASE_TERMS,
                added_terms=all_control_terms,
                stage="All controls jointly",
                stage_type="Joint KHB decomposition",
            )
        ]
    )

    # --------------------------------------------------------
    # C. KHB: sequential theoretical blocks
    # --------------------------------------------------------

    print("KHB 순차 블록 분해 추정")

    khb_sequential_rows = []
    prior_terms = BASE_TERMS

    for block in CONTROL_BLOCKS:
        khb_sequential_rows.append(
            khb_decompose(
                model_df=model_df,
                reduced_terms=prior_terms,
                added_terms=block["terms"],
                stage=block["block_ko"],
                stage_type="Sequential KHB decomposition",
            )
        )

        prior_terms = f"{prior_terms} + {block['terms']}"

    khb_sequential_table = pd.DataFrame(khb_sequential_rows)

    # --------------------------------------------------------
    # D. KHB: country FE 기준, 각 블록 단독 추가
    # --------------------------------------------------------

    print("KHB 단일 블록 비교 추정")

    khb_single_block_rows = []

    for block in CONTROL_BLOCKS:
        khb_single_block_rows.append(
            khb_decompose(
                model_df=model_df,
                reduced_terms=BASE_TERMS,
                added_terms=block["terms"],
                stage=block["block_ko"],
                stage_type="Single-block KHB contrast",
            )
        )

    khb_single_block_table = pd.DataFrame(khb_single_block_rows)

    # --------------------------------------------------------
    # E. Readme / interpretation guide
    # --------------------------------------------------------

    guide_table = pd.DataFrame(
        [
            {
                "Item": "Blockwise models",
                "Interpretation": (
                    "M0-M5 use the same complete-case sample. "
                    "Observe which added block changes the public-sector "
                    "coefficient, OR, and adjusted probability difference."
                ),
            },
            {
                "Item": "KHB all controls jointly",
                "Interpretation": (
                    "Use this to distinguish the raw reduced-to-full change "
                    "from the rescaling component and the added-covariate component."
                ),
            },
            {
                "Item": "KHB sequential blocks",
                "Interpretation": (
                    "Shows incremental contributions in the stated theoretical order. "
                    "Results are order-dependent and should not be interpreted as "
                    "causal mediation."
                ),
            },
            {
                "Item": "KHB single-block contrasts",
                "Interpretation": (
                    "Shows how each block changes the coefficient when added "
                    "to the country-FE baseline alone. These contrasts are not additive."
                ),
            },
            {
                "Item": "Manuscript wording",
                "Interpretation": (
                    "Use 'rescaling-adjusted decomposition of the coefficient change' "
                    "or 'composition-related contribution', not 'causal mediation'."
                ),
            },
        ]
    )

    # --------------------------------------------------------
    # F. Save
    # --------------------------------------------------------

    tables = {
        "01_BlockwiseSummary": blockwise_table,
        "02_BlockwiseCoefficients": full_coefficients_table,
        "03_KHB_AllControls": khb_all_controls,
        "04_KHB_Sequential": khb_sequential_table,
        "05_KHB_SingleBlock": khb_single_block_table,
        "06_InterpretationGuide": guide_table,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, table in tables.items():
            table.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    for sheet_name, table in tables.items():
        save_csv(
            dataframe=table,
            file_name=f"{sheet_name}.csv",
        )

    format_excel(OUTPUT_XLSX)

    # --------------------------------------------------------
    # G. Console summary
    # --------------------------------------------------------

    print("\n[Blockwise public-sector effect]")
    print(
        blockwise_table[
            [
                "Model",
                "Added block",
                "N",
                "Public-sector beta (logit)",
                "OR [95% CI]",
                "p-value label",
                "Significance",
                "AAPD (public - non-public, pp)",
                "Naive delta beta vs prior",
            ]
        ].to_string(index=False)
    )

    print("\n[KHB all-controls decomposition]")
    print(khb_all_controls.to_string(index=False))

    print("\n[KHB sequential blocks]")
    print(
        khb_sequential_table[
            [
                "Stage",
                "Raw reduced beta",
                "Rescaled reduced beta",
                "Full beta",
                "Scaling / rescaling component",
                "Added-covariate component",
                "Direction note",
            ]
        ].to_string(index=False)
    )

    print("\n" + "=" * 76)
    print("분석 완료")
    print(f"결과 파일: {OUTPUT_XLSX}")
    print(f"결과 폴더: {OUTPUT_DIR}")
    print("=" * 76)


if __name__ == "__main__":
    main()
