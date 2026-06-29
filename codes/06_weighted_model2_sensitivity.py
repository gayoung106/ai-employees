from __future__ import annotations

from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# 06_weighted_model2_sensitivity.py
#
# 목적:
#   비가중 Model 2와 w1 가중 Model 2를 동일 표본에서 비교하는
#   가중치 민감도 분석입니다.
#
# 주의:
#   - w1은 전처리 단계에서 weight_w1으로 저장되어 있다고 가정합니다.
#   - 본 분석은 국가 고정효과 + 국가 단위 cluster-robust SE를 유지합니다.
#   - 가중분석은 메인 결과가 아니라 민감도 분석으로 사용합니다.
# ============================================================


# ------------------------------------------------------------
# 0. 경로 설정
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "06_weighted_model2_sensitivity"
OUTPUT_XLSX = OUTPUT_DIR / "06_weighted_model2_sensitivity.xlsx"


# ------------------------------------------------------------
# 1. 분석 변수
# ------------------------------------------------------------

OUTCOMES = {
    "detailed_explanation": "상세 설명 제공",
    "personal_data_access": "개인정보 접근권 제공",
    "automated_analysis_access": "자동화 분석 결과 접근권 제공",
    "not_informed": "기술 사용 사실 무고지",
}

WEIGHT_COLUMN = "weight_w1"

CONTROL_COLUMNS = [
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

CATEGORICAL_CONTROL_COLUMNS = [
    "gender",
    "occupation_code",
    "workplace_size",
]

REQUIRED_COLUMNS = [
    "country_fe",
    "public_sector",
    WEIGHT_COLUMN,
    *OUTCOMES.keys(),
    *CONTROL_COLUMNS,
]


# ------------------------------------------------------------
# 2. 보조 함수
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    """Pandas nullable numeric dtype을 NumPy float64로 변환."""
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    """결측 제거 후 범주형 변수를 NumPy int64로 변환."""
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    """Patsy 호환을 위해 일반 object 문자열로 변환."""
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 45))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. Model 2 분석 표본 생성
# ------------------------------------------------------------

def prepare_model2_data(
    dataframe: pd.DataFrame,
    outcome: str,
) -> pd.DataFrame:
    """
    outcome별 Model 2 완전사례 표본을 생성합니다.

    Model 2:
      outcome ~ public_sector + country FE + age + gender
              + education_age + occupation + workplace_size
              + digital_skill_job
    """
    columns = [
        outcome,
        "public_sector",
        "country_fe",
        WEIGHT_COLUMN,
        *CONTROL_COLUMNS,
    ]

    model_df = dataframe[columns].copy()

    # 이항·연속형 변수
    for column in [
        outcome,
        "public_sector",
        WEIGHT_COLUMN,
        "age",
        "education_age",
        "digital_skill_job",
    ]:
        model_df[column] = to_float64(model_df[column])

    # 범주형 통제변수
    for column in CATEGORICAL_CONTROL_COLUMNS:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    # 유효값 범위 필터
    model_df = model_df[
        model_df[outcome].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
        & model_df[WEIGHT_COLUMN].gt(0)
    ].copy()

    model_df = model_df.dropna().copy()

    # 가중치 평균을 1로 정규화:
    # 계수는 바뀌지 않으며, 가중 표본 크기의 해석 혼동을 줄입니다.
    model_df["weight_normalized"] = (
        model_df[WEIGHT_COLUMN]
        / model_df[WEIGHT_COLUMN].mean()
    )

    # statsmodels/patsy 호환을 위한 기본 dtype 변환
    model_df[outcome] = to_int64(model_df[outcome])
    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in [
        WEIGHT_COLUMN,
        "weight_normalized",
        "age",
        "education_age",
        "digital_skill_job",
    ]:
        model_df[column] = to_float64(model_df[column])

    for column in CATEGORICAL_CONTROL_COLUMNS:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


def model2_formula(outcome: str) -> str:
    return (
        f"{outcome} ~ public_sector + C(country_fe)"
        " + age + C(gender) + education_age"
        " + C(occupation_code) + C(workplace_size)"
        " + digital_skill_job"
    )


# ------------------------------------------------------------
# 4. 추정 함수
# ------------------------------------------------------------

def fit_unweighted_logit(
    model_df: pd.DataFrame,
    formula: str,
):
    """
    기존 Model 2와 같은 비가중 Logit.
    국가 단위 cluster-robust SE를 적용합니다.
    """
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.logit(
            formula=formula,
            data=model_df,
            missing="raise",
        )

        result = model.fit(
            disp=False,
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    converged = getattr(result, "mle_retvals", {}).get(
        "converged",
        np.nan,
    )

    return result, converged, " | ".join(warning_messages)


def fit_weighted_glm(
    model_df: pd.DataFrame,
    formula: str,
):
    """
    w1 정규화 가중치를 사용한 이항 GLM.

    sampling weight를 statsmodels에서 직접 설계기반으로 추정하는
    기능은 제한적이므로, 여기서는 weighted pseudo-likelihood
    민감도 분석으로 해석합니다.
    """
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.glm(
            formula=formula,
            data=model_df,
            family=sm.families.Binomial(),
            freq_weights=model_df["weight_normalized"],
            missing="raise",
        )

        result = model.fit(
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    converged = getattr(result, "converged", np.nan)

    return result, converged, " | ".join(warning_messages)


def extract_public_sector_effect(
    result,
    outcome: str,
    outcome_label: str,
    model_type: str,
    model_df: pd.DataFrame,
    converged,
    warning_text: str,
) -> dict:
    """
    public_sector의 OR, 95% CI, 평균 조정예측확률을 추출합니다.
    """
    parameter = "public_sector"

    coefficient = float(result.params[parameter])
    standard_error = float(result.bse[parameter])
    p_value = float(result.pvalues[parameter])

    confidence_interval = result.conf_int().loc[parameter]
    ci_low = float(confidence_interval.iloc[0])
    ci_high = float(confidence_interval.iloc[1])

    non_public_data = model_df.copy()
    public_data = model_df.copy()

    non_public_data["public_sector"] = 0
    public_data["public_sector"] = 1

    predicted_non_public = result.predict(non_public_data)
    predicted_public = result.predict(public_data)

    if model_type == "Weighted GLM (w1 normalized)":
        weights = model_df["weight_normalized"]
        adjusted_non_public = float(
            np.average(predicted_non_public, weights=weights)
        )
        adjusted_public = float(
            np.average(predicted_public, weights=weights)
        )
    else:
        adjusted_non_public = float(predicted_non_public.mean())
        adjusted_public = float(predicted_public.mean())

    return {
        "결과변수": outcome,
        "결과변수 라벨": outcome_label,
        "모형": model_type,
        "표본 N": int(len(model_df)),
        "국가 수": int(model_df["country_fe"].nunique()),
        "수렴 여부": converged,
        "경고 메시지": warning_text,
        "공공부문 계수(log-odds)": round(coefficient, 4),
        "표준오차": round(standard_error, 4),
        "오즈비(OR)": round(math.exp(coefficient), 3),
        "OR 95% CI 하한": round(math.exp(ci_low), 3),
        "OR 95% CI 상한": round(math.exp(ci_high), 3),
        "p-value": p_value,
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
        "비공공부문 조정예측확률(%)": round(
            adjusted_non_public * 100,
            2,
        ),
        "공공부문 조정예측확률(%)": round(
            adjusted_public * 100,
            2,
        ),
        "조정확률 차이(%p, 공공-비공공)": round(
            (adjusted_public - adjusted_non_public) * 100,
            2,
        ),
    }


def make_weight_summary(
    model_df: pd.DataFrame,
    outcome: str,
    outcome_label: str,
) -> dict:
    """outcome별 Model 2 표본에서 가중치 분포를 확인합니다."""
    weights = model_df[WEIGHT_COLUMN]

    return {
        "결과변수": outcome,
        "결과변수 라벨": outcome_label,
        "표본 N": len(model_df),
        "가중치 평균": round(float(weights.mean()), 4),
        "가중치 표준편차": round(float(weights.std(ddof=1)), 4),
        "가중치 최소값": round(float(weights.min()), 4),
        "가중치 P25": round(float(weights.quantile(0.25)), 4),
        "가중치 중앙값": round(float(weights.median()), 4),
        "가중치 P75": round(float(weights.quantile(0.75)), 4),
        "가중치 최대값": round(float(weights.max()), 4),
    }


# ------------------------------------------------------------
# 5. 데이터 로드 및 점검
# ------------------------------------------------------------

if not INPUT_PATH.exists():
    raise FileNotFoundError(
        "전처리 파일을 찾을 수 없습니다.\n"
        f"확인 경로: {INPUT_PATH}"
    )

df = pd.read_parquet(INPUT_PATH)

missing_columns = sorted(
    set(REQUIRED_COLUMNS) - set(df.columns)
)

if missing_columns:
    raise KeyError(
        "clean_data.parquet에 필요한 변수가 없습니다:\n"
        + "\n".join(f"- {column}" for column in missing_columns)
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 72)
print("w1 가중 Model 2 민감도 분석 시작")
print(f"전처리 데이터 행 수: {len(df):,}")
print("=" * 72)


# ------------------------------------------------------------
# 6. 분석 실행
# ------------------------------------------------------------

effect_rows = []
weight_rows = []
status_rows = []

for outcome, outcome_label in OUTCOMES.items():
    print(f"\n분석 중: {outcome}")

    model_df = prepare_model2_data(
        dataframe=df,
        outcome=outcome,
    )

    formula = model2_formula(outcome)

    weight_rows.append(
        make_weight_summary(
            model_df=model_df,
            outcome=outcome,
            outcome_label=outcome_label,
        )
    )

    # 비가중 Model 2: 가중 Model 2와 동일 표본에서 재추정
    try:
        unweighted_result, unweighted_converged, unweighted_warning = (
            fit_unweighted_logit(
                model_df=model_df,
                formula=formula,
            )
        )

        effect_rows.append(
            extract_public_sector_effect(
                result=unweighted_result,
                outcome=outcome,
                outcome_label=outcome_label,
                model_type="Unweighted Model 2 (same N)",
                model_df=model_df,
                converged=unweighted_converged,
                warning_text=unweighted_warning,
            )
        )
    except Exception as error:
        status_rows.append(
            {
                "결과변수": outcome,
                "모형": "Unweighted Model 2 (same N)",
                "상태": "FAILED",
                "사유": f"{type(error).__name__}: {error}",
            }
        )

    # w1 가중 Model 2
    try:
        weighted_result, weighted_converged, weighted_warning = (
            fit_weighted_glm(
                model_df=model_df,
                formula=formula,
            )
        )

        effect_rows.append(
            extract_public_sector_effect(
                result=weighted_result,
                outcome=outcome,
                outcome_label=outcome_label,
                model_type="Weighted GLM (w1 normalized)",
                model_df=model_df,
                converged=weighted_converged,
                warning_text=weighted_warning,
            )
        )
    except Exception as error:
        status_rows.append(
            {
                "결과변수": outcome,
                "모형": "Weighted GLM (w1 normalized)",
                "상태": "FAILED",
                "사유": f"{type(error).__name__}: {error}",
            }
        )


# ------------------------------------------------------------
# 7. 결과 구성
# ------------------------------------------------------------

effects_table = pd.DataFrame(effect_rows)
weight_summary_table = pd.DataFrame(weight_rows)

if status_rows:
    status_table = pd.DataFrame(status_rows)
else:
    status_table = pd.DataFrame(
        columns=["결과변수", "모형", "상태", "사유"]
    )

# 메인 결과의 비교표: 상세 설명 중심
detailed_explanation_table = effects_table.loc[
    effects_table["결과변수"] == "detailed_explanation"
].copy()


# ------------------------------------------------------------
# 8. 저장
# ------------------------------------------------------------

tables = {
    "01_가중치민감도_전체": effects_table,
    "02_상세설명_가중치비교": detailed_explanation_table,
    "03_가중치분포": weight_summary_table,
    "04_오류및경고": status_table,
}

with pd.ExcelWriter(
    OUTPUT_XLSX,
    engine="openpyxl",
) as writer:
    for sheet_name, table in tables.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

for sheet_name, table in tables.items():
    safe_name = (
        sheet_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )

    save_csv(
        dataframe=table,
        file_name=f"{safe_name}.csv",
    )

format_excel(OUTPUT_XLSX)


# ------------------------------------------------------------
# 9. 콘솔 출력
# ------------------------------------------------------------

summary_columns = [
    "결과변수 라벨",
    "모형",
    "표본 N",
    "수렴 여부",
    "오즈비(OR)",
    "OR 95% CI 하한",
    "OR 95% CI 상한",
    "p-value 표기",
    "유의성",
    "비공공부문 조정예측확률(%)",
    "공공부문 조정예측확률(%)",
    "조정확률 차이(%p, 공공-비공공)",
]

print("\n[가중치 민감도 분석: 공공부문 효과]")
print(
    effects_table[
        summary_columns
    ].to_string(index=False)
)

print("\n" + "=" * 72)
print("w1 가중 Model 2 민감도 분석 완료")
print(f"엑셀 파일: {OUTPUT_XLSX}")
print(f"결과 폴더: {OUTPUT_DIR}")
print("=" * 72)
