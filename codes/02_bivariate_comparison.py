from __future__ import annotations

from pathlib import Path
import math

import numpy as np
import pandas as pd
from scipy.stats import chi2_contingency, ttest_ind
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# 02_bivariate_comparison.py
#
# 공공부문 vs 비공공부문 이변량 비교 분석
#
# 프로젝트 구조
# ai-employees/
# ├─ codes/
# │  └─ 02_bivariate_comparison.py
# ├─ clean_data/
# │  └─ clean_data.parquet
# └─ results/
#    └─ 02_bivariate_comparison/
# ============================================================


# ------------------------------------------------------------
# 0. 경로 설정
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "02_bivariate_comparison"
OUTPUT_XLSX = OUTPUT_DIR / "02_bivariate_comparison.xlsx"


# ------------------------------------------------------------
# 1. 분석 변수 정의
# ------------------------------------------------------------

OUTCOME_VARIABLES = {
    "detailed_explanation": "상세 설명 제공",
    "personal_data_access": "개인정보 접근권 제공",
    "automated_analysis_access": "자동화 분석 결과 접근권 제공",
    "not_informed": "기술 사용 사실 무고지",
    "basic_notification_only": "단순 고지",
}

EXPOSURE_VARIABLES = {
    "automated_monitoring_exposure": "자동화 감시 경험",
    "automated_performance_management_exposure": "자동화 성과평가 경험",
}

CONTINUOUS_VARIABLES = {
    "age": "연령",
    "education_age": "전일제 교육 종료 연령",
    "digital_skill_job": "직무 수행 디지털 역량",
}

CATEGORICAL_VARIABLES = {
    "gender": {
        "label": "성별",
        "categories": {
            1: "Man",
            2: "Woman",
            3: "None of the above / Non-binary / Prefer not to say",
        },
    },
    "occupation_code": {
        "label": "직업 유형",
        "categories": {
            10: "Employed professional",
            11: "General management / director / top management",
            12: "Middle management / other management",
            13: "Working mainly at a desk",
            14: "Not at a desk but travelling",
            15: "Service job",
            16: "Supervisor",
            17: "Skilled manual worker",
            18: "Other employed manual worker",
        },
    },
    "workplace_size": {
        "label": "조직 규모",
        "categories": {
            1: "1",
            2: "2-9",
            3: "10-49",
            4: "50-250",
            5: "More than 250",
        },
    },
}


# ------------------------------------------------------------
# 2. 공통 함수
# ------------------------------------------------------------

def as_numeric(series: pd.Series) -> pd.Series:
    """숫자가 아닌 값은 결측치로 변환합니다."""
    return pd.to_numeric(series, errors="coerce")


def p_value_label(p_value: float) -> str:
    """p-value 표시 형식."""
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    """유의성 별표."""
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    """한글이 깨지지 않도록 UTF-8-SIG 형식으로 CSV를 저장합니다."""
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    """생성된 엑셀 파일에 기본 서식을 적용합니다."""
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            max_length = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(text) + 2)

            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(
                max_length,
                42,
            )

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. 이항변수 비교: 카이제곱, Phi, 오즈비
# ------------------------------------------------------------

def compare_binary(
    dataframe: pd.DataFrame,
    variable: str,
    label: str,
) -> dict:
    """
    공공부문(1)과 비공공부문(0)을 비교합니다.

    이항 결과변수는 0/1만 유효값으로 사용합니다.
    """
    work = dataframe[["public_sector", variable]].copy()
    work["public_sector"] = as_numeric(work["public_sector"])
    work[variable] = as_numeric(work[variable])

    work = work[
        work["public_sector"].isin([0, 1])
        & work[variable].isin([0, 1])
    ].copy()

    contingency = pd.crosstab(
        work["public_sector"],
        work[variable],
    ).reindex(
        index=[0, 1],
        columns=[0, 1],
        fill_value=0,
    )

    non_public_no = int(contingency.loc[0, 0])
    non_public_yes = int(contingency.loc[0, 1])
    public_no = int(contingency.loc[1, 0])
    public_yes = int(contingency.loc[1, 1])

    non_public_n = non_public_no + non_public_yes
    public_n = public_no + public_yes
    total_n = non_public_n + public_n

    non_public_pct = (
        non_public_yes / non_public_n * 100
        if non_public_n > 0
        else np.nan
    )
    public_pct = (
        public_yes / public_n * 100
        if public_n > 0
        else np.nan
    )

    chi_square = np.nan
    p_value = np.nan
    phi = np.nan
    minimum_expected = np.nan

    # 전체 응답이 모두 0 또는 모두 1이면 검정할 수 없음
    if (
        non_public_n > 0
        and public_n > 0
        and (non_public_yes + public_yes) > 0
        and (non_public_no + public_no) > 0
    ):
        chi_square, p_value, _, expected = chi2_contingency(
            contingency,
            correction=False,
        )
        phi = math.sqrt(chi_square / total_n)
        minimum_expected = float(expected.min())

    # 오즈비: Public sector / Non-public sector
    # 0 cell이 있으면 0.5 보정을 적용
    a = float(public_yes)
    b = float(public_no)
    c = float(non_public_yes)
    d = float(non_public_no)

    if min(a, b, c, d) == 0:
        a += 0.5
        b += 0.5
        c += 0.5
        d += 0.5

    odds_ratio = (a * d) / (b * c)

    return {
        "변수명": variable,
        "변수 라벨": label,
        "비공공부문 유효 N": non_public_n,
        "비공공부문 예(1) N": non_public_yes,
        "비공공부문 예(1) 비율(%)": round(non_public_pct, 2),
        "공공부문 유효 N": public_n,
        "공공부문 예(1) N": public_yes,
        "공공부문 예(1) 비율(%)": round(public_pct, 2),
        "비율 차이(%p, 공공-비공공)": round(
            public_pct - non_public_pct,
            2,
        ),
        "카이제곱": (
            round(float(chi_square), 3)
            if not pd.isna(chi_square)
            else np.nan
        ),
        "p-value": (
            float(p_value)
            if not pd.isna(p_value)
            else np.nan
        ),
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
        "Phi 효과크기": (
            round(float(phi), 4)
            if not pd.isna(phi)
            else np.nan
        ),
        "공공부문 대비 오즈비": round(float(odds_ratio), 3),
        "최소 기대빈도": (
            round(float(minimum_expected), 2)
            if not pd.isna(minimum_expected)
            else np.nan
        ),
    }


# ------------------------------------------------------------
# 4. 연속형 변수 비교: Welch t-test, Cohen's d
# ------------------------------------------------------------

def compare_continuous(
    dataframe: pd.DataFrame,
    variable: str,
    label: str,
) -> dict:
    """공공부문과 비공공부문의 평균 차이를 비교합니다."""
    work = dataframe[["public_sector", variable]].copy()
    work["public_sector"] = as_numeric(work["public_sector"])
    work[variable] = as_numeric(work[variable])

    work = work[
        work["public_sector"].isin([0, 1])
        & work[variable].notna()
    ].copy()

    non_public = work.loc[
        work["public_sector"] == 0,
        variable,
    ]
    public = work.loc[
        work["public_sector"] == 1,
        variable,
    ]

    non_public_mean = non_public.mean()
    public_mean = public.mean()

    non_public_sd = non_public.std(ddof=1)
    public_sd = public.std(ddof=1)

    t_statistic = np.nan
    p_value = np.nan
    cohens_d = np.nan

    if len(non_public) >= 2 and len(public) >= 2:
        t_statistic, p_value = ttest_ind(
            public,
            non_public,
            equal_var=False,
        )

        pooled_variance = (
            ((len(public) - 1) * public_sd**2)
            + ((len(non_public) - 1) * non_public_sd**2)
        ) / (len(public) + len(non_public) - 2)

        pooled_sd = math.sqrt(pooled_variance)

        if pooled_sd > 0:
            cohens_d = (
                public_mean - non_public_mean
            ) / pooled_sd

    return {
        "변수명": variable,
        "변수 라벨": label,
        "비공공부문 유효 N": len(non_public),
        "비공공부문 평균": round(float(non_public_mean), 2),
        "비공공부문 표준편차": round(float(non_public_sd), 2),
        "공공부문 유효 N": len(public),
        "공공부문 평균": round(float(public_mean), 2),
        "공공부문 표준편차": round(float(public_sd), 2),
        "평균 차이(공공-비공공)": round(
            float(public_mean - non_public_mean),
            2,
        ),
        "Welch t": (
            round(float(t_statistic), 3)
            if not pd.isna(t_statistic)
            else np.nan
        ),
        "p-value": (
            float(p_value)
            if not pd.isna(p_value)
            else np.nan
        ),
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
        "Cohen's d": (
            round(float(cohens_d), 3)
            if not pd.isna(cohens_d)
            else np.nan
        ),
    }


# ------------------------------------------------------------
# 5. 범주형 변수 비교: 분포표, 카이제곱, Cramer's V
# ------------------------------------------------------------

def compare_categorical(
    dataframe: pd.DataFrame,
    variable: str,
    label: str,
    categories: dict[int, str],
) -> tuple[pd.DataFrame, dict]:
    """공공부문과 비공공부문의 범주형 변수 분포를 비교합니다."""
    work = dataframe[["public_sector", variable]].copy()
    work["public_sector"] = as_numeric(work["public_sector"])
    work[variable] = as_numeric(work[variable])

    work = work[
        work["public_sector"].isin([0, 1])
        & work[variable].isin(list(categories.keys()))
    ].copy()

    contingency = pd.crosstab(
        work["public_sector"],
        work[variable],
    ).reindex(
        index=[0, 1],
        columns=list(categories.keys()),
        fill_value=0,
    )

    non_public_total = int(contingency.loc[0].sum())
    public_total = int(contingency.loc[1].sum())

    distribution_rows = []

    for code, category_label in categories.items():
        non_public_n = int(contingency.loc[0, code])
        public_n = int(contingency.loc[1, code])

        non_public_pct = (
            non_public_n / non_public_total * 100
            if non_public_total > 0
            else np.nan
        )
        public_pct = (
            public_n / public_total * 100
            if public_total > 0
            else np.nan
        )

        distribution_rows.append(
            {
                "변수명": variable,
                "변수 라벨": label,
                "코드": code,
                "범주": category_label,
                "비공공부문 N": non_public_n,
                "비공공부문 비율(%)": round(
                    non_public_pct,
                    2,
                ),
                "공공부문 N": public_n,
                "공공부문 비율(%)": round(
                    public_pct,
                    2,
                ),
                "비율 차이(%p, 공공-비공공)": round(
                    public_pct - non_public_pct,
                    2,
                ),
            }
        )

    # 전체 빈도가 0인 범주는 카이제곱 계산에서 제외
    test_table = contingency.loc[
        :,
        contingency.sum(axis=0) > 0,
    ]

    chi_square = np.nan
    p_value = np.nan
    dof = np.nan
    cramers_v = np.nan
    minimum_expected = np.nan

    if test_table.shape[1] >= 2 and test_table.to_numpy().sum() > 0:
        chi_square, p_value, dof, expected = chi2_contingency(
            test_table,
            correction=False,
        )

        denominator = (
            test_table.to_numpy().sum()
            * (min(test_table.shape) - 1)
        )

        if denominator > 0:
            cramers_v = math.sqrt(chi_square / denominator)

        minimum_expected = float(expected.min())

    test_result = {
        "변수명": variable,
        "변수 라벨": label,
        "비공공부문 유효 N": non_public_total,
        "공공부문 유효 N": public_total,
        "카이제곱": (
            round(float(chi_square), 3)
            if not pd.isna(chi_square)
            else np.nan
        ),
        "자유도": dof,
        "p-value": (
            float(p_value)
            if not pd.isna(p_value)
            else np.nan
        ),
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
        "Cramer's V": (
            round(float(cramers_v), 4)
            if not pd.isna(cramers_v)
            else np.nan
        ),
        "최소 기대빈도": (
            round(float(minimum_expected), 2)
            if not pd.isna(minimum_expected)
            else np.nan
        ),
    }

    return pd.DataFrame(distribution_rows), test_result


# ------------------------------------------------------------
# 6. 데이터 불러오기 및 점검
# ------------------------------------------------------------

if not INPUT_PATH.exists():
    raise FileNotFoundError(
        "전처리 파일을 찾을 수 없습니다.\n"
        f"확인 경로: {INPUT_PATH}"
    )

df = pd.read_parquet(INPUT_PATH)

required_columns = [
    "respondent_id",
    "country_fe",
    "public_sector",
    *OUTCOME_VARIABLES.keys(),
    *EXPOSURE_VARIABLES.keys(),
    *CONTINUOUS_VARIABLES.keys(),
    *CATEGORICAL_VARIABLES.keys(),
]

missing_columns = sorted(
    set(required_columns) - set(df.columns)
)

if missing_columns:
    raise KeyError(
        "clean_data.parquet에 필요한 변수가 없습니다:\n"
        + "\n".join(f"- {column}" for column in missing_columns)
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 70)
print("공공부문·비공공부문 이변량 비교 분석 시작")
print(f"분석 데이터 행 수: {len(df):,}")
print("=" * 70)


# ------------------------------------------------------------
# 7. 표본 구성
# ------------------------------------------------------------

public_sector_numeric = as_numeric(df["public_sector"])

sample_summary = pd.DataFrame(
    [
        {
            "집단": "전체",
            "N": len(df),
            "전체 대비 비율(%)": 100.00,
        },
        {
            "집단": "비공공부문",
            "N": int((public_sector_numeric == 0).sum()),
            "전체 대비 비율(%)": round(
                (public_sector_numeric == 0).mean() * 100,
                2,
            ),
        },
        {
            "집단": "공공부문",
            "N": int((public_sector_numeric == 1).sum()),
            "전체 대비 비율(%)": round(
                (public_sector_numeric == 1).mean() * 100,
                2,
            ),
        },
    ]
)


# ------------------------------------------------------------
# 8. 핵심 결과변수 비교
# ------------------------------------------------------------

outcome_table = pd.DataFrame(
    [
        compare_binary(df, variable, label)
        for variable, label in OUTCOME_VARIABLES.items()
    ]
)


# ------------------------------------------------------------
# 9. 자동화 관리 경험 비교
# ------------------------------------------------------------

exposure_table = pd.DataFrame(
    [
        compare_binary(df, variable, label)
        for variable, label in EXPOSURE_VARIABLES.items()
    ]
)


# ------------------------------------------------------------
# 10. 연속형 통제변수 비교
# ------------------------------------------------------------

continuous_table = pd.DataFrame(
    [
        compare_continuous(df, variable, label)
        for variable, label in CONTINUOUS_VARIABLES.items()
    ]
)


# ------------------------------------------------------------
# 11. 범주형 통제변수 비교
# ------------------------------------------------------------

categorical_distribution_list = []
categorical_test_list = []

for variable, variable_info in CATEGORICAL_VARIABLES.items():
    distribution, test_result = compare_categorical(
        df,
        variable,
        variable_info["label"],
        variable_info["categories"],
    )

    categorical_distribution_list.append(distribution)
    categorical_test_list.append(test_result)

categorical_distribution_table = pd.concat(
    categorical_distribution_list,
    ignore_index=True,
)

categorical_test_table = pd.DataFrame(categorical_test_list)


# ------------------------------------------------------------
# 12. 국가별 표본 구성
# ------------------------------------------------------------

country_df = df[["country_fe", "public_sector"]].copy()
country_df["country_fe"] = (
    country_df["country_fe"]
    .astype("string")
    .fillna("Missing")
)
country_df["public_sector"] = as_numeric(country_df["public_sector"])
country_df = country_df[
    country_df["public_sector"].isin([0, 1])
].copy()

country_sector_table = (
    country_df.groupby(
        ["country_fe", "public_sector"],
        dropna=False,
    )
    .size()
    .unstack(fill_value=0)
    .reindex(columns=[0, 1], fill_value=0)
    .reset_index()
)

country_sector_table = country_sector_table.rename(
    columns={
        "country_fe": "국가",
        0: "비공공부문 N",
        1: "공공부문 N",
    }
)

country_sector_table["전체 N"] = (
    country_sector_table["비공공부문 N"]
    + country_sector_table["공공부문 N"]
)

country_sector_table["공공부문 비율(%)"] = (
    country_sector_table["공공부문 N"]
    / country_sector_table["전체 N"]
    * 100
).round(2)

country_sector_table = country_sector_table.sort_values(
    by="국가"
).reset_index(drop=True)


# ------------------------------------------------------------
# 13. 결과 저장
# ------------------------------------------------------------

tables = {
    "00_표본구성": sample_summary,
    "01_핵심결과_차이검정": outcome_table,
    "02_자동화관리경험_차이검정": exposure_table,
    "03_연속형통제_차이검정": continuous_table,
    "04_범주형통제_분포": categorical_distribution_table,
    "05_범주형통제_차이검정": categorical_test_table,
    "06_국가별부문표본": country_sector_table,
}

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    for sheet_name, table in tables.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

for sheet_name, table in tables.items():
    safe_name = (
        sheet_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )

    save_csv(table, f"{safe_name}.csv")

format_excel(OUTPUT_XLSX)


# ------------------------------------------------------------
# 14. 콘솔 요약
# ------------------------------------------------------------

print("\n[핵심 결과변수: 공공부문 vs 비공공부문]")
print(
    outcome_table[
        [
            "변수 라벨",
            "비공공부문 예(1) 비율(%)",
            "공공부문 예(1) 비율(%)",
            "비율 차이(%p, 공공-비공공)",
            "카이제곱",
            "p-value 표기",
            "유의성",
            "Phi 효과크기",
            "공공부문 대비 오즈비",
        ]
    ].to_string(index=False)
)

print("\n[자동화 관리 경험: 공공부문 vs 비공공부문]")
print(
    exposure_table[
        [
            "변수 라벨",
            "비공공부문 예(1) 비율(%)",
            "공공부문 예(1) 비율(%)",
            "비율 차이(%p, 공공-비공공)",
            "카이제곱",
            "p-value 표기",
            "유의성",
            "Phi 효과크기",
        ]
    ].to_string(index=False)
)

print("\n[연속형 통제변수: 공공부문 vs 비공공부문]")
print(
    continuous_table[
        [
            "변수 라벨",
            "비공공부문 평균",
            "공공부문 평균",
            "평균 차이(공공-비공공)",
            "Welch t",
            "p-value 표기",
            "유의성",
            "Cohen's d",
        ]
    ].to_string(index=False)
)

print("\n" + "=" * 70)
print("이변량 비교 분석 완료")
print(f"엑셀 파일: {OUTPUT_XLSX}")
print(f"결과 폴더: {OUTPUT_DIR}")
print("=" * 70)
