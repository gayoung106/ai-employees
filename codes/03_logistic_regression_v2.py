from __future__ import annotations

from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# 03_logistic_regression_v2.py
#
# 공공부문 여부와 정보공개·데이터 접근권의 관계:
# 국가 고정효과 이항 로지스틱 회귀분석
#
# pandas nullable dtype(Int64, string[python])를 statsmodels가
# 읽지 못하는 문제를 피하도록 모든 분석변수를 NumPy 기본 dtype으로
# 강제 변환한 버전입니다.
# ============================================================


# ------------------------------------------------------------
# 0. 경로 설정
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "03_logistic_regression"
OUTPUT_XLSX = OUTPUT_DIR / "03_logistic_regression.xlsx"


# ------------------------------------------------------------
# 1. 변수 정의
# ------------------------------------------------------------

OUTCOMES = {
    "detailed_explanation": {
        "label": "상세 설명 제공",
        "role": "메인 결과변수",
    },
    "personal_data_access": {
        "label": "개인정보 접근권 제공",
        "role": "메인 결과변수",
    },
    "automated_analysis_access": {
        "label": "자동화 분석 결과 접근권 제공",
        "role": "메인 결과변수",
    },
    "not_informed": {
        "label": "기술 사용 사실 무고지",
        "role": "메인 결과변수",
    },
    "basic_notification_only": {
        "label": "단순 고지",
        "role": "보조 결과변수",
    },
}

MODEL_2_CONTROLS = [
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

MODEL_3_EXPOSURES = [
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
]

REQUIRED_COLUMNS = [
    "public_sector",
    "country_fe",
    *OUTCOMES.keys(),
    *MODEL_2_CONTROLS,
    *MODEL_3_EXPOSURES,
]

CATEGORICAL_CONTROL_COLUMNS = [
    "gender",
    "occupation_code",
    "workplace_size",
]

NUMERIC_COLUMNS = [
    "public_sector",
    "age",
    "education_age",
    "digital_skill_job",
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
]


# ------------------------------------------------------------
# 2. 공통 함수
# ------------------------------------------------------------

def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def to_float64(series: pd.Series) -> pd.Series:
    """pandas nullable Int64 등을 NumPy float64로 강제 변환."""
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    """
    결측치가 없는 범주형 변수를 NumPy int64로 강제 변환.
    이 함수 호출 전 결측치를 제거해야 합니다.
    """
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    """pandas StringDtype이 아닌 일반 object 문자열로 강제 변환."""
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(
        values,
        index=series.index,
        dtype=object,
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 42))

            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = width

    workbook.save(excel_path)


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


# ------------------------------------------------------------
# 3. 분석 표본 생성
# ------------------------------------------------------------

def prepare_analysis_data(
    dataframe: pd.DataFrame,
    outcome: str,
    model_number: int,
) -> pd.DataFrame:
    """
    Model 1:
        public_sector + country fixed effects

    Model 2:
        Model 1 + age + gender + education + occupation
        + workplace size + digital skill

    Model 3:
        Model 2 + automated monitoring + automated performance management
    """
    required = [
        outcome,
        "public_sector",
        "country_fe",
    ]

    if model_number >= 2:
        required.extend(MODEL_2_CONTROLS)

    if model_number >= 3:
        required.extend(MODEL_3_EXPOSURES)

    model_df = dataframe[required].copy()

    # 결과변수와 수치형 변수를 NumPy float64로 변환
    model_df[outcome] = to_float64(model_df[outcome])

    for column in NUMERIC_COLUMNS:
        if column in model_df.columns:
            model_df[column] = to_float64(model_df[column])

    # 국가 변수는 일반 object 문자열로 변환
    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    # 범주형 통제변수는 우선 float64로 바꿔 결측치 확인 가능하게 함
    for column in CATEGORICAL_CONTROL_COLUMNS:
        if column in model_df.columns:
            model_df[column] = to_float64(model_df[column])

    # 핵심 이항변수는 0/1만 남김
    model_df = model_df[
        model_df[outcome].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
    ].copy()

    # Model 2 이상: 유효 범주만 유지
    if model_number >= 2:
        model_df = model_df[
            model_df["gender"].isin([1.0, 2.0, 3.0])
            & model_df["occupation_code"].isin(
                [float(x) for x in range(10, 19)]
            )
            & model_df["workplace_size"].isin(
                [1.0, 2.0, 3.0, 4.0, 5.0]
            )
        ].copy()

    # Model 3: 자동화 노출 변수는 0/1만 유지
    if model_number >= 3:
        model_df = model_df[
            model_df["automated_monitoring_exposure"].isin([0.0, 1.0])
            & model_df[
                "automated_performance_management_exposure"
            ].isin([0.0, 1.0])
        ].copy()

    # 연속형 변수의 결측을 포함해 남은 결측치 제거
    model_df = model_df.dropna().copy()

    # statsmodels/patsy 호환을 위해 다시 NumPy 기본 dtype으로 강제
    model_df[outcome] = to_int64(model_df[outcome])
    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in [
        "age",
        "education_age",
        "digital_skill_job",
        "automated_monitoring_exposure",
        "automated_performance_management_exposure",
    ]:
        if column in model_df.columns:
            model_df[column] = to_float64(model_df[column])

    for column in CATEGORICAL_CONTROL_COLUMNS:
        if column in model_df.columns:
            model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    # 설명변수 범주가 하나뿐인 국가가 있더라도 국가 FE 추정은 가능하나,
    # 전체적으로 국가가 최소 두 개 이상인지 확인
    if model_df["country_fe"].nunique() < 2:
        raise ValueError(
            "국가 고정효과를 추정하려면 두 개 이상의 국가가 필요합니다."
        )

    return model_df


# ------------------------------------------------------------
# 4. 모형식 및 적합
# ------------------------------------------------------------

def make_formula(outcome: str, model_number: int) -> str:
    predictors = [
        "public_sector",
        "C(country_fe)",
    ]

    if model_number >= 2:
        predictors.extend(
            [
                "age",
                "C(gender)",
                "education_age",
                "C(occupation_code)",
                "C(workplace_size)",
                "digital_skill_job",
            ]
        )

    if model_number >= 3:
        predictors.extend(
            [
                "automated_monitoring_exposure",
                "automated_performance_management_exposure",
            ]
        )

    return f"{outcome} ~ " + " + ".join(predictors)


def fit_logit_model(
    model_df: pd.DataFrame,
    formula: str,
):
    """
    국가 고정효과 로지스틱 회귀.
    우선 국가 단위 cluster-robust SE를 사용하고,
    실패하면 HC3 robust SE로 대체합니다.
    """
    base_model = smf.logit(
        formula=formula,
        data=model_df,
        missing="raise",
    )

    try:
        result = base_model.fit(
            disp=False,
            maxiter=300,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )
        covariance_type = "Cluster-robust SE by country"

    except Exception as cluster_error:
        warnings.warn(
            "국가 군집-강건 표준오차 추정에 실패하여 "
            "HC3 강건 표준오차로 대체합니다.\n"
            f"원인: {cluster_error}"
        )

        result = base_model.fit(
            disp=False,
            maxiter=300,
            cov_type="HC3",
        )
        covariance_type = "HC3 robust SE"

    return result, covariance_type


# ------------------------------------------------------------
# 5. 결과 추출
# ------------------------------------------------------------

def extract_public_sector_effect(
    result,
    outcome: str,
    outcome_label: str,
    outcome_role: str,
    model_number: int,
    model_df: pd.DataFrame,
    covariance_type: str,
) -> dict:
    parameter = "public_sector"

    coefficient = float(result.params[parameter])
    standard_error = float(result.bse[parameter])
    p_value = float(result.pvalues[parameter])

    conf_int = result.conf_int().loc[parameter]
    ci_low = float(conf_int.iloc[0])
    ci_high = float(conf_int.iloc[1])

    odds_ratio = math.exp(coefficient)
    or_low = math.exp(ci_low)
    or_high = math.exp(ci_high)

    # Average adjusted predictions:
    # 모든 관측치에서 public_sector만 0/1로 바꾼 뒤 예측값 평균
    non_public_data = model_df.copy()
    public_data = model_df.copy()

    non_public_data["public_sector"] = 0
    public_data["public_sector"] = 1

    predicted_non_public = float(
        result.predict(non_public_data).mean()
    )
    predicted_public = float(
        result.predict(public_data).mean()
    )

    return {
        "결과변수": outcome,
        "결과변수 라벨": outcome_label,
        "결과변수 구분": outcome_role,
        "모형": f"Model {model_number}",
        "표본 N": int(result.nobs),
        "국가 수": int(model_df["country_fe"].nunique()),
        "공분산 추정": covariance_type,
        "공공부문 계수(log-odds)": round(coefficient, 4),
        "표준오차": round(standard_error, 4),
        "오즈비(OR)": round(odds_ratio, 3),
        "OR 95% CI 하한": round(or_low, 3),
        "OR 95% CI 상한": round(or_high, 3),
        "p-value": p_value,
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
        "비공공부문 조정예측확률(%)": round(
            predicted_non_public * 100,
            2,
        ),
        "공공부문 조정예측확률(%)": round(
            predicted_public * 100,
            2,
        ),
        "조정확률 차이(%p, 공공-비공공)": round(
            (predicted_public - predicted_non_public) * 100,
            2,
        ),
        "Pseudo R²": round(float(result.prsquared), 4),
        "Log-Likelihood": round(float(result.llf), 2),
        "AIC": round(float(result.aic), 2),
        "BIC": round(float(result.bic), 2),
    }


def extract_all_coefficients(
    result,
    outcome: str,
    outcome_label: str,
    model_number: int,
    covariance_type: str,
) -> pd.DataFrame:
    confidence_intervals = result.conf_int()

    rows = []

    for parameter in result.params.index:
        coefficient = float(result.params[parameter])
        standard_error = float(result.bse[parameter])
        p_value = float(result.pvalues[parameter])

        conf_int = confidence_intervals.loc[parameter]
        ci_low = float(conf_int.iloc[0])
        ci_high = float(conf_int.iloc[1])

        rows.append(
            {
                "결과변수": outcome,
                "결과변수 라벨": outcome_label,
                "모형": f"Model {model_number}",
                "공분산 추정": covariance_type,
                "변수": parameter,
                "계수(log-odds)": round(coefficient, 4),
                "표준오차": round(standard_error, 4),
                "오즈비(OR)": round(math.exp(coefficient), 3),
                "OR 95% CI 하한": round(math.exp(ci_low), 3),
                "OR 95% CI 상한": round(math.exp(ci_high), 3),
                "p-value": p_value,
                "p-value 표기": p_value_label(p_value),
                "유의성": significance_mark(p_value),
            }
        )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 6. 데이터 불러오기 및 점검
# ------------------------------------------------------------

if not INPUT_PATH.exists():
    raise FileNotFoundError(
        "전처리 파일을 찾을 수 없습니다.\n"
        f"확인 경로: {INPUT_PATH}"
    )

df = pd.read_parquet(INPUT_PATH)

missing_columns = sorted(
    set(REQUIRED_COLUMNS) - set(df.columns)
)

if missing_columns:
    raise KeyError(
        "clean_data.parquet에 필요한 변수가 없습니다:\n"
        + "\n".join(f"- {column}" for column in missing_columns)
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 72)
print("국가 고정효과 로지스틱 회귀분석 시작")
print(f"전처리 데이터 행 수: {len(df):,}")
print("=" * 72)


# ------------------------------------------------------------
# 7. 회귀분석 실행
# ------------------------------------------------------------

effect_rows = []
coefficient_tables = []
model_sample_rows = []

for outcome, outcome_info in OUTCOMES.items():
    for model_number in [1, 2, 3]:
        print(f"\n분석 중: {outcome} | Model {model_number}")

        model_df = prepare_analysis_data(
            dataframe=df,
            outcome=outcome,
            model_number=model_number,
        )

        formula = make_formula(
            outcome=outcome,
            model_number=model_number,
        )

        if len(model_df) < 100:
            raise ValueError(
                f"{outcome}, Model {model_number}: "
                f"유효 표본이 너무 적습니다 (N={len(model_df)})."
            )

        result, covariance_type = fit_logit_model(
            model_df=model_df,
            formula=formula,
        )

        effect_rows.append(
            extract_public_sector_effect(
                result=result,
                outcome=outcome,
                outcome_label=outcome_info["label"],
                outcome_role=outcome_info["role"],
                model_number=model_number,
                model_df=model_df,
                covariance_type=covariance_type,
            )
        )

        coefficient_tables.append(
            extract_all_coefficients(
                result=result,
                outcome=outcome,
                outcome_label=outcome_info["label"],
                model_number=model_number,
                covariance_type=covariance_type,
            )
        )

        model_sample_rows.append(
            {
                "결과변수": outcome,
                "결과변수 라벨": outcome_info["label"],
                "모형": f"Model {model_number}",
                "회귀식": formula,
                "표본 N": int(result.nobs),
                "국가 수": int(model_df["country_fe"].nunique()),
                "공공부문 N": int(
                    (model_df["public_sector"] == 1).sum()
                ),
                "비공공부문 N": int(
                    (model_df["public_sector"] == 0).sum()
                ),
                "공분산 추정": covariance_type,
            }
        )


# ------------------------------------------------------------
# 8. 결과표 구성
# ------------------------------------------------------------

public_sector_effects = pd.DataFrame(effect_rows)

main_public_sector_effects = public_sector_effects.loc[
    public_sector_effects["결과변수 구분"] == "메인 결과변수"
].copy()

adjusted_probabilities = public_sector_effects[
    [
        "결과변수",
        "결과변수 라벨",
        "결과변수 구분",
        "모형",
        "표본 N",
        "비공공부문 조정예측확률(%)",
        "공공부문 조정예측확률(%)",
        "조정확률 차이(%p, 공공-비공공)",
        "오즈비(OR)",
        "OR 95% CI 하한",
        "OR 95% CI 상한",
        "p-value 표기",
        "유의성",
    ]
].copy()

model_samples = pd.DataFrame(model_sample_rows)

all_coefficients = pd.concat(
    coefficient_tables,
    ignore_index=True,
)


# ------------------------------------------------------------
# 9. 파일 저장
# ------------------------------------------------------------

tables = {
    "01_공공부문효과_전체": public_sector_effects,
    "02_공공부문효과_메인": main_public_sector_effects,
    "03_조정예측확률": adjusted_probabilities,
    "04_모형별표본": model_samples,
    "05_전체회귀계수": all_coefficients,
}

with pd.ExcelWriter(
    OUTPUT_XLSX,
    engine="openpyxl",
) as writer:
    for sheet_name, table in tables.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

for sheet_name, table in tables.items():
    safe_name = (
        sheet_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )

    save_csv(
        dataframe=table,
        file_name=f"{safe_name}.csv",
    )

format_excel(OUTPUT_XLSX)


# ------------------------------------------------------------
# 10. 콘솔 요약
# ------------------------------------------------------------

console_columns = [
    "결과변수 라벨",
    "모형",
    "표본 N",
    "오즈비(OR)",
    "OR 95% CI 하한",
    "OR 95% CI 상한",
    "p-value 표기",
    "유의성",
    "비공공부문 조정예측확률(%)",
    "공공부문 조정예측확률(%)",
    "조정확률 차이(%p, 공공-비공공)",
]

print("\n[공공부문 효과: 메인 결과변수]")
print(
    main_public_sector_effects[
        console_columns
    ].to_string(index=False)
)

print("\n" + "=" * 72)
print("국가 고정효과 로지스틱 회귀분석 완료")
print(f"엑셀 파일: {OUTPUT_XLSX}")
print(f"결과 폴더: {OUTPUT_DIR}")
print("=" * 72)
