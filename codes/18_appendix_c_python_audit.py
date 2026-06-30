from __future__ import annotations

from pathlib import Path
import importlib.util
import math
import sys
import warnings

import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import t as student_t


# ============================================================
# 18_appendix_c_python_audit.py
#
# 목적
# - Python-only 환경에서 Appendix C의 WCLR 결과를 재점검
# - 17_wild_cluster_bootstrap_logit.py의 WCLR 결과에 대해
#   (1) restricted logit 수렴 여부,
#   (2) Monte-Carlo seed 안정성,
#   (3) leave-one-country-out cluster jackknife
#   를 별도로 확인
#
# 핵심 원칙
# - WCLR-C/WCLR-S는 17번 스크립트와 동일한 구현을 사용하므로
#   "외부 소프트웨어 검증"이 아니라 재현성·수렴성·난수오차 점검이다.
# - cluster jackknife는 bootstrap 코드와 독립적으로 국가 하나씩을 제외해
#   조정 로짓모형을 다시 적합하는 별도 소표본 클러스터 추론 점검이다.
# - restricted logit이 수렴하지 않는 결과변수는 WCLR p-value를
#   Appendix C 후보표에서 자동으로 비워 둔다.
#
# 실행:
#   python .\18_appendix_c_python_audit.py
#
# 전제:
# - codes/17_wild_cluster_bootstrap_logit.py가 같은 폴더에 있어야 함
# - clean_data/clean_data.parquet가 프로젝트 루트에 있어야 함
#
# 출력:
# results/18_appendix_c_python_audit/
#   ├─ 18_appendix_c_python_audit.xlsx
#   ├─ 01_AppendixC_Candidate.csv
#   ├─ 02_WCLR_MonteCarlo_Audit.csv
#   ├─ 03_Cluster_Jackknife_Summary.csv
#   ├─ 04_LeaveOneCountry_Estimates.csv
#   └─ 05_Notes.csv
# ============================================================


# ------------------------------------------------------------
# 0. Paths
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

WCLR_SCRIPT_PATH = (
    SCRIPT_DIR / "17_wild_cluster_bootstrap_logit.py"
)

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"

OUTPUT_DIR = (
    PROJECT_ROOT / "results" / "18_appendix_c_python_audit"
)
OUTPUT_XLSX = OUTPUT_DIR / "18_appendix_c_python_audit.xlsx"


# ------------------------------------------------------------
# 1. Audit settings
#
# 초기 실행에는 아래 값이 현실적인 균형입니다.
# 최종 원고 직전에는 FINAL_WCLR_REPS를 99_999로 올려도 됩니다.
# ------------------------------------------------------------

FINAL_WCLR_REPS = 29_999
AUDIT_WCLR_REPS = 9_999
AUDIT_SEEDS = (
    20260630,  # 기존 Appendix C 산출과 같은 seed
    20260701,
    20260702,
)

# RQ3은 기존 실행에서 restricted logit이 미수렴했으므로,
# WCLR p-value를 재계산하지 않고 수렴 실패 여부만 확인합니다.
WCLR_CANDIDATE_OUTCOMES = (
    "detailed_explanation",
    "personal_data_access",
    "not_informed",
)


# ------------------------------------------------------------
# 2. General utilities
# ------------------------------------------------------------

def load_wclr_module():
    """
    17번 WCLR 스크립트를 모듈로 불러옵니다.
    main()은 __name__ == '__main__'일 때만 실행되므로,
    import 시 분석이 자동으로 재실행되지는 않습니다.
    """
    if not WCLR_SCRIPT_PATH.exists():
        raise FileNotFoundError(
            "17_wild_cluster_bootstrap_logit.py를 찾을 수 없습니다.\n"
            f"확인 경로: {WCLR_SCRIPT_PATH}\n"
            "18번 파일과 17번 파일을 같은 codes 폴더에 두세요."
        )

    module_name = "wclr_reference_module"

    spec = importlib.util.spec_from_file_location(
        module_name,
        WCLR_SCRIPT_PATH,
    )

    if spec is None or spec.loader is None:
        raise ImportError(
            "17번 스크립트의 import specification을 만들지 못했습니다."
        )

    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)

    return module


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""

    if p_value < 0.001:
        return "< .001"

    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""

    if p_value < 0.001:
        return "***"

    if p_value < 0.01:
        return "**"

    if p_value < 0.05:
        return "*"

    return ""


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 58))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

        for row_number in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 24

    workbook.save(excel_path)


def safe_converged(result) -> bool:
    """
    statsmodels optimizer 결과의 convergence flag를 안전하게 읽습니다.
    """
    return bool(
        getattr(result, "mle_retvals", {}).get(
            "converged",
            False,
        )
    )


# ------------------------------------------------------------
# 3. Conventional adjusted model
# ------------------------------------------------------------

def fit_conventional_adjusted_model(
    wclr_module,
    outcome: str,
    model_df: pd.DataFrame,
):
    """
    Table 2와 동일한 조정 로짓모형을 국가 단위 cluster-robust SE로 적합.
    """
    formula = wclr_module.adjusted_formula(outcome)

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.logit(
            formula=formula,
            data=model_df,
            missing="raise",
        )

        result = model.fit(
            disp=False,
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(
                    dtype=object
                ),
            },
        )

    warnings_text = " | ".join(
        str(item.message)
        for item in warning_list
    )

    parameter = "public_sector"
    ci = result.conf_int().loc[parameter]

    return {
        "Result": result,
        "Outcome": outcome,
        "N": int(result.nobs),
        "Countries": int(model_df["country_fe"].nunique()),
        "Beta": float(result.params[parameter]),
        "OR": float(math.exp(result.params[parameter])),
        "CI lower": float(math.exp(ci.iloc[0])),
        "CI upper": float(math.exp(ci.iloc[1])),
        "Conventional cluster p": float(result.pvalues[parameter]),
        "Full model converged": safe_converged(result),
        "Warnings": warnings_text,
    }


# ------------------------------------------------------------
# 4. Restricted-model convergence check
# ------------------------------------------------------------

def check_restricted_logit_convergence(
    wclr_module,
    outcome: str,
    model_df: pd.DataFrame,
) -> dict:
    """
    H0: public_sector coefficient = 0의 restricted logit을 직접 적합.

    WCLR은 restricted model을 기반으로 하므로,
    이 모형이 수렴하지 않으면 해당 outcome의 WCLR p-value는
    보고하지 않는 것이 원칙입니다.
    """
    y, x, clusters, design_columns = (
        wclr_module.build_reordered_design(
            outcome=outcome,
            model_df=model_df,
        )
    )

    # x[:, 0] = public_sector; restricted H0에서는 제거
    x_restricted = x[:, 1:]

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        restricted_model = sm.Logit(y, x_restricted)

        restricted_result = restricted_model.fit(
            disp=False,
            maxiter=500,
        )

    warnings_text = " | ".join(
        str(item.message)
        for item in warning_list
    )

    return {
        "Outcome": outcome,
        "Restricted model converged": safe_converged(
            restricted_result
        ),
        "Restricted warnings": warnings_text,
        "Design parameters": int(x.shape[1]),
        "Design columns": " | ".join(design_columns),
    }


# ------------------------------------------------------------
# 5. Cluster jackknife
# ------------------------------------------------------------

def cluster_jackknife(
    wclr_module,
    outcome: str,
    model_df: pd.DataFrame,
) -> tuple[dict, pd.DataFrame]:
    """
    국가 하나씩을 제외한 27개 재적합으로 cluster jackknife SE를 계산.

    Formula:
        SE_JK = sqrt((G - 1) / G * sum_g(beta[-g] - mean(beta[-g]))^2)

    Test statistic:
        t_JK = beta_full / SE_JK
        p_JK = two-sided t distribution with df = G - 1

    이는 WCLR과 다른 독립적인 소표본 클러스터 점검입니다.
    """
    formula = wclr_module.adjusted_formula(outcome)
    target = "public_sector"

    full_model = smf.logit(
        formula=formula,
        data=model_df,
        missing="raise",
    )

    full_result = full_model.fit(
        disp=False,
        maxiter=500,
    )

    beta_full = float(full_result.params[target])

    countries = sorted(
        model_df["country_fe"].astype(str).unique().tolist()
    )

    leave_one_rows = []

    for country in countries:
        subset = model_df.loc[
            model_df["country_fe"].astype(str) != country
        ].copy()

        warning_text = ""

        try:
            with warnings.catch_warnings(record=True) as warning_list:
                warnings.simplefilter("always")

                result_minus_country = smf.logit(
                    formula=formula,
                    data=subset,
                    missing="raise",
                ).fit(
                    disp=False,
                    maxiter=500,
                )

            warning_text = " | ".join(
                str(item.message)
                for item in warning_list
            )

            converged = safe_converged(result_minus_country)

            beta_minus_country = float(
                result_minus_country.params[target]
            )

            leave_one_rows.append(
                {
                    "Outcome": outcome,
                    "Excluded country": country,
                    "N (-country)": int(result_minus_country.nobs),
                    "Converged": converged,
                    "Public-sector beta": beta_minus_country,
                    "OR": float(math.exp(beta_minus_country)),
                    "Warning": warning_text,
                    "Error": "",
                }
            )

        except Exception as error:
            leave_one_rows.append(
                {
                    "Outcome": outcome,
                    "Excluded country": country,
                    "N (-country)": int(len(subset)),
                    "Converged": False,
                    "Public-sector beta": np.nan,
                    "OR": np.nan,
                    "Warning": warning_text,
                    "Error": repr(error),
                }
            )

    leave_one_df = pd.DataFrame(leave_one_rows)

    successful = leave_one_df.loc[
        leave_one_df["Converged"]
        & leave_one_df["Public-sector beta"].notna()
    ].copy()

    g_total = len(countries)
    g_success = len(successful)

    if g_success != g_total:
        summary = {
            "Outcome": outcome,
            "Full beta": beta_full,
            "Jackknife SE": np.nan,
            "Jackknife t": np.nan,
            "Jackknife p (df=G-1)": np.nan,
            "Jackknife successful clusters": g_success,
            "Jackknife total clusters": g_total,
            "Jackknife status": (
                "NOT REPORTED: at least one leave-one-country model "
                "failed to converge."
            ),
        }

        return summary, leave_one_df

    beta_minus = successful[
        "Public-sector beta"
    ].to_numpy(dtype=float)

    beta_minus_mean = float(beta_minus.mean())

    jackknife_variance = (
        (g_total - 1) / g_total
        * np.sum((beta_minus - beta_minus_mean) ** 2)
    )

    jackknife_se = float(math.sqrt(jackknife_variance))
    jackknife_t = float(beta_full / jackknife_se)

    jackknife_p = float(
        2.0
        * student_t.sf(
            abs(jackknife_t),
            df=g_total - 1,
        )
    )

    summary = {
        "Outcome": outcome,
        "Full beta": beta_full,
        "Jackknife SE": jackknife_se,
        "Jackknife t": jackknife_t,
        "Jackknife p (df=G-1)": jackknife_p,
        "Jackknife successful clusters": g_success,
        "Jackknife total clusters": g_total,
        "Jackknife status": "REPORTED",
    }

    return summary, leave_one_df


# ------------------------------------------------------------
# 6. WCLR Monte-Carlo audit
# ------------------------------------------------------------

def run_wclr_monte_carlo_audit(
    wclr_module,
    outcome: str,
    model_df: pd.DataFrame,
) -> tuple[dict, pd.DataFrame]:
    """
    동일 WCLR 구현을 서로 다른 random seed에서 반복 실행해
    Monte-Carlo 오차가 결론을 바꾸지 않는지 점검합니다.

    주의:
    - 이는 '독립 패키지 검증'이 아니라 randomization stability audit.
    - restricted logit이 미수렴하면 해당 outcome은 자동으로 무효 처리.
    """
    convergence = check_restricted_logit_convergence(
        wclr_module=wclr_module,
        outcome=outcome,
        model_df=model_df,
    )

    if not convergence["Restricted model converged"]:
        summary = {
            "Outcome": outcome,
            "Restricted model converged": False,
            "Audit runs requested": len(AUDIT_SEEDS),
            "Audit runs valid": 0,
            "WCLR-C mean": np.nan,
            "WCLR-C SD": np.nan,
            "WCLR-C min": np.nan,
            "WCLR-C max": np.nan,
            "WCLR-S mean": np.nan,
            "WCLR-S SD": np.nan,
            "WCLR-S min": np.nan,
            "WCLR-S max": np.nan,
            "Audit status": (
                "NOT REPORTED: restricted logit did not converge."
            ),
        }

        empty_df = pd.DataFrame(
            [
                {
                    "Outcome": outcome,
                    "Seed": np.nan,
                    "Bootstrap replications": AUDIT_WCLR_REPS,
                    "Restricted model converged": False,
                    "WCLR-C p": np.nan,
                    "WCLR-S p": np.nan,
                    "Status": (
                        "Skipped: restricted logit did not converge."
                    ),
                }
            ]
        )

        return summary, empty_df

    y, x, clusters, _ = wclr_module.build_reordered_design(
        outcome=outcome,
        model_df=model_df,
    )

    draw_rows = []

    for seed in AUDIT_SEEDS:
        try:
            result = wclr_module.restricted_wclr_p_values(
                y=y,
                x=x,
                clusters=clusters,
                reps=AUDIT_WCLR_REPS,
                seed=seed,
            )

            is_valid = bool(
                result["Restricted logit converged"]
            )

            draw_rows.append(
                {
                    "Outcome": outcome,
                    "Seed": seed,
                    "Bootstrap replications": AUDIT_WCLR_REPS,
                    "Restricted model converged": is_valid,
                    "WCLR-C p": (
                        result["WCLR-C symmetric p"]
                        if is_valid
                        else np.nan
                    ),
                    "WCLR-S p": (
                        result["WCLR-S symmetric p"]
                        if is_valid
                        else np.nan
                    ),
                    "Status": (
                        "Valid"
                        if is_valid
                        else (
                            "Invalid: restricted logit did not "
                            "converge."
                        )
                    ),
                }
            )

        except Exception as error:
            draw_rows.append(
                {
                    "Outcome": outcome,
                    "Seed": seed,
                    "Bootstrap replications": AUDIT_WCLR_REPS,
                    "Restricted model converged": False,
                    "WCLR-C p": np.nan,
                    "WCLR-S p": np.nan,
                    "Status": f"Failed: {error!r}",
                }
            )

    draw_df = pd.DataFrame(draw_rows)

    valid_draws = draw_df.loc[
        draw_df["Restricted model converged"]
        & draw_df["WCLR-C p"].notna()
        & draw_df["WCLR-S p"].notna()
    ].copy()

    valid_n = len(valid_draws)

    if valid_n == 0:
        summary = {
            "Outcome": outcome,
            "Restricted model converged": True,
            "Audit runs requested": len(AUDIT_SEEDS),
            "Audit runs valid": 0,
            "WCLR-C mean": np.nan,
            "WCLR-C SD": np.nan,
            "WCLR-C min": np.nan,
            "WCLR-C max": np.nan,
            "WCLR-S mean": np.nan,
            "WCLR-S SD": np.nan,
            "WCLR-S min": np.nan,
            "WCLR-S max": np.nan,
            "Audit status": "NOT REPORTED: no successful audit draws.",
        }

        return summary, draw_df

    summary = {
        "Outcome": outcome,
        "Restricted model converged": True,
        "Audit runs requested": len(AUDIT_SEEDS),
        "Audit runs valid": valid_n,
        "WCLR-C mean": float(valid_draws["WCLR-C p"].mean()),
        "WCLR-C SD": float(valid_draws["WCLR-C p"].std(ddof=1))
        if valid_n > 1
        else 0.0,
        "WCLR-C min": float(valid_draws["WCLR-C p"].min()),
        "WCLR-C max": float(valid_draws["WCLR-C p"].max()),
        "WCLR-S mean": float(valid_draws["WCLR-S p"].mean()),
        "WCLR-S SD": float(valid_draws["WCLR-S p"].std(ddof=1))
        if valid_n > 1
        else 0.0,
        "WCLR-S min": float(valid_draws["WCLR-S p"].min()),
        "WCLR-S max": float(valid_draws["WCLR-S p"].max()),
        "Audit status": "REPORTED",
    }

    return summary, draw_df


def run_final_wclr(
    wclr_module,
    outcome: str,
    model_df: pd.DataFrame,
) -> dict:
    """
    Appendix C 후보표에 넣을 high-replication WCLR 결과를 산출.
    restricted logit 미수렴 outcome에는 결측을 반환합니다.
    """
    convergence = check_restricted_logit_convergence(
        wclr_module=wclr_module,
        outcome=outcome,
        model_df=model_df,
    )

    if not convergence["Restricted model converged"]:
        return {
            "Outcome": outcome,
            "Final WCLR reps": FINAL_WCLR_REPS,
            "Final WCLR-C p": np.nan,
            "Final WCLR-S p": np.nan,
            "Final WCLR status": (
                "NOT REPORTED: restricted logit did not converge."
            ),
        }

    y, x, clusters, _ = wclr_module.build_reordered_design(
        outcome=outcome,
        model_df=model_df,
    )

    result = wclr_module.restricted_wclr_p_values(
        y=y,
        x=x,
        clusters=clusters,
        reps=FINAL_WCLR_REPS,
        seed=20260630,
    )

    if not result["Restricted logit converged"]:
        return {
            "Outcome": outcome,
            "Final WCLR reps": FINAL_WCLR_REPS,
            "Final WCLR-C p": np.nan,
            "Final WCLR-S p": np.nan,
            "Final WCLR status": (
                "NOT REPORTED: restricted logit did not converge."
            ),
        }

    return {
        "Outcome": outcome,
        "Final WCLR reps": FINAL_WCLR_REPS,
        "Final WCLR-C p": float(result["WCLR-C symmetric p"]),
        "Final WCLR-S p": float(result["WCLR-S symmetric p"]),
        "Final WCLR status": "REPORTED",
    }


# ------------------------------------------------------------
# 7. Main
# ------------------------------------------------------------

def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            "전처리 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {INPUT_PATH}"
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wclr_module = load_wclr_module()

    raw_df = pd.read_parquet(INPUT_PATH)
    model_df = wclr_module.prepare_common_sample(raw_df)

    if len(model_df) != 9_719:
        raise ValueError(
            "공통 완전사례 표본 N이 논문과 다릅니다.\n"
            f"Expected N=9,719 | Actual N={len(model_df):,}"
        )

    if model_df["country_fe"].nunique() != 27:
        raise ValueError(
            "국가 클러스터 수가 논문과 다릅니다.\n"
            "Expected G=27 | "
            f"Actual G={model_df['country_fe'].nunique()}"
        )

    print("=" * 78)
    print("Appendix C Python-only audit 시작")
    print(f"공통 완전사례 표본 N: {len(model_df):,}")
    print(
        "국가 클러스터 수 G: "
        f"{model_df['country_fe'].nunique()}"
    )
    print(
        "Final WCLR reps: "
        f"{FINAL_WCLR_REPS:,} | "
        "Audit reps per seed: "
        f"{AUDIT_WCLR_REPS:,}"
    )
    print("=" * 78)

    conventional_rows = []
    restricted_rows = []
    jackknife_rows = []
    leave_one_tables = []
    monte_carlo_rows = []
    monte_carlo_draw_tables = []
    final_wclr_rows = []

    for outcome, outcome_label in wclr_module.OUTCOMES.items():
        print(f"\n분석 중: {outcome}")

        conventional = fit_conventional_adjusted_model(
            wclr_module=wclr_module,
            outcome=outcome,
            model_df=model_df,
        )

        conventional_rows.append(
            {
                "Outcome": outcome,
                "Outcome label": outcome_label,
                **{
                    key: value
                    for key, value in conventional.items()
                    if key != "Result"
                },
            }
        )

        restricted = check_restricted_logit_convergence(
            wclr_module=wclr_module,
            outcome=outcome,
            model_df=model_df,
        )

        restricted_rows.append(restricted)

        jackknife_summary, leave_one_df = cluster_jackknife(
            wclr_module=wclr_module,
            outcome=outcome,
            model_df=model_df,
        )

        jackknife_rows.append(jackknife_summary)
        leave_one_tables.append(leave_one_df)

        monte_carlo_summary, monte_carlo_draws = (
            run_wclr_monte_carlo_audit(
                wclr_module=wclr_module,
                outcome=outcome,
                model_df=model_df,
            )
        )

        monte_carlo_rows.append(monte_carlo_summary)
        monte_carlo_draw_tables.append(monte_carlo_draws)

        if outcome in WCLR_CANDIDATE_OUTCOMES:
            final_wclr_rows.append(
                run_final_wclr(
                    wclr_module=wclr_module,
                    outcome=outcome,
                    model_df=model_df,
                )
            )
        else:
            final_wclr_rows.append(
                {
                    "Outcome": outcome,
                    "Final WCLR reps": FINAL_WCLR_REPS,
                    "Final WCLR-C p": np.nan,
                    "Final WCLR-S p": np.nan,
                    "Final WCLR status": (
                        "NOT RUN: outcome excluded because restricted "
                        "model was previously non-convergent."
                    ),
                }
            )

    conventional_df = pd.DataFrame(conventional_rows)
    restricted_df = pd.DataFrame(restricted_rows)
    jackknife_df = pd.DataFrame(jackknife_rows)
    leave_one_df = pd.concat(
        leave_one_tables,
        ignore_index=True,
    )
    monte_carlo_df = pd.DataFrame(monte_carlo_rows)
    monte_carlo_draws_df = pd.concat(
        monte_carlo_draw_tables,
        ignore_index=True,
    )
    final_wclr_df = pd.DataFrame(final_wclr_rows)

    appendix_candidate = (
        conventional_df
        .merge(
            restricted_df[
                [
                    "Outcome",
                    "Restricted model converged",
                    "Restricted warnings",
                ]
            ],
            on="Outcome",
            how="left",
        )
        .merge(
            final_wclr_df,
            on="Outcome",
            how="left",
        )
        .merge(
            jackknife_df[
                [
                    "Outcome",
                    "Jackknife SE",
                    "Jackknife t",
                    "Jackknife p (df=G-1)",
                    "Jackknife status",
                ]
            ],
            on="Outcome",
            how="left",
        )
        .merge(
            monte_carlo_df[
                [
                    "Outcome",
                    "WCLR-C mean",
                    "WCLR-C min",
                    "WCLR-C max",
                    "WCLR-S mean",
                    "WCLR-S min",
                    "WCLR-S max",
                    "Audit status",
                ]
            ],
            on="Outcome",
            how="left",
        )
    )

    appendix_candidate["Conventional p label"] = (
        appendix_candidate["Conventional cluster p"].map(
            p_value_label
        )
    )

    appendix_candidate["Final WCLR-C p label"] = (
        appendix_candidate["Final WCLR-C p"].map(
            p_value_label
        )
    )

    appendix_candidate["Final WCLR-S p label"] = (
        appendix_candidate["Final WCLR-S p"].map(
            p_value_label
        )
    )

    appendix_candidate["Jackknife p label"] = (
        appendix_candidate["Jackknife p (df=G-1)"].map(
            p_value_label
        )
    )

    appendix_candidate["Publication rule"] = np.where(
        appendix_candidate["Restricted model converged"]
        & appendix_candidate["Final WCLR-S p"].notna(),
        (
            "WCLR reportable; report WCLR-S as primary "
            "and WCLR-C for comparison."
        ),
        (
            "Do not report WCLR p-values; use conventional "
            "clustered results and jackknife check only."
        ),
    )

    notes_df = pd.DataFrame(
        [
            {
                "Item": "Interpretation of this audit",
                "Detail": (
                    "This is a Python-only reproducibility and "
                    "small-cluster audit. It does not claim that a second "
                    "external software implementation reproduced WCLR."
                ),
            },
            {
                "Item": "WCLR reporting rule",
                "Detail": (
                    "Report WCLR-C/WCLR-S only when the restricted logit "
                    "model converges. A non-convergent restricted model "
                    "invalidates restricted-bootstrap p-values."
                ),
            },
            {
                "Item": "Independent check",
                "Detail": (
                    "The leave-one-country cluster jackknife refits "
                    "27 adjusted logit models and is independent of the "
                    "linearized WCLR calculation."
                ),
            },
            {
                "Item": "Primary WCLR statistic",
                "Detail": (
                    "Use WCLR-S as the primary bootstrap sensitivity "
                    "statistic and WCLR-C as a transparent companion "
                    "statistic only for convergent restricted models."
                ),
            },
            {
                "Item": "RQ3 handling",
                "Detail": (
                    "If the restricted RQ3 logit remains non-convergent, "
                    "enter em dashes for WCLR-C/WCLR-S in Appendix C and "
                    "state the reason in a footnote."
                ),
            },
            {
                "Item": "Final-replication setting",
                "Detail": (
                    f"Final WCLR results use {FINAL_WCLR_REPS:,} "
                    "Rademacher draws with seed 20260630. "
                    "Monte-Carlo audit uses "
                    f"{len(AUDIT_SEEDS)} seeds x "
                    f"{AUDIT_WCLR_REPS:,} draws."
                ),
            },
        ]
    )

    tables = {
        "01_AppendixC_Candidate": appendix_candidate,
        "02_WCLR_MonteCarlo": monte_carlo_df,
        "03_WCLR_AuditDraws": monte_carlo_draws_df,
        "04_Jackknife_Summary": jackknife_df,
        "05_LeaveOneCountry": leave_one_df,
        "06_Restricted_Check": restricted_df,
        "07_Notes": notes_df,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    for sheet_name, dataframe in tables.items():
        dataframe.to_csv(
            OUTPUT_DIR / f"{sheet_name}.csv",
            index=False,
            encoding="utf-8-sig",
        )

    format_excel(OUTPUT_XLSX)

    print("\n[Appendix C candidate]")
    display_columns = [
        "Outcome label",
        "Conventional p label",
        "Restricted model converged",
        "Final WCLR-C p label",
        "Final WCLR-S p label",
        "Jackknife p label",
        "Publication rule",
    ]

    print(
        appendix_candidate[
            display_columns
        ].to_string(index=False)
    )

    print("\n" + "=" * 78)
    print("분석 완료")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"Results folder: {OUTPUT_DIR}")
    print("=" * 78)


if __name__ == "__main__":
    main()
