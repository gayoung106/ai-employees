from __future__ import annotations

from pathlib import Path
import math
import re
import warnings

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import patsy
import pyreadstat
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.special import expit
from scipy.stats import chi2_contingency, fisher_exact, norm
from statsmodels.stats.multitest import multipletests


# ============================================================
# 09_occupation_mechanism_detailed_explanation.py
#
# 목적
# 1) 공공·비공공부문 간 직업 유형 구성 차이를 확인
# 2) 직업 유형별 상세 설명 제공률을 확인
# 3) 직업 유형별로 보아도 공공부문 계수가 대체로 음(-)의
#    방향을 유지하는지, 상호작용 로지스틱 회귀와 조정예측확률로 확인
#
# 주의
# - 이것은 '직업 유형이 공공부문 효과를 매개한다'는 분석이 아닙니다.
# - 직업 유형은 공공부문 고용과 구조적으로 연관된 구성요인일 수 있습니다.
# - 따라서 이 분석은 Model 1 -> Model 2 변화의 원인을 설명하기 위한
#   "직업구성 및 조건부 이질성 진단"입니다.
# ============================================================


# ------------------------------------------------------------
# 0. 경로
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
RAW_SAV_PATH = PROJECT_ROOT / "datas" / "raw_data.sav"

OUTPUT_DIR = (
    PROJECT_ROOT
    / "results"
    / "09_occupation_mechanism_detailed_explanation"
)
OUTPUT_XLSX = OUTPUT_DIR / "09_occupation_mechanism_detailed_explanation.xlsx"
OUTPUT_PNG = OUTPUT_DIR / "Figure_A1_OccupationSpecific_AAPD.png"


# ------------------------------------------------------------
# 1. 분석 정의
# ------------------------------------------------------------

OUTCOME = "detailed_explanation"
OUTCOME_LABEL = "상세 설명 제공"

REQUIRED_COLUMNS = [
    OUTCOME,
    "public_sector",
    "country_fe",
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

BASELINE_FORMULA = (
    f"{OUTCOME} ~ public_sector + C(country_fe)"
    " + age + C(gender) + education_age"
    " + C(occupation_code) + C(workplace_size)"
    " + digital_skill_job"
)

INTERACTION_FORMULA = (
    f"{OUTCOME} ~ public_sector * C(occupation_code)"
    " + C(country_fe) + age + C(gender)"
    " + education_age + C(workplace_size)"
    " + digital_skill_job"
)


# ------------------------------------------------------------
# 2. 공통 함수
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 55))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

        for row_number in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 24

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. 자료 준비
# ------------------------------------------------------------

def prepare_complete_case_sample(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    기존 Model 2와 동일한 완전사례 표본(N=9,719)을 재현합니다.
    """
    model_df = dataframe[REQUIRED_COLUMNS].copy()

    for column in [
        OUTCOME,
        "public_sector",
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    model_df = model_df[
        model_df[OUTCOME].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
    ].copy()

    model_df = model_df.dropna().copy()

    model_df[OUTCOME] = to_int64(model_df[OUTCOME])
    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in ["age", "education_age", "digital_skill_job"]:
        model_df[column] = to_float64(model_df[column])

    for column in ["gender", "occupation_code", "workplace_size"]:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


def load_occupation_labels() -> dict[int, str]:
    """
    원자료 SAV의 d15a value label을 사용해 직업 유형 명칭을 가져옵니다.
    원자료 파일 또는 metadata 접근에 실패하면 코드명으로 대체합니다.
    """
    fallback = {
        code: f"Occupation code {code}"
        for code in range(10, 19)
    }

    if not RAW_SAV_PATH.exists():
        return fallback

    try:
        _, metadata = pyreadstat.read_sav(
            RAW_SAV_PATH,
            metadataonly=True,
        )

        value_labels = metadata.variable_value_labels.get("d15a", {})

        if not value_labels:
            return fallback

        labels = {}

        for code in range(10, 19):
            label = (
                value_labels.get(float(code))
                or value_labels.get(code)
                or fallback[code]
            )

            labels[code] = str(label)

        return labels

    except Exception:
        return fallback


# ------------------------------------------------------------
# 4. 기술통계: 직업 구성과 원자료 비율
# ------------------------------------------------------------

def chi_square_or_fisher(table: np.ndarray) -> tuple[str, float, float]:
    """
    2x2 표에서 기대빈도 5 미만이 있으면 Fisher exact test,
    아니면 Pearson chi-square test를 사용합니다.
    """
    try:
        chi2, p_value, _, expected = chi2_contingency(
            table,
            correction=False,
        )

        if (expected < 5).any():
            _, fisher_p = fisher_exact(table)
            return "Fisher exact", np.nan, float(fisher_p)

        return "Pearson chi-square", float(chi2), float(p_value)

    except ValueError:
        return "Not estimable", np.nan, np.nan


def build_occupation_composition(
    model_df: pd.DataFrame,
    occupation_labels: dict[int, str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    직업 유형의 부문별 분포와 전체 chi-square 검정 결과를 생성합니다.
    """
    occupation_codes = sorted(model_df["occupation_code"].unique())

    nonpublic_total = int((model_df["public_sector"] == 0).sum())
    public_total = int((model_df["public_sector"] == 1).sum())

    rows = []

    for code in occupation_codes:
        subset = model_df.loc[
            model_df["occupation_code"] == code
        ]

        nonpublic_n = int((subset["public_sector"] == 0).sum())
        public_n = int((subset["public_sector"] == 1).sum())

        rows.append(
            {
                "Occupation code": int(code),
                "Occupation label": occupation_labels[int(code)],
                "Non-public N": nonpublic_n,
                "Non-public sector composition (%)": round(
                    nonpublic_n / nonpublic_total * 100,
                    2,
                ),
                "Public N": public_n,
                "Public sector composition (%)": round(
                    public_n / public_total * 100,
                    2,
                ),
                "Public - non-public composition difference (pp)": round(
                    public_n / public_total * 100
                    - nonpublic_n / nonpublic_total * 100,
                    2,
                ),
                "Total N": int(len(subset)),
            }
        )

    composition_table = pd.DataFrame(rows)

    crosstab = pd.crosstab(
        model_df["public_sector"],
        model_df["occupation_code"],
    )

    chi2, p_value, df, _ = chi2_contingency(
        crosstab,
        correction=False,
    )

    n_total = crosstab.to_numpy().sum()
    min_dimension = min(crosstab.shape) - 1
    cramers_v = math.sqrt(
        chi2 / (n_total * min_dimension)
    )

    overall_test = pd.DataFrame(
        [
            {
                "Comparison": "Public vs. non-public occupation distribution",
                "N": int(n_total),
                "Chi-square": round(float(chi2), 3),
                "df": int(df),
                "p-value": float(p_value),
                "p-value label": p_value_label(float(p_value)),
                "Significance": significance_mark(float(p_value)),
                "Cramer's V": round(float(cramers_v), 4),
            }
        ]
    )

    return composition_table, overall_test


def build_raw_detail_rates(
    model_df: pd.DataFrame,
    occupation_labels: dict[int, str],
) -> pd.DataFrame:
    """
    직업 유형별 상세 설명 제공률의 비조정 비교표를 생성합니다.
    """
    rows = []

    for code in sorted(model_df["occupation_code"].unique()):
        subset = model_df.loc[
            model_df["occupation_code"] == code
        ]

        nonpublic = subset.loc[subset["public_sector"] == 0]
        public = subset.loc[subset["public_sector"] == 1]

        nonpublic_yes = int(nonpublic[OUTCOME].sum())
        public_yes = int(public[OUTCOME].sum())

        nonpublic_n = len(nonpublic)
        public_n = len(public)

        table = np.array(
            [
                [public_yes, public_n - public_yes],
                [nonpublic_yes, nonpublic_n - nonpublic_yes],
            ]
        )

        test_name, chi2, p_value = chi_square_or_fisher(table)

        rows.append(
            {
                "Occupation code": int(code),
                "Occupation label": occupation_labels[int(code)],
                "Non-public N": int(nonpublic_n),
                "Non-public detailed explanation N": int(nonpublic_yes),
                "Non-public detailed explanation (%)": round(
                    nonpublic_yes / nonpublic_n * 100,
                    2,
                ),
                "Public N": int(public_n),
                "Public detailed explanation N": int(public_yes),
                "Public detailed explanation (%)": round(
                    public_yes / public_n * 100,
                    2,
                ),
                "Public - non-public difference (pp)": round(
                    public_yes / public_n * 100
                    - nonpublic_yes / nonpublic_n * 100,
                    2,
                ),
                "Test": test_name,
                "Chi-square": (
                    np.nan
                    if pd.isna(chi2)
                    else round(chi2, 3)
                ),
                "p-value": p_value,
                "p-value label": p_value_label(p_value),
                "Significance": significance_mark(p_value),
            }
        )

    raw_rates = pd.DataFrame(rows)

    valid_p = raw_rates["p-value"].notna()

    raw_rates["FDR-adjusted p-value"] = np.nan
    raw_rates["FDR-adjusted p-value label"] = ""
    raw_rates["FDR significance"] = ""

    if valid_p.any():
        adjusted = multipletests(
            raw_rates.loc[valid_p, "p-value"],
            alpha=0.05,
            method="fdr_bh",
        )

        raw_rates.loc[
            valid_p,
            "FDR-adjusted p-value",
        ] = adjusted[1]

        raw_rates.loc[
            valid_p,
            "FDR-adjusted p-value label",
        ] = [
            p_value_label(p_value)
            for p_value in adjusted[1]
        ]

        raw_rates.loc[
            valid_p,
            "FDR significance",
        ] = [
            significance_mark(p_value)
            for p_value in adjusted[1]
        ]

    return raw_rates


def build_country_coverage(
    model_df: pd.DataFrame,
    occupation_labels: dict[int, str],
) -> pd.DataFrame:
    """
    직업별 조정예측확률이 얼마나 많은 국가에서
    공공·비공공 양 부문을 모두 관측하는지 점검합니다.
    """
    rows = []

    for code in sorted(model_df["occupation_code"].unique()):
        subset = model_df.loc[
            model_df["occupation_code"] == code
        ]

        country_sector = pd.crosstab(
            subset["country_fe"],
            subset["public_sector"],
        )

        countries_total = int(country_sector.shape[0])
        countries_both = int(
            (
                (country_sector.get(0, 0) > 0)
                & (country_sector.get(1, 0) > 0)
            ).sum()
        )

        rows.append(
            {
                "Occupation code": int(code),
                "Occupation label": occupation_labels[int(code)],
                "Countries represented": countries_total,
                "Countries with both sectors": countries_both,
                "Both-sector country coverage (%)": round(
                    countries_both / countries_total * 100,
                    2,
                ) if countries_total > 0 else np.nan,
            }
        )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 5. 상호작용 로지스틱 회귀
# ------------------------------------------------------------

def fit_glm(
    formula: str,
    model_df: pd.DataFrame,
):
    """
    국가 단위 cluster-robust SE를 적용한 이항 GLM.
    로지스틱 MLE와 동일한 링크 함수를 사용합니다.
    """
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.glm(
            formula=formula,
            data=model_df,
            family=sm.families.Binomial(),
            missing="raise",
        )

        try:
            result = model.fit(
                maxiter=500,
                cov_type="cluster",
                cov_kwds={
                    "groups": model_df["country_fe"].to_numpy(dtype=object),
                },
            )
            covariance_type = "Cluster-robust SE by country"

        except Exception as cluster_error:
            result = model.fit(
                maxiter=500,
                cov_type="HC3",
            )
            covariance_type = (
                "HC3 robust SE "
                f"(cluster estimation failed: {type(cluster_error).__name__})"
            )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    return (
        result,
        covariance_type,
        " | ".join(warning_messages),
    )


def extract_public_sector_effect(
    result,
    specification: str,
    model_df: pd.DataFrame,
    covariance_type: str,
    warning_text: str,
) -> dict:
    """
    기본 범주 직업유형에서의 공공부문 계수 또는
    비상호작용 모형의 공공부문 계수를 정리합니다.
    """
    parameter = "public_sector"

    beta = float(result.params[parameter])
    se = float(result.bse[parameter])
    p_value = float(result.pvalues[parameter])

    ci = result.conf_int().loc[parameter]
    ci_low = float(ci.iloc[0])
    ci_high = float(ci.iloc[1])

    return {
        "Specification": specification,
        "N": int(result.nobs),
        "Countries": int(model_df["country_fe"].nunique()),
        "Covariance estimator": covariance_type,
        "Warning": warning_text,
        "Public-sector beta": round(beta, 4),
        "SE": round(se, 4),
        "OR": round(math.exp(beta), 3),
        "OR 95% CI lower": round(math.exp(ci_low), 3),
        "OR 95% CI upper": round(math.exp(ci_high), 3),
        "p-value": p_value,
        "p-value label": p_value_label(p_value),
        "Significance": significance_mark(p_value),
    }


def interaction_wald_test(result) -> pd.DataFrame:
    """
    public_sector × occupation_code 상호작용 항들이 모두 0인지
    국가 군집강건 공분산행렬로 Wald 검정합니다.
    """
    parameter_names = list(result.params.index)

    interaction_indices = [
        index
        for index, name in enumerate(parameter_names)
        if (
            "public_sector" in name
            and "occupation_code" in name
        )
    ]

    if not interaction_indices:
        return pd.DataFrame(
            [
                {
                    "Test": "Public sector × occupation interaction",
                    "Status": "Not found",
                    "Number of interaction parameters": 0,
                    "Wald chi-square": np.nan,
                    "df": np.nan,
                    "p-value": np.nan,
                    "p-value label": "",
                    "Significance": "",
                }
            ]
        )

    r_matrix = np.zeros(
        (len(interaction_indices), len(parameter_names))
    )

    for row_index, parameter_index in enumerate(interaction_indices):
        r_matrix[row_index, parameter_index] = 1.0

    test_result = result.wald_test(r_matrix, scalar=True)

    statistic = float(np.asarray(test_result.statistic).squeeze())
    p_value = float(np.asarray(test_result.pvalue).squeeze())

    return pd.DataFrame(
        [
            {
                "Test": "Public sector × occupation interaction",
                "Status": "Estimated",
                "Number of interaction parameters": len(interaction_indices),
                "Wald chi-square": round(statistic, 3),
                "df": len(interaction_indices),
                "p-value": p_value,
                "p-value label": p_value_label(p_value),
                "Significance": significance_mark(p_value),
                "Interaction parameters": " | ".join(
                    parameter_names[index]
                    for index in interaction_indices
                ),
            }
        ]
    )


# ------------------------------------------------------------
# 6. 조정예측확률 및 직업별 공공부문 차이
# ------------------------------------------------------------

def get_design_matrix(
    result,
    new_data: pd.DataFrame,
) -> np.ndarray:
    """
    추정된 GLM과 동일한 Patsy design_info를 사용하여
    새 데이터의 설계행렬을 생성합니다.
    """
    design_info = result.model.data.design_info

    matrix = patsy.build_design_matrices(
        [design_info],
        new_data,
        return_type="dataframe",
    )[0]

    return matrix.to_numpy(dtype=float)


def standardized_margin(
    result,
    new_data: pd.DataFrame,
) -> tuple[float, float, float, np.ndarray]:
    """
    평균 조정예측확률과 델타법 95% CI.

    표준화 방식:
    - 직업유형 내부의 실제 국가·연령·성별·교육·조직규모·디지털역량 분포를 유지
    - public_sector만 0 또는 1로 반사실적으로 변경
    """
    x_matrix = get_design_matrix(result, new_data)

    beta = result.params.to_numpy(dtype=float)
    covariance = (
        result.cov_params()
        .loc[result.params.index, result.params.index]
        .to_numpy(dtype=float)
    )

    linear_predictor = x_matrix @ beta
    predicted = expit(linear_predictor)

    probability = float(predicted.mean())

    gradient = (
        x_matrix
        * (predicted * (1 - predicted))[:, None]
    ).mean(axis=0)

    variance = float(gradient @ covariance @ gradient.T)
    standard_error = math.sqrt(max(variance, 0.0))

    ci_low = max(0.0, probability - 1.96 * standard_error)
    ci_high = min(1.0, probability + 1.96 * standard_error)

    return probability, ci_low, ci_high, gradient


def build_adjusted_margins(
    interaction_result,
    model_df: pd.DataFrame,
    occupation_labels: dict[int, str],
) -> pd.DataFrame:
    """
    직업별로 공공/비공공부문 조정예측확률과 차이를 계산합니다.
    """
    rows = []

    for code in sorted(model_df["occupation_code"].unique()):
        occupation_df = model_df.loc[
            model_df["occupation_code"] == code
        ].copy()

        counterfactual_nonpublic = occupation_df.copy()
        counterfactual_public = occupation_df.copy()

        counterfactual_nonpublic["public_sector"] = 0
        counterfactual_public["public_sector"] = 1

        (
            probability_nonpublic,
            ci_low_nonpublic,
            ci_high_nonpublic,
            gradient_nonpublic,
        ) = standardized_margin(
            interaction_result,
            counterfactual_nonpublic,
        )

        (
            probability_public,
            ci_low_public,
            ci_high_public,
            gradient_public,
        ) = standardized_margin(
            interaction_result,
            counterfactual_public,
        )

        covariance = (
            interaction_result.cov_params()
            .loc[
                interaction_result.params.index,
                interaction_result.params.index,
            ]
            .to_numpy(dtype=float)
        )

        difference = probability_public - probability_nonpublic
        difference_gradient = gradient_public - gradient_nonpublic

        difference_variance = float(
            difference_gradient
            @ covariance
            @ difference_gradient.T
        )

        difference_se = math.sqrt(
            max(difference_variance, 0.0)
        )

        difference_ci_low = difference - 1.96 * difference_se
        difference_ci_high = difference + 1.96 * difference_se

        if difference_se == 0:
            z_value = np.nan
            p_value = np.nan
        else:
            z_value = difference / difference_se
            p_value = float(
                2 * norm.sf(abs(z_value))
            )

        rows.append(
            {
                "Occupation code": int(code),
                "Occupation label": occupation_labels[int(code)],
                "Occupation N": int(len(occupation_df)),
                "Non-public adjusted probability (%)": round(
                    probability_nonpublic * 100,
                    2,
                ),
                "Non-public 95% CI lower (%)": round(
                    ci_low_nonpublic * 100,
                    2,
                ),
                "Non-public 95% CI upper (%)": round(
                    ci_high_nonpublic * 100,
                    2,
                ),
                "Public adjusted probability (%)": round(
                    probability_public * 100,
                    2,
                ),
                "Public 95% CI lower (%)": round(
                    ci_low_public * 100,
                    2,
                ),
                "Public 95% CI upper (%)": round(
                    ci_high_public * 100,
                    2,
                ),
                "AAPD (public - non-public, pp)": round(
                    difference * 100,
                    2,
                ),
                "AAPD 95% CI lower (pp)": round(
                    difference_ci_low * 100,
                    2,
                ),
                "AAPD 95% CI upper (pp)": round(
                    difference_ci_high * 100,
                    2,
                ),
                "AAPD z": round(z_value, 3),
                "AAPD p-value": p_value,
                "AAPD p-value label": p_value_label(p_value),
                "AAPD significance": significance_mark(p_value),
            }
        )

    margins = pd.DataFrame(rows)

    valid_p = margins["AAPD p-value"].notna()

    margins["AAPD FDR-adjusted p-value"] = np.nan
    margins["AAPD FDR-adjusted p-value label"] = ""
    margins["AAPD FDR significance"] = ""

    if valid_p.any():
        adjusted = multipletests(
            margins.loc[valid_p, "AAPD p-value"],
            alpha=0.05,
            method="fdr_bh",
        )

        margins.loc[
            valid_p,
            "AAPD FDR-adjusted p-value",
        ] = adjusted[1]

        margins.loc[
            valid_p,
            "AAPD FDR-adjusted p-value label",
        ] = [
            p_value_label(p_value)
            for p_value in adjusted[1]
        ]

        margins.loc[
            valid_p,
            "AAPD FDR significance",
        ] = [
            significance_mark(p_value)
            for p_value in adjusted[1]
        ]

    return margins


# ------------------------------------------------------------
# 7. 그림
# ------------------------------------------------------------

def create_margin_plot(margins: pd.DataFrame) -> None:
    """
    직업별 조정확률 차이(AAPD)와 95% CI를 표시합니다.
    기본 matplotlib 스타일과 색상만 사용합니다.
    """
    plot_df = margins.sort_values(
        "AAPD (public - non-public, pp)"
    ).reset_index(drop=True)

    y_positions = np.arange(len(plot_df))

    point_estimates = plot_df[
        "AAPD (public - non-public, pp)"
    ].to_numpy(dtype=float)

    ci_lower = plot_df[
        "AAPD 95% CI lower (pp)"
    ].to_numpy(dtype=float)

    ci_upper = plot_df[
        "AAPD 95% CI upper (pp)"
    ].to_numpy(dtype=float)

    lower_errors = point_estimates - ci_lower
    upper_errors = ci_upper - point_estimates

    labels = [
        (
            f"{code}: {label}"
            if len(label) <= 42
            else f"{code}: {label[:39]}..."
        )
        for code, label in zip(
            plot_df["Occupation code"],
            plot_df["Occupation label"],
        )
    ]

    plt.figure(figsize=(11, 7))
    plt.errorbar(
        point_estimates,
        y_positions,
        xerr=[lower_errors, upper_errors],
        fmt="o",
        capsize=3,
    )
    plt.axvline(x=0, linewidth=1)
    plt.yticks(y_positions, labels)
    plt.xlabel(
        "Adjusted probability difference in detailed explanation "
        "(public − non-public, percentage points)"
    )
    plt.ylabel("Occupation category")
    plt.title(
        "Occupation-Specific Adjusted Differences in Detailed Explanation"
    )
    plt.tight_layout()
    plt.savefig(
        OUTPUT_PNG,
        dpi=300,
        bbox_inches="tight",
    )
    plt.close()


# ------------------------------------------------------------
# 8. 실행
# ------------------------------------------------------------

def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            "전처리 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {INPUT_PATH}"
        )

    raw_df = pd.read_parquet(INPUT_PATH)

    missing_columns = sorted(
        set(REQUIRED_COLUMNS) - set(raw_df.columns)
    )

    if missing_columns:
        raise KeyError(
            "clean_data.parquet에 필요한 변수가 없습니다:\n"
            + "\n".join(
                f"- {column}"
                for column in missing_columns
            )
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    model_df = prepare_complete_case_sample(raw_df)
    occupation_labels = load_occupation_labels()

    print("=" * 76)
    print("직업 유형 구성 및 직업별 공공부문 효과 분석 시작")
    print(f"분석표본 N: {len(model_df):,}")
    print(f"국가 수: {model_df['country_fe'].nunique()}")
    print("=" * 76)

    # A. Occupation composition
    composition_table, composition_test = build_occupation_composition(
        model_df=model_df,
        occupation_labels=occupation_labels,
    )

    # B. Raw detailed-explanation rates by occupation
    raw_rates_table = build_raw_detail_rates(
        model_df=model_df,
        occupation_labels=occupation_labels,
    )

    # C. Country coverage
    country_coverage_table = build_country_coverage(
        model_df=model_df,
        occupation_labels=occupation_labels,
    )

    # D. Baseline Model 2 and interaction model
    print("기준 Model 2 추정")
    (
        baseline_result,
        baseline_covariance_type,
        baseline_warning,
    ) = fit_glm(
        formula=BASELINE_FORMULA,
        model_df=model_df,
    )

    print("직업 유형 상호작용 모형 추정")
    (
        interaction_result,
        interaction_covariance_type,
        interaction_warning,
    ) = fit_glm(
        formula=INTERACTION_FORMULA,
        model_df=model_df,
    )

    model_summary = pd.DataFrame(
        [
            extract_public_sector_effect(
                result=baseline_result,
                specification="Baseline Model 2 (no interaction)",
                model_df=model_df,
                covariance_type=baseline_covariance_type,
                warning_text=baseline_warning,
            ),
            extract_public_sector_effect(
                result=interaction_result,
                specification=(
                    "Occupation interaction model "
                    "(reference occupation only)"
                ),
                model_df=model_df,
                covariance_type=interaction_covariance_type,
                warning_text=interaction_warning,
            ),
        ]
    )

    interaction_test_table = interaction_wald_test(
        interaction_result
    )

    # E. Adjusted probabilities / differences by occupation
    print("직업 유형별 조정예측확률 및 공공부문 차이 산출")
    adjusted_margins_table = build_adjusted_margins(
        interaction_result=interaction_result,
        model_df=model_df,
        occupation_labels=occupation_labels,
    )

    create_margin_plot(adjusted_margins_table)

    # F. Full interaction coefficient table
    ci = interaction_result.conf_int()
    coefficient_rows = []

    for parameter in interaction_result.params.index:
        beta = float(interaction_result.params[parameter])
        p_value = float(interaction_result.pvalues[parameter])
        ci_low = float(ci.loc[parameter].iloc[0])
        ci_high = float(ci.loc[parameter].iloc[1])

        coefficient_rows.append(
            {
                "Parameter": parameter,
                "Beta": round(beta, 5),
                "SE": round(float(interaction_result.bse[parameter]), 5),
                "OR": round(math.exp(beta), 3),
                "OR 95% CI lower": round(math.exp(ci_low), 3),
                "OR 95% CI upper": round(math.exp(ci_high), 3),
                "p-value": p_value,
                "p-value label": p_value_label(p_value),
                "Significance": significance_mark(p_value),
            }
        )

    interaction_coefficients_table = pd.DataFrame(
        coefficient_rows
    )

    # G. Readme
    interpretation_guide = pd.DataFrame(
        [
            {
                "Item": "Occupation composition",
                "Interpretation": (
                    "Shows whether public- and non-public-sector workers "
                    "are distributed differently across occupation categories."
                ),
            },
            {
                "Item": "Raw rates",
                "Interpretation": (
                    "Unadjusted detailed-explanation rates within each occupation. "
                    "Use descriptively; individual occupation-level p-values are "
                    "supplemented by FDR-adjusted values."
                ),
            },
            {
                "Item": "Interaction Wald test",
                "Interpretation": (
                    "Tests whether the adjusted public-sector association differs "
                    "jointly across occupation categories."
                ),
            },
            {
                "Item": "Adjusted margins",
                "Interpretation": (
                    "Within each occupation, covariates retain their observed "
                    "distribution while public_sector is counterfactually set "
                    "to 0 and 1. AAPD is public minus non-public."
                ),
            },
            {
                "Item": "Interpretive limit",
                "Interpretation": (
                    "Do not call occupation a mediator or claim causal mechanisms. "
                    "This is a composition and conditional-heterogeneity diagnostic."
                ),
            },
        ]
    )

    # H. Save outputs
    tables = {
        "01_OccupationComposition": composition_table,
        "02_CompositionOverallTest": composition_test,
        "03_RawDetailRates": raw_rates_table,
        "04_CountryCoverage": country_coverage_table,
        "05_ModelSummary": model_summary,
        "06_InteractionWaldTest": interaction_test_table,
        "07_AdjustedMargins": adjusted_margins_table,
        "08_InteractionCoefficients": interaction_coefficients_table,
        "09_InterpretationGuide": interpretation_guide,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    for sheet_name, dataframe in tables.items():
        save_csv(
            dataframe=dataframe,
            file_name=f"{sheet_name}.csv",
        )

    format_excel(OUTPUT_XLSX)

    # I. Console output
    print("\n[직업 유형 구성 차이]")
    print(composition_test.to_string(index=False))

    print("\n[직업별 비조정 상세 설명 제공률]")
    print(
        raw_rates_table[
            [
                "Occupation code",
                "Occupation label",
                "Non-public detailed explanation (%)",
                "Public detailed explanation (%)",
                "Public - non-public difference (pp)",
                "FDR-adjusted p-value label",
                "FDR significance",
            ]
        ].to_string(index=False)
    )

    print("\n[직업 유형 상호작용 Wald 검정]")
    print(interaction_test_table.to_string(index=False))

    print("\n[직업별 조정확률 차이]")
    print(
        adjusted_margins_table[
            [
                "Occupation code",
                "Occupation label",
                "Non-public adjusted probability (%)",
                "Public adjusted probability (%)",
                "AAPD (public - non-public, pp)",
                "AAPD 95% CI lower (pp)",
                "AAPD 95% CI upper (pp)",
                "AAPD FDR-adjusted p-value label",
                "AAPD FDR significance",
            ]
        ].to_string(index=False)
    )

    print("\n" + "=" * 76)
    print("분석 완료")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"Figure: {OUTPUT_PNG}")
    print("=" * 76)


if __name__ == "__main__":
    main()
