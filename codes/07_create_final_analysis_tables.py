from __future__ import annotations

from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 07_create_final_analysis_tables.py
#
# 목적:
#   results 폴더에 저장된 기존 분석 결과를 불러와
#   논문용 최종 분석표(Excel + CSV)를 자동 생성합니다.
#
# 생성 표:
#   Table 1. 공공·비공공부문 표본 특성 및 이변량 비교
#   Table 2. 국가 고정효과 로지스틱 회귀 결과
#   Table 3. 상세 설명 제공 결과의 강건성 분석
#   Appendix. 결측치, 분석표본, 결과변수 정의
#
# 실행 전제:
#   아래 분석이 이미 results/에 존재해야 합니다.
#   - 01_descriptive_statistics
#   - 02_bivariate_comparison
#   - 03_logistic_regression
#   - 04_model3_sensitivity_diagnostics
#   - 05_loco_detailed_explanation
#   - 06_weighted_model2_sensitivity
# ============================================================


# ------------------------------------------------------------
# 0. 경로
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
RESULTS_ROOT = PROJECT_ROOT / "results"

OUTPUT_DIR = RESULTS_ROOT / "07_final_analysis_tables"
OUTPUT_XLSX = OUTPUT_DIR / "final_analysis_tables.xlsx"


# ------------------------------------------------------------
# 1. 파일 경로 후보
#    운영체제/한글 파일명 차이를 고려해 후보를 순차 탐색합니다.
# ------------------------------------------------------------

FILE_CANDIDATES = {
    "sample_summary": [
        RESULTS_ROOT / "01_descriptive_statistics" / "00_표본요약.csv",
    ],
    "missingness": [
        RESULTS_ROOT / "01_descriptive_statistics" / "01_결측치현황.csv",
    ],
    "outcome_descriptives": [
        RESULTS_ROOT / "01_descriptive_statistics" / "02_핵심결과변수_부문별.csv",
    ],
    "continuous_descriptives": [
        RESULTS_ROOT / "01_descriptive_statistics" / "04_연속형통제변수.csv",
    ],
    "categorical_descriptives": [
        RESULTS_ROOT / "01_descriptive_statistics" / "05_범주형통제변수.csv",
    ],
    "binary_tests": [
        RESULTS_ROOT / "02_bivariate_comparison" / "01_핵심결과_차이검정.csv",
    ],
    "exposure_tests": [
        RESULTS_ROOT / "02_bivariate_comparison" / "02_자동화관리경험_차이검정.csv",
    ],
    "continuous_tests": [
        RESULTS_ROOT / "02_bivariate_comparison" / "03_연속형통제_차이검정.csv",
    ],
    "categorical_distribution": [
        RESULTS_ROOT / "02_bivariate_comparison" / "04_범주형통제_분포.csv",
    ],
    "categorical_tests": [
        RESULTS_ROOT / "02_bivariate_comparison" / "05_범주형통제_차이검정.csv",
    ],
    "main_regression": [
        RESULTS_ROOT / "03_logistic_regression" / "02_공공부문효과_메인.csv",
    ],
    "same_sample": [
        RESULTS_ROOT
        / "04_model3_sensitivity_diagnostics"
        / "01_동일표본_공공부문효과.csv",
    ],
    "loco_summary": [
        RESULTS_ROOT
        / "05_loco_detailed_explanation"
        / "02_LOCO_요약.csv",
    ],
    "weighted": [
        RESULTS_ROOT
        / "06_weighted_model2_sensitivity"
        / "01_가중치민감도_전체.csv",
    ],
}


# ------------------------------------------------------------
# 2. 공통 함수
# ------------------------------------------------------------

def find_file(candidates: Iterable[Path], file_key: str) -> Path:
    """후보 경로 중 실제 존재하는 첫 파일을 반환합니다."""
    for path in candidates:
        if path.exists():
            return path

    paths_text = "\n".join(f"- {path}" for path in candidates)

    raise FileNotFoundError(
        f"[{file_key}] 결과 파일을 찾지 못했습니다.\n"
        f"확인한 경로:\n{paths_text}\n\n"
        "기존 분석 스크립트를 모두 정상 실행했는지 확인하세요."
    )


def read_result_csv(file_key: str) -> pd.DataFrame:
    """UTF-8-SIG CSV를 읽습니다."""
    path = find_file(FILE_CANDIDATES[file_key], file_key)
    return pd.read_csv(path, encoding="utf-8-sig")


def safe_value(
    dataframe: pd.DataFrame,
    filter_column: str,
    filter_value: str,
    target_column: str,
    default=np.nan,
):
    """필터 조건에 맞는 첫 번째 값을 안전하게 가져옵니다."""
    subset = dataframe.loc[
        dataframe[filter_column].astype(str) == str(filter_value)
    ]

    if subset.empty or target_column not in subset.columns:
        return default

    return subset.iloc[0][target_column]


def format_p(p_value) -> str:
    """p-value를 논문 표기 형식으로 정리합니다."""
    if pd.isna(p_value):
        return ""

    p_value = float(p_value)

    if p_value < 0.001:
        return "< .001"

    return f"{p_value:.3f}"


def format_mean_sd(mean, sd) -> str:
    if pd.isna(mean) or pd.isna(sd):
        return ""
    return f"{float(mean):.2f} ({float(sd):.2f})"


def format_n_pct(n, pct) -> str:
    if pd.isna(n) or pd.isna(pct):
        return ""
    return f"{int(n):,} ({float(pct):.2f}%)"


def format_or_ci(or_value, ci_low, ci_high) -> str:
    if pd.isna(or_value) or pd.isna(ci_low) or pd.isna(ci_high):
        return ""
    return (
        f"{float(or_value):.3f} "
        f"[{float(ci_low):.3f}, {float(ci_high):.3f}]"
    )


def format_effect_size(value, digits: int = 3) -> str:
    if pd.isna(value):
        return ""
    return f"{float(value):.{digits}f}"


def first_non_null(*values):
    """첫 번째 결측이 아닌 값을 반환."""
    for value in values:
        if not pd.isna(value):
            return value
    return np.nan


def make_star(p_value) -> str:
    if pd.isna(p_value):
        return ""
    p_value = float(p_value)
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def blank_row() -> dict:
    return {
        "Domain": "",
        "Variable": "",
        "Non-public sector": "",
        "Public sector": "",
        "Difference / Test": "",
        "p-value": "",
        "Effect size": "",
        "Notes": "",
    }


# ------------------------------------------------------------
# 3. Table 1: 기술통계 및 이변량 비교
# ------------------------------------------------------------

def build_table_1(
    outcome_descriptives: pd.DataFrame,
    continuous_descriptives: pd.DataFrame,
    categorical_descriptives: pd.DataFrame,
    binary_tests: pd.DataFrame,
    exposure_tests: pd.DataFrame,
    continuous_tests: pd.DataFrame,
    categorical_distribution: pd.DataFrame,
    categorical_tests: pd.DataFrame,
) -> pd.DataFrame:
    """공공·비공공부문 기술통계 및 이변량 비교 표를 구성합니다."""

    rows: list[dict] = []

    # --------------------------------------------------------
    # A. 정보공개·접근권 결과변수
    # --------------------------------------------------------

    rows.append(
        {
            "Domain": "A. Information disclosure and access",
            "Variable": "",
            "Non-public sector": "",
            "Public sector": "",
            "Difference / Test": "",
            "p-value": "",
            "Effect size": "",
            "Notes": "",
        }
    )

    outcome_order = [
        "상세 설명 제공",
        "개인정보 접근권 제공",
        "자동화 분석 결과 접근권 제공",
        "기술 사용 사실 무고지",
        "단순 고지",
    ]

    for label in outcome_order:
        test_row = binary_tests.loc[
            binary_tests["변수 라벨"] == label
        ]

        if test_row.empty:
            continue

        test_row = test_row.iloc[0]

        rows.append(
            {
                "Domain": "",
                "Variable": label,
                "Non-public sector": format_n_pct(
                    test_row["비공공부문 예(1) N"],
                    test_row["비공공부문 예(1) 비율(%)"],
                ),
                "Public sector": format_n_pct(
                    test_row["공공부문 예(1) N"],
                    test_row["공공부문 예(1) 비율(%)"],
                ),
                "Difference / Test": (
                    f"{float(test_row['비율 차이(%p, 공공-비공공)']):+.2f} pp; "
                    f"χ²={float(test_row['카이제곱']):.3f}"
                ),
                "p-value": (
                    f"{test_row['p-value 표기']}"
                    f"{test_row['유의성']}"
                ),
                "Effect size": (
                    f"Phi={format_effect_size(test_row['Phi 효과크기'], 4)}"
                ),
                "Notes": "Binary outcome",
            }
        )

    rows.append(blank_row())

    # --------------------------------------------------------
    # B. 자동화 관리 경험
    # --------------------------------------------------------

    rows.append(
        {
            "Domain": "B. Automated management exposure",
            "Variable": "",
            "Non-public sector": "",
            "Public sector": "",
            "Difference / Test": "",
            "p-value": "",
            "Effect size": "",
            "Notes": "",
        }
    )

    for _, test_row in exposure_tests.iterrows():
        rows.append(
            {
                "Domain": "",
                "Variable": test_row["변수 라벨"],
                "Non-public sector": format_n_pct(
                    test_row["비공공부문 예(1) N"],
                    test_row["비공공부문 예(1) 비율(%)"],
                ),
                "Public sector": format_n_pct(
                    test_row["공공부문 예(1) N"],
                    test_row["공공부문 예(1) 비율(%)"],
                ),
                "Difference / Test": (
                    f"{float(test_row['비율 차이(%p, 공공-비공공)']):+.2f} pp; "
                    f"χ²={float(test_row['카이제곱']):.3f}"
                ),
                "p-value": (
                    f"{test_row['p-value 표기']}"
                    f"{test_row['유의성']}"
                ),
                "Effect size": (
                    f"Phi={format_effect_size(test_row['Phi 효과크기'], 4)}"
                ),
                "Notes": "Supplementary exposure",
            }
        )

    rows.append(blank_row())

    # --------------------------------------------------------
    # C. 연속형 통제변수
    # --------------------------------------------------------

    rows.append(
        {
            "Domain": "C. Continuous covariates",
            "Variable": "",
            "Non-public sector": "",
            "Public sector": "",
            "Difference / Test": "",
            "p-value": "",
            "Effect size": "",
            "Notes": "",
        }
    )

    continuous_order = [
        "연령",
        "전일제 교육 종료 연령",
        "직무 수행 디지털 역량",
    ]

    for label in continuous_order:
        test_row = continuous_tests.loc[
            continuous_tests["변수 라벨"] == label
        ]

        if test_row.empty:
            continue

        test_row = test_row.iloc[0]

        rows.append(
            {
                "Domain": "",
                "Variable": label,
                "Non-public sector": format_mean_sd(
                    test_row["비공공부문 평균"],
                    test_row["비공공부문 표준편차"],
                ),
                "Public sector": format_mean_sd(
                    test_row["공공부문 평균"],
                    test_row["공공부문 표준편차"],
                ),
                "Difference / Test": (
                    f"{float(test_row['평균 차이(공공-비공공)']):+.2f}; "
                    f"t={float(test_row['Welch t']):.3f}"
                ),
                "p-value": (
                    f"{test_row['p-value 표기']}"
                    f"{test_row['유의성']}"
                ),
                "Effect size": (
                    f'd={format_effect_size(test_row["Cohen's d"])}' 
                ),
                "Notes": "Mean (SD)",
            }
        )

    rows.append(blank_row())

    # --------------------------------------------------------
    # D. 범주형 통제변수
    # --------------------------------------------------------

    rows.append(
        {
            "Domain": "D. Categorical covariates",
            "Variable": "",
            "Non-public sector": "",
            "Public sector": "",
            "Difference / Test": "",
            "p-value": "",
            "Effect size": "",
            "Notes": "",
        }
    )

    categorical_order = [
        "성별",
        "직업 유형",
        "조직 규모",
    ]

    for variable_label in categorical_order:
        test_row = categorical_tests.loc[
            categorical_tests["변수 라벨"] == variable_label
        ]

        if test_row.empty:
            continue

        test_row = test_row.iloc[0]

        rows.append(
            {
                "Domain": "",
                "Variable": variable_label,
                "Non-public sector": "",
                "Public sector": "",
                "Difference / Test": (
                    f"χ²={float(test_row['카이제곱']):.3f}"
                ),
                "p-value": (
                    f"{test_row['p-value 표기']}"
                    f"{test_row['유의성']}"
                ),
                "Effect size": (
                    f"Cramer's V="
                    f'{format_effect_size(test_row["Cramer's V"], 4)}' 
                ),
                "Notes": "Category distribution shown below",
            }
        )

        category_rows = categorical_distribution.loc[
            categorical_distribution["변수 라벨"] == variable_label
        ].copy()

        for _, category_row in category_rows.iterrows():
            rows.append(
                {
                    "Domain": "",
                    "Variable": f"   {category_row['범주']}",
                    "Non-public sector": format_n_pct(
                        category_row["비공공부문 N"],
                        category_row["비공공부문 비율(%)"],
                    ),
                    "Public sector": format_n_pct(
                        category_row["공공부문 N"],
                        category_row["공공부문 비율(%)"],
                    ),
                    "Difference / Test": (
                        f"{float(category_row['비율 차이(%p, 공공-비공공)']):+.2f} pp"
                    ),
                    "p-value": "",
                    "Effect size": "",
                    "Notes": "",
                }
            )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 4. Table 2: 메인 회귀결과
# ------------------------------------------------------------

def build_table_2(main_regression: pd.DataFrame) -> pd.DataFrame:
    """
    메인 회귀표:
    Outcome별 Model 1 / Model 2 결과를 wide format으로 정리합니다.
    """

    outcome_order = [
        "상세 설명 제공",
        "개인정보 접근권 제공",
        "자동화 분석 결과 접근권 제공",
        "기술 사용 사실 무고지",
    ]

    rows: list[dict] = []

    for outcome_label in outcome_order:
        subset = main_regression.loc[
            main_regression["결과변수 라벨"] == outcome_label
        ].copy()

        model_1 = subset.loc[subset["모형"] == "Model 1"]
        model_2 = subset.loc[subset["모형"] == "Model 2"]

        if model_1.empty or model_2.empty:
            continue

        model_1 = model_1.iloc[0]
        model_2 = model_2.iloc[0]

        rows.append(
            {
                "Outcome": outcome_label,
                "Model 1 OR [95% CI]": format_or_ci(
                    model_1["오즈비(OR)"],
                    model_1["OR 95% CI 하한"],
                    model_1["OR 95% CI 상한"],
                ),
                "Model 1 p": (
                    f"{model_1['p-value 표기']}{model_1['유의성']}"
                ),
                "Model 1 AAPD (pp)": (
                    f"{float(model_1['조정확률 차이(%p, 공공-비공공)']):+.2f}"
                ),
                "Model 1 N": int(model_1["표본 N"]),
                "Model 2 OR [95% CI]": format_or_ci(
                    model_2["오즈비(OR)"],
                    model_2["OR 95% CI 하한"],
                    model_2["OR 95% CI 상한"],
                ),
                "Model 2 p": (
                    f"{model_2['p-value 표기']}{model_2['유의성']}"
                ),
                "Model 2 AAPD (pp)": (
                    f"{float(model_2['조정확률 차이(%p, 공공-비공공)']):+.2f}"
                ),
                "Model 2 N": int(model_2["표본 N"]),
                "Country FE": "Yes",
                "Individual/workplace controls": "No / Yes",
            }
        )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 5. Table 3: 상세 설명 제공 강건성 분석
# ------------------------------------------------------------

def build_table_3(
    main_regression: pd.DataFrame,
    same_sample: pd.DataFrame,
    weighted: pd.DataFrame,
    loco_summary: pd.DataFrame,
) -> pd.DataFrame:
    """상세 설명 제공의 핵심 강건성 결과를 정리합니다."""

    rows: list[dict] = []

    # Main Model 2
    main_model_2 = main_regression.loc[
        (main_regression["결과변수"] == "detailed_explanation")
        & (main_regression["모형"] == "Model 2")
    ]

    if not main_model_2.empty:
        row = main_model_2.iloc[0]
        rows.append(
            {
                "Specification": "Main Model 2",
                "N": int(row["표본 N"]),
                "OR [95% CI]": format_or_ci(
                    row["오즈비(OR)"],
                    row["OR 95% CI 하한"],
                    row["OR 95% CI 상한"],
                ),
                "p-value": f"{row['p-value 표기']}{row['유의성']}",
                "AAPD (pp)": (
                    f"{float(row['조정확률 차이(%p, 공공-비공공)']):+.2f}"
                ),
                "Interpretation / note": (
                    "Country FE + demographic, occupational, and workplace controls"
                ),
            }
        )

    # Same-sample restricted Model 2
    restricted_model_2 = same_sample.loc[
        (
            same_sample["결과변수"] == "detailed_explanation"
        )
        & (
            same_sample["모형"] == "Restricted Model 2 (same N)"
        )
    ]

    if not restricted_model_2.empty:
        row = restricted_model_2.iloc[0]
        rows.append(
            {
                "Specification": "Restricted Model 2 (same N)",
                "N": int(row["표본 N"]),
                "OR [95% CI]": format_or_ci(
                    row["오즈비(OR)"],
                    row["OR 95% CI 하한"],
                    row["OR 95% CI 상한"],
                ),
                "p-value": f"{row['p-value 표기']}{row['유의성']}",
                "AAPD (pp)": (
                    f"{float(row['조정확률 차이(%p, 공공-비공공)']):+.2f}"
                ),
                "Interpretation / note": (
                    "Same complete-case sample used for Model 3"
                ),
            }
        )

    # Model 3
    model_3 = same_sample.loc[
        (same_sample["결과변수"] == "detailed_explanation")
        & (same_sample["모형"] == "Model 3")
    ]

    if not model_3.empty:
        row = model_3.iloc[0]
        rows.append(
            {
                "Specification": "Model 3",
                "N": int(row["표본 N"]),
                "OR [95% CI]": format_or_ci(
                    row["오즈비(OR)"],
                    row["OR 95% CI 하한"],
                    row["OR 95% CI 상한"],
                ),
                "p-value": f"{row['p-value 표기']}{row['유의성']}",
                "AAPD (pp)": (
                    f"{float(row['조정확률 차이(%p, 공공-비공공)']):+.2f}"
                ),
                "Interpretation / note": (
                    "Model 2 + automated monitoring and performance-management exposure"
                ),
            }
        )

    # Weighted Model 2
    weighted_model = weighted.loc[
        (weighted["결과변수"] == "detailed_explanation")
        & (
            weighted["모형"] == "Weighted GLM (w1 normalized)"
        )
    ]

    if not weighted_model.empty:
        row = weighted_model.iloc[0]
        rows.append(
            {
                "Specification": "Weighted Model 2 (w1)",
                "N": int(row["표본 N"]),
                "OR [95% CI]": format_or_ci(
                    row["오즈비(OR)"],
                    row["OR 95% CI 하한"],
                    row["OR 95% CI 상한"],
                ),
                "p-value": f"{row['p-value 표기']}{row['유의성']}",
                "AAPD (pp)": (
                    f"{float(row['조정확률 차이(%p, 공공-비공공)']):+.2f}"
                ),
                "Interpretation / note": (
                    "Weighted pseudo-likelihood sensitivity analysis"
                ),
            }
        )

    # LOCO
    if not loco_summary.empty:
        row = loco_summary.iloc[0]
        rows.append(
            {
                "Specification": "Leave-one-country-out",
                "N": "Model-specific",
                "OR [95% CI]": (
                    f"Range: {float(row['반복 OR 최솟값']):.3f}"
                    f"–{float(row['반복 OR 최댓값']):.3f}"
                ),
                "p-value": (
                    f"{int(row['p < .05 반복 수'])}/"
                    f"{int(row['성공한 국가 제외 반복 수'])} "
                    "replications p < .05"
                ),
                "AAPD (pp)": "",
                "Interpretation / note": (
                    f"OR < 1 in {int(row['OR < 1 반복 수'])}/"
                    f"{int(row['성공한 국가 제외 반복 수'])} "
                    "replications"
                ),
            }
        )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 6. Appendix: 표본 및 결측치
# ------------------------------------------------------------

def build_appendix_sample(
    sample_summary: pd.DataFrame,
    missingness: pd.DataFrame,
) -> pd.DataFrame:
    """분석표본과 결측치 요약을 한 시트에 배치합니다."""
    sample_table = sample_summary.copy()
    sample_table.insert(0, "Section", "Analysis sample")

    missing_table = missingness.copy()
    missing_table.insert(0, "Section", "Missingness")

    # 열 구조를 맞추기 위해 별도 표를 간단히 정리
    sample_table = sample_table.rename(
        columns={
            "집단": "Variable / Group",
            "N": "N",
            "전체 대비 비율(%)": "Percent",
        }
    )[
        ["Section", "Variable / Group", "N", "Percent"]
    ]

    missing_table = missing_table.rename(
        columns={
            "변수명": "Variable / Group",
            "결측 N": "N",
            "결측 비율(%)": "Percent",
        }
    )[
        ["Section", "Variable / Group", "N", "Percent"]
    ]

    return pd.concat(
        [sample_table, missing_table],
        ignore_index=True,
    )


# ------------------------------------------------------------
# 7. Workbook formatting
# ------------------------------------------------------------

def format_workbook(excel_path: Path) -> None:
    """논문용 분석표에 맞게 기본 서식을 적용합니다."""

    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    section_fill = PatternFill(
        fill_type="solid",
        fgColor="EAF2F8",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        # Header
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Body
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        # Table 1 section rows
        if worksheet.title == "Table1_Characteristics":
            for row_idx in range(2, worksheet.max_row + 1):
                domain_value = worksheet.cell(row=row_idx, column=1).value
                variable_value = worksheet.cell(row=row_idx, column=2).value

                if domain_value and not variable_value:
                    for cell in worksheet[row_idx]:
                        cell.font = Font(bold=True)
                        cell.fill = section_fill

        # Width
        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 58))

            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = width

        # Row height
        for row_idx in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 24

    workbook.save(excel_path)


# ------------------------------------------------------------
# 8. Main
# ------------------------------------------------------------

def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 72)
    print("최종 분석표 생성 시작")
    print(f"결과 원본 폴더: {RESULTS_ROOT}")
    print("=" * 72)

    # Read all source outputs
    sample_summary = read_result_csv("sample_summary")
    missingness = read_result_csv("missingness")

    outcome_descriptives = read_result_csv("outcome_descriptives")
    continuous_descriptives = read_result_csv("continuous_descriptives")
    categorical_descriptives = read_result_csv(
        "categorical_descriptives"
    )

    binary_tests = read_result_csv("binary_tests")
    exposure_tests = read_result_csv("exposure_tests")
    continuous_tests = read_result_csv("continuous_tests")
    categorical_distribution = read_result_csv(
        "categorical_distribution"
    )
    categorical_tests = read_result_csv("categorical_tests")

    main_regression = read_result_csv("main_regression")
    same_sample = read_result_csv("same_sample")
    loco_summary = read_result_csv("loco_summary")
    weighted = read_result_csv("weighted")

    # Build tables
    table_1 = build_table_1(
        outcome_descriptives=outcome_descriptives,
        continuous_descriptives=continuous_descriptives,
        categorical_descriptives=categorical_descriptives,
        binary_tests=binary_tests,
        exposure_tests=exposure_tests,
        continuous_tests=continuous_tests,
        categorical_distribution=categorical_distribution,
        categorical_tests=categorical_tests,
    )

    table_2 = build_table_2(main_regression)

    table_3 = build_table_3(
        main_regression=main_regression,
        same_sample=same_sample,
        weighted=weighted,
        loco_summary=loco_summary,
    )

    appendix_sample = build_appendix_sample(
        sample_summary=sample_summary,
        missingness=missingness,
    )

    # Save workbook
    tables = {
        "Table1_Characteristics": table_1,
        "Table2_MainRegression": table_2,
        "Table3_Robustness": table_3,
        "Appendix_SampleMissing": appendix_sample,
        "Raw_MainRegression": main_regression,
        "Raw_WeightedSensitivity": weighted,
        "Raw_LOCO": loco_summary,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    # Save core paper tables as CSV too
    table_1.to_csv(
        OUTPUT_DIR / "Table1_Characteristics.csv",
        index=False,
        encoding="utf-8-sig",
    )

    table_2.to_csv(
        OUTPUT_DIR / "Table2_MainRegression.csv",
        index=False,
        encoding="utf-8-sig",
    )

    table_3.to_csv(
        OUTPUT_DIR / "Table3_Robustness.csv",
        index=False,
        encoding="utf-8-sig",
    )

    format_workbook(OUTPUT_XLSX)

    print("\n[생성 완료]")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"CSV folder: {OUTPUT_DIR}")

    print("\n[Table 2: Main regression]")
    print(table_2.to_string(index=False))

    print("\n[Table 3: Robustness]")
    print(table_3.to_string(index=False))

    print("=" * 72)
    print("최종 분석표 생성 완료")
    print("=" * 72)


if __name__ == "__main__":
    main()
