from __future__ import annotations

from pathlib import Path
import math

import numpy as np
import pandas as pd
import patsy
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import t as student_t


# ============================================================
# 17_wild_cluster_bootstrap_logit.py
#
# 목적
# - Table 2의 조정 로짓모형(RQ1~RQ4)에 대해
#   restricted wild-cluster linearized bootstrap(WCLR) p-value를 산출
# - 국가 단위 클러스터 수 G=27의 소표본 추론 문제를 점검
#
# 방법
# - H0: public_sector coefficient = 0을 부과한 restricted logit 적합
# - 국가별 empirical score에 Rademacher weight(+1/-1)를 적용
# - WCLR-C(classic)와 WCLR-S(transformed score)를 모두 산출
#
# 출력
# - results/17_wild_cluster_bootstrap_logit/
#   ├─ 17_wild_cluster_bootstrap_logit.xlsx
#   ├─ 01_AppendixTable_A3.csv
#   ├─ 02_Full_WCLR_Results.csv
#   ├─ 03_Design_Diagnostics.csv
#   └─ 04_Notes.csv
#
# 주의
# - 11_common_sample_main_models.py와 동일한
#   master complete-case sample을 재현하도록 작성
# - WCLR p-value는 Table 2의 OR/AAPD를 대체하지 않고,
#   국가 27개 클러스터의 소표본 추론을 점검하는 부록 분석임
# ============================================================


# ------------------------------------------------------------
# 0. Paths
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "17_wild_cluster_bootstrap_logit"
OUTPUT_XLSX = OUTPUT_DIR / "17_wild_cluster_bootstrap_logit.xlsx"


# ------------------------------------------------------------
# 1. Analysis settings
# ------------------------------------------------------------

RANDOM_SEED = 20260630
BOOTSTRAP_REPS = 9_999
TARGET_PARAMETER = "public_sector"

OUTCOMES = {
    "detailed_explanation": "상세 설명 경험 (RQ1)",
    "personal_data_access": "개인정보 접근 경험 (RQ2)",
    "automated_analysis_access": "자동화 분석 결과 접근 경험 (RQ3)",
    "not_informed": "기술 사용 사실 무고지 경험 (RQ4)",
}

CONTROL_COLUMNS = [
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

CATEGORICAL_CONTROL_COLUMNS = [
    "gender",
    "occupation_code",
    "workplace_size",
]

REQUIRED_COLUMNS = [
    TARGET_PARAMETER,
    "country_fe",
    *OUTCOMES.keys(),
    *CONTROL_COLUMNS,
]


# ------------------------------------------------------------
# 2. Utilities
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")

    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")

    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)

    return pd.Series(values, index=series.index, dtype=object)


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""

    if p_value < 0.001:
        return "< .001"

    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""

    if p_value < 0.001:
        return "***"

    if p_value < 0.01:
        return "**"

    if p_value < 0.05:
        return "*"

    return ""


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 55))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

        for row_number in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 24

    workbook.save(excel_path)


def safe_inverse(matrix: np.ndarray, label: str) -> np.ndarray:
    """
    Information matrix가 수치적으로 거의 특이한 경우에는
    Moore-Penrose generalized inverse를 사용합니다.
    """
    try:
        return np.linalg.inv(matrix)
    except np.linalg.LinAlgError:
        print(
            f"[주의] {label}가 특이행렬입니다. "
            "Moore-Penrose generalized inverse를 사용합니다."
        )
        return np.linalg.pinv(matrix)


# ------------------------------------------------------------
# 3. Common complete-case sample
# ------------------------------------------------------------

def prepare_common_sample(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    11_common_sample_main_models.py와 동일한 master complete-case sample.
    """
    model_df = dataframe[REQUIRED_COLUMNS].copy()

    numeric_columns = [
        TARGET_PARAMETER,
        *OUTCOMES.keys(),
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]

    for column in numeric_columns:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    valid_outcome_mask = np.ones(len(model_df), dtype=bool)

    for outcome in OUTCOMES:
        valid_outcome_mask &= model_df[outcome].isin([0.0, 1.0])

    model_df = model_df[
        valid_outcome_mask
        & model_df[TARGET_PARAMETER].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
    ].copy()

    model_df = model_df.dropna().copy()

    for outcome in OUTCOMES:
        model_df[outcome] = to_int64(model_df[outcome])

    model_df[TARGET_PARAMETER] = to_int64(
        model_df[TARGET_PARAMETER]
    )

    for column in ["age", "education_age", "digital_skill_job"]:
        model_df[column] = to_float64(model_df[column])

    for column in CATEGORICAL_CONTROL_COLUMNS:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


# ------------------------------------------------------------
# 4. Formula and design matrix
# ------------------------------------------------------------

def adjusted_formula(outcome: str) -> str:
    """
    11_common_sample_main_models.py의 조정모형과 동일.
    """
    return (
        f"{outcome} ~ {TARGET_PARAMETER} + C(country_fe)"
        " + age + C(gender) + education_age"
        " + C(occupation_code) + C(workplace_size)"
        " + digital_skill_job"
    )


def build_reordered_design(
    outcome: str,
    model_df: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, list[str]]:
    """
    WCLR-S restricted-score 변환의 표기에 맞춰
    검정 대상 public_sector를 design matrix 첫 번째 열로 배치합니다.
    """
    y_df, x_df = patsy.dmatrices(
        adjusted_formula(outcome),
        data=model_df,
        return_type="dataframe",
        NA_action="raise",
    )

    if TARGET_PARAMETER not in x_df.columns:
        raise KeyError(
            f"Design matrix에서 {TARGET_PARAMETER!r}를 찾지 못했습니다.\n"
            f"Columns: {list(x_df.columns)}"
        )

    ordered_columns = [
        TARGET_PARAMETER,
        *[
            column
            for column in x_df.columns
            if column != TARGET_PARAMETER
        ],
    ]

    x = x_df.loc[:, ordered_columns].to_numpy(dtype=np.float64)
    y = y_df.iloc[:, 0].to_numpy(dtype=np.float64)

    clusters = (
        model_df.loc[x_df.index, "country_fe"]
        .astype(str)
        .to_numpy(dtype=object)
    )

    return y, x, clusters, ordered_columns


# ------------------------------------------------------------
# 5. Score / information matrix helpers
# ------------------------------------------------------------

def cluster_scores_and_information(
    y: np.ndarray,
    x: np.ndarray,
    probabilities: np.ndarray,
    cluster_codes: np.ndarray,
    n_clusters: int,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Returns
    -------
    scores:
        G x k matrix; row g is empirical score vector s_g(beta).
    information:
        G x k x k array; information[g] is J_g(beta).
    """
    n_parameters = x.shape[1]

    scores = np.zeros(
        (n_clusters, n_parameters),
        dtype=np.float64,
    )

    information = np.zeros(
        (n_clusters, n_parameters, n_parameters),
        dtype=np.float64,
    )

    residual = y - probabilities
    logistic_weight = probabilities * (1.0 - probabilities)

    for g in range(n_clusters):
        mask = cluster_codes == g
        x_g = x[mask]
        residual_g = residual[mask]
        weight_g = logistic_weight[mask]

        scores[g] = x_g.T @ residual_g
        information[g] = x_g.T @ (weight_g[:, None] * x_g)

    return scores, information


def cv1_variance(
    information: np.ndarray,
    scores: np.ndarray,
    n_observations: int,
    n_parameters: int,
    n_clusters: int,
) -> np.ndarray:
    """
    Cluster-robust CV1 sandwich variance:
      G/(G-1) * (N-1)/(N-k) * J^{-1} sum(s_g s_g') J^{-1}
    """
    information_total = information.sum(axis=0)

    information_inverse = safe_inverse(
        information_total,
        label="CV1 information matrix",
    )

    finite_sample_factor = (
        n_clusters / (n_clusters - 1)
        * (n_observations - 1) / (n_observations - n_parameters)
    )

    return (
        finite_sample_factor
        * information_inverse
        @ (scores.T @ scores)
        @ information_inverse
    )


# ------------------------------------------------------------
# 6. Restricted WCLR bootstrap
# ------------------------------------------------------------

def restricted_wclr_p_values(
    y: np.ndarray,
    x: np.ndarray,
    clusters: np.ndarray,
    reps: int,
    seed: int,
) -> dict:
    """
    H0: beta_public_sector = 0에 대한 restricted WCLR-C / WCLR-S.

    - WCLR-C: restricted empirical scores를 직접 사용
    - WCLR-S: nuisance parameter 추정으로 발생한 score distortion을
               leave-one-cluster linearization으로 보정
    - G=27이므로 Rademacher weights 사용
    """
    n_observations, n_parameters = x.shape

    cluster_labels, cluster_codes = np.unique(
        clusters,
        return_inverse=True,
    )

    n_clusters = len(cluster_labels)

    if n_clusters < 13:
        raise ValueError(
            "이 스크립트는 G>=13일 때의 Rademacher weights를 사용합니다. "
            f"현재 G={n_clusters}입니다."
        )

    # x[:, 0] = public_sector
    x_restricted = x[:, 1:]

    # --------------------------------------------------------
    # A. Unrestricted full model: observed coefficient / t
    # --------------------------------------------------------
    full_model = sm.Logit(y, x)

    full_result = full_model.fit(
        disp=False,
        maxiter=500,
    )

    # Table 2의 conventional cluster-robust z p-value와 대조용
    full_cluster_result = full_model.fit(
        disp=False,
        maxiter=500,
        cov_type="cluster",
        cov_kwds={
            "groups": clusters,
            "use_correction": True,
        },
    )

    beta_hat = np.asarray(full_result.params, dtype=np.float64)
    p_hat = np.asarray(full_result.predict(), dtype=np.float64)

    scores_hat, information_hat = cluster_scores_and_information(
        y=y,
        x=x,
        probabilities=p_hat,
        cluster_codes=cluster_codes,
        n_clusters=n_clusters,
    )

    variance_cv1 = cv1_variance(
        information=information_hat,
        scores=scores_hat,
        n_observations=n_observations,
        n_parameters=n_parameters,
        n_clusters=n_clusters,
    )

    se_cv1 = float(math.sqrt(variance_cv1[0, 0]))
    t_observed = float(beta_hat[0] / se_cv1)

    p_cv1_t = float(
        2.0 * student_t.sf(
            abs(t_observed),
            df=n_clusters - 1,
        )
    )

    # --------------------------------------------------------
    # B. Restricted model under H0: public_sector = 0
    # --------------------------------------------------------
    restricted_model = sm.Logit(y, x_restricted)

    restricted_result = restricted_model.fit(
        disp=False,
        maxiter=500,
    )

    beta_restricted = np.asarray(
        restricted_result.params,
        dtype=np.float64,
    )

    beta_tilde = np.concatenate(
        [
            np.array([0.0], dtype=np.float64),
            beta_restricted,
        ]
    )

    linear_predictor_tilde = x @ beta_tilde
    p_tilde = 1.0 / (1.0 + np.exp(-linear_predictor_tilde))

    # Restricted nuisance score / information: (k-1) dimensions
    scores_r, information_r = cluster_scores_and_information(
        y=y,
        x=x_restricted,
        probabilities=p_tilde,
        cluster_codes=cluster_codes,
        n_clusters=n_clusters,
    )

    # Restricted probabilities evaluated on full X:
    # target score is retained, so dimensions are k
    scores_r2, information_r2 = cluster_scores_and_information(
        y=y,
        x=x,
        probabilities=p_tilde,
        cluster_codes=cluster_codes,
        n_clusters=n_clusters,
    )

    information_r_total = information_r.sum(axis=0)
    information_r2_total = information_r2.sum(axis=0)

    information_r2_inverse = safe_inverse(
        information_r2_total,
        label="Restricted full information matrix",
    )

    score_r_total = scores_r.sum(axis=0)

    # --------------------------------------------------------
    # C. WCLR-S transformed restricted scores
    #
    # 국가 fixed effect 때문에 특정 국가를 하나 제외하면
    # 해당 국가 dummy column이 0이 됩니다.
    # 따라서 leave-one-country nuisance information matrix에는
    # generalized inverse(pinv)를 사용합니다.
    # --------------------------------------------------------
    scores_r2_s = np.zeros_like(scores_r2)

    for g in range(n_clusters):
        leave_one_information = (
            information_r_total - information_r[g]
        )

        leave_one_score = score_r_total - scores_r[g]

        beta_leave_one_out_nuisance = (
            np.linalg.pinv(leave_one_information)
            @ leave_one_score
        )

        # Target parameter(public_sector)는 첫 번째 열.
        # Formula analogous to restricted transformed-score equation:
        # s_dot_g = s_tilde_g - J_tilde_g[:, nuisance] b_tilde_(g)
        scores_r2_s[g] = (
            scores_r2[g]
            - information_r2[g][:, 1:]
            @ beta_leave_one_out_nuisance
        )

    # --------------------------------------------------------
    # D. Bootstrap setup
    # --------------------------------------------------------
    rng = np.random.default_rng(seed)

    # B x G Rademacher weights: +1 / -1 with equal probability
    wild_weights = rng.choice(
        np.array([-1.0, 1.0]),
        size=(reps, n_clusters),
        replace=True,
    )

    finite_sample_factor = (
        n_clusters / (n_clusters - 1)
        * (n_observations - 1) / (n_observations - n_parameters)
    )

    # First row of J^{-1}; sufficient for Var(beta*_public_sector)
    target_influence = information_r2_inverse[0, :]

    def bootstrap_t_statistics(
        bootstrap_scores: np.ndarray,
    ) -> np.ndarray:
        """
        Nonlinear logit model을 bootstrap 반복마다 다시 적합하지 않고,
        empirical score와 information matrix를 이용해
        WCLR t-statistic을 계산합니다.
        """
        # b* = J^{-1} sum_g(v_g* s_g)
        score_sums = wild_weights @ bootstrap_scores
        beta_star = score_sums @ information_r2_inverse.T

        # beta*_public_sector의 bootstrap CV1 variance만 계산
        target_variance_numerator = np.zeros(
            reps,
            dtype=np.float64,
        )

        for g in range(n_clusters):
            # w_g* = v_g*s_g - J_g*b*
            w_g = (
                wild_weights[:, [g]]
                * bootstrap_scores[g][None, :]
                - beta_star @ information_r2[g].T
            )

            projected_w_g = w_g @ target_influence
            target_variance_numerator += projected_w_g**2

        bootstrap_se_target = np.sqrt(
            finite_sample_factor * target_variance_numerator
        )

        if np.any(bootstrap_se_target <= 0.0):
            raise RuntimeError(
                "Non-positive bootstrap standard error encountered."
            )

        return beta_star[:, 0] / bootstrap_se_target

    t_bootstrap_classic = bootstrap_t_statistics(scores_r2)
    t_bootstrap_score = bootstrap_t_statistics(scores_r2_s)

    # Symmetric two-sided bootstrap p-value:
    # mean(|t*| > |t_observed|)
    p_wclr_c = float(
        np.mean(
            np.abs(t_bootstrap_classic) > abs(t_observed)
        )
    )

    p_wclr_s = float(
        np.mean(
            np.abs(t_bootstrap_score) > abs(t_observed)
        )
    )

    # Equal-tail p-value는 검증·감사용으로 함께 저장
    def equal_tail_p_value(t_bootstrap: np.ndarray) -> float:
        upper_tail = np.mean(t_bootstrap > t_observed)
        lower_tail = np.mean(t_bootstrap <= t_observed)

        return float(
            min(1.0, 2.0 * min(upper_tail, lower_tail))
        )

    return {
        "N": n_observations,
        "Countries": n_clusters,
        "Bootstrap replications": reps,
        "Bootstrap weights": "Rademacher",
        "Public-sector beta": float(beta_hat[0]),
        "OR": float(math.exp(beta_hat[0])),
        "Table 2 cluster-robust SE": float(
            full_cluster_result.bse[0]
        ),
        "Table 2 cluster-robust z p": float(
            full_cluster_result.pvalues[0]
        ),
        "Conventional CV1 SE": se_cv1,
        "Conventional CV1 t": t_observed,
        "Conventional CV1 t(df=G-1) p": p_cv1_t,
        "WCLR-C symmetric p": p_wclr_c,
        "WCLR-S symmetric p": p_wclr_s,
        "WCLR-C equal-tail p": equal_tail_p_value(
            t_bootstrap_classic
        ),
        "WCLR-S equal-tail p": equal_tail_p_value(
            t_bootstrap_score
        ),
        "Full logit converged": bool(
            full_result.mle_retvals.get("converged", False)
        ),
        "Restricted logit converged": bool(
            restricted_result.mle_retvals.get("converged", False)
        ),
    }


# ------------------------------------------------------------
# 7. Run
# ------------------------------------------------------------

def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            "전처리 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {INPUT_PATH}"
        )

    raw_df = pd.read_parquet(INPUT_PATH)

    missing_columns = sorted(
        set(REQUIRED_COLUMNS) - set(raw_df.columns)
    )

    if missing_columns:
        raise KeyError(
            "clean_data.parquet에 필요한 변수가 없습니다:\n"
            + "\n".join(
                f"- {column}"
                for column in missing_columns
            )
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    model_df = prepare_common_sample(raw_df)

    n_countries = model_df["country_fe"].nunique()

    if n_countries != 27:
        raise ValueError(
            "Expected 27 country clusters for this analysis, "
            f"but found {n_countries}."
        )

    print("=" * 76)
    print("Restricted wild-cluster linearized bootstrap 시작")
    print(f"공통 분석표본 N: {len(model_df):,}")
    print(f"국가 클러스터 수 G: {n_countries}")
    print(f"Bootstrap replications: {BOOTSTRAP_REPS:,}")
    print("=" * 76)

    result_rows = []
    design_diagnostics = []

    for outcome, outcome_label in OUTCOMES.items():
        print(f"\n분석 중: {outcome}")

        y, x, clusters, design_columns = build_reordered_design(
            outcome=outcome,
            model_df=model_df,
        )

        if design_columns[0] != TARGET_PARAMETER:
            raise AssertionError(
                "public_sector must be the first design-matrix column."
            )

        result = restricted_wclr_p_values(
            y=y,
            x=x,
            clusters=clusters,
            reps=BOOTSTRAP_REPS,
            seed=RANDOM_SEED,
        )

        result_rows.append(
            {
                "Outcome": outcome,
                "Outcome label": outcome_label,
                **result,
                "WCLR-C p label": p_value_label(
                    result["WCLR-C symmetric p"]
                ),
                "WCLR-C significance": significance_mark(
                    result["WCLR-C symmetric p"]
                ),
                "WCLR-S p label": p_value_label(
                    result["WCLR-S symmetric p"]
                ),
                "WCLR-S significance": significance_mark(
                    result["WCLR-S symmetric p"]
                ),
            }
        )

        design_diagnostics.append(
            {
                "Outcome": outcome,
                "N": len(y),
                "Number of model parameters": x.shape[1],
                "First / target column": design_columns[0],
                "Design columns": " | ".join(design_columns),
            }
        )

    results_df = pd.DataFrame(result_rows)
    diagnostics_df = pd.DataFrame(design_diagnostics)

    appendix_table = results_df[
        [
            "Outcome label",
            "Public-sector beta",
            "OR",
            "Table 2 cluster-robust z p",
            "WCLR-C symmetric p",
            "WCLR-S symmetric p",
            "N",
            "Countries",
            "Bootstrap replications",
            "Bootstrap weights",
        ]
    ].copy()

    appendix_table = appendix_table.rename(
        columns={
            "Outcome label": "Outcome",
            "Table 2 cluster-robust z p": (
                "Conventional cluster-robust z p"
            ),
            "WCLR-C symmetric p": "WCLR-C p",
            "WCLR-S symmetric p": "WCLR-S p",
        }
    )

    notes_df = pd.DataFrame(
        [
            {
                "Item": "Primary bootstrap inference",
                "Detail": (
                    "Use WCLR-S symmetric p-value as the primary "
                    "wild-cluster-bootstrap sensitivity result."
                ),
            },
            {
                "Item": "Supplementary bootstrap inference",
                "Detail": (
                    "Report WCLR-C symmetric p-value alongside WCLR-S "
                    "in the appendix for transparency."
                ),
            },
            {
                "Item": "Null hypothesis",
                "Detail": (
                    "H0: public_sector coefficient = 0 in each "
                    "adjusted country-FE logit model."
                ),
            },
            {
                "Item": "Model specification",
                "Detail": (
                    "Country FE, age, gender, education age, occupation, "
                    "workplace size, and digital skill."
                ),
            },
            {
                "Item": "Interpretation",
                "Detail": (
                    "WCLR p-values are a small-cluster inference check; "
                    "they do not change ORs or AAPDs from Table 2."
                ),
            },
        ]
    )

    tables = {
        "01_AppendixTable_A3": appendix_table,
        "02_Full_WCLR_Results": results_df,
        "03_Design_Diagnostics": diagnostics_df,
        "04_Notes": notes_df,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    for sheet_name, dataframe in tables.items():
        save_csv(
            dataframe=dataframe,
            file_name=f"{sheet_name}.csv",
        )

    format_excel(OUTPUT_XLSX)

    print("\n[Appendix Table A3 candidate]")
    print(appendix_table.to_string(index=False))

    print("\n" + "=" * 76)
    print("분석 완료")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"Results folder: {OUTPUT_DIR}")
    print("=" * 76)


if __name__ == "__main__":
    main()
