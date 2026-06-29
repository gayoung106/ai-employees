from __future__ import annotations

from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# 04_model3_sensitivity_diagnostics.py
#
# 목적
# 1) Model 3 완전사례 표본에서 Model 2와 Model 3을 동일 표본으로 비교
# 2) automated_analysis_access Model 3 수렴 경고 원인 진단
#
# 해석 원칙
# - Model 2 = 메인 분석
# - Model 3 = 자동화 감시·성과평가 경험을 추가한 민감도 분석
# ============================================================


# ------------------------------------------------------------
# 0. 경로 설정
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "04_model3_sensitivity_diagnostics"
OUTPUT_XLSX = OUTPUT_DIR / "04_model3_sensitivity_diagnostics.xlsx"


# ------------------------------------------------------------
# 1. 분석 변수
# ------------------------------------------------------------

OUTCOMES = {
    "detailed_explanation": "상세 설명 제공",
    "personal_data_access": "개인정보 접근권 제공",
    "automated_analysis_access": "자동화 분석 결과 접근권 제공",
    "not_informed": "기술 사용 사실 무고지",
}

MODEL_2_CONTROLS = [
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

MODEL_3_EXPOSURES = [
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
]

CATEGORICAL_CONTROLS = [
    "gender",
    "occupation_code",
    "workplace_size",
]

NUMERIC_COLUMNS = [
    "public_sector",
    "age",
    "education_age",
    "digital_skill_job",
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
]

REQUIRED_COLUMNS = [
    "country_fe",
    "public_sector",
    *OUTCOMES.keys(),
    *MODEL_2_CONTROLS,
    *MODEL_3_EXPOSURES,
]


# ------------------------------------------------------------
# 2. 보조 함수
# ------------------------------------------------------------

def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 45))

            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = width

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. Model 3 완전사례 표본 생성
# ------------------------------------------------------------

def prepare_model3_complete_case(
    dataframe: pd.DataFrame,
    outcome: str,
) -> pd.DataFrame:
    """
    Model 3에 필요한 모든 변수가 유효한 표본만 남깁니다.
    이 동일 표본에서:
      - Restricted Model 2
      - Model 3
    을 비교합니다.
    """
    columns = [
        outcome,
        "public_sector",
        "country_fe",
        *MODEL_2_CONTROLS,
        *MODEL_3_EXPOSURES,
    ]

    model_df = dataframe[columns].copy()

    model_df[outcome] = to_float64(model_df[outcome])

    for column in NUMERIC_COLUMNS:
        if column in model_df.columns:
            model_df[column] = to_float64(model_df[column])

    for column in CATEGORICAL_CONTROLS:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    model_df = model_df[
        model_df[outcome].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(x) for x in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
        & model_df["automated_monitoring_exposure"].isin([0.0, 1.0])
        & model_df[
            "automated_performance_management_exposure"
        ].isin([0.0, 1.0])
    ].copy()

    model_df = model_df.dropna().copy()

    # Statsmodels/Patsy와의 호환을 위한 기본 NumPy dtype 강제
    model_df[outcome] = to_int64(model_df[outcome])
    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in [
        "age",
        "education_age",
        "digital_skill_job",
        "automated_monitoring_exposure",
        "automated_performance_management_exposure",
    ]:
        model_df[column] = to_float64(model_df[column])

    for column in CATEGORICAL_CONTROLS:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


def formula_model2(outcome: str) -> str:
    return (
        f"{outcome} ~ public_sector + C(country_fe) + age + C(gender)"
        " + education_age + C(occupation_code) + C(workplace_size)"
        " + digital_skill_job"
    )


def formula_model3(outcome: str) -> str:
    return (
        formula_model2(outcome)
        + " + automated_monitoring_exposure"
        + " + automated_performance_management_exposure"
    )


# ------------------------------------------------------------
# 4. 회귀모형 적합
# ------------------------------------------------------------

def fit_model(model_df: pd.DataFrame, formula: str):
    """
    국가 군집-강건 표준오차를 우선 적용.
    최적화 수렴여부와 경고 메시지를 함께 반환.
    """
    captured_warnings = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.logit(
            formula=formula,
            data=model_df,
            missing="raise",
        )

        try:
            result = model.fit(
                disp=False,
                maxiter=500,
                cov_type="cluster",
                cov_kwds={
                    "groups": model_df["country_fe"].to_numpy(dtype=object),
                },
            )
            covariance_type = "Cluster-robust SE by country"

        except Exception as cluster_error:
            result = model.fit(
                disp=False,
                maxiter=500,
                cov_type="HC3",
            )
            covariance_type = (
                "HC3 robust SE "
                f"(cluster SE failed: {type(cluster_error).__name__})"
            )

        for warning_item in warning_list:
            captured_warnings.append(str(warning_item.message))

    convergence_info = getattr(result, "mle_retvals", {})
    converged = convergence_info.get("converged", np.nan)

    return result, covariance_type, converged, " | ".join(captured_warnings)


def extract_effect(
    result,
    outcome: str,
    outcome_label: str,
    model_name: str,
    model_df: pd.DataFrame,
    covariance_type: str,
    converged,
    warning_text: str,
) -> dict:
    param = "public_sector"

    coefficient = float(result.params[param])
    standard_error = float(result.bse[param])
    p_value = float(result.pvalues[param])

    ci = result.conf_int().loc[param]
    ci_low = float(ci.iloc[0])
    ci_high = float(ci.iloc[1])

    non_public = model_df.copy()
    public = model_df.copy()

    non_public["public_sector"] = 0
    public["public_sector"] = 1

    predicted_non_public = float(result.predict(non_public).mean())
    predicted_public = float(result.predict(public).mean())

    return {
        "결과변수": outcome,
        "결과변수 라벨": outcome_label,
        "모형": model_name,
        "표본 N": int(result.nobs),
        "국가 수": int(model_df["country_fe"].nunique()),
        "수렴 여부": converged,
        "경고 메시지": warning_text,
        "공분산 추정": covariance_type,
        "공공부문 계수(log-odds)": round(coefficient, 4),
        "표준오차": round(standard_error, 4),
        "오즈비(OR)": round(math.exp(coefficient), 3),
        "OR 95% CI 하한": round(math.exp(ci_low), 3),
        "OR 95% CI 상한": round(math.exp(ci_high), 3),
        "p-value": p_value,
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
        "비공공부문 조정예측확률(%)": round(
            predicted_non_public * 100,
            2,
        ),
        "공공부문 조정예측확률(%)": round(
            predicted_public * 100,
            2,
        ),
        "조정확률 차이(%p, 공공-비공공)": round(
            (predicted_public - predicted_non_public) * 100,
            2,
        ),
        "Pseudo R²": round(float(result.prsquared), 4),
        "AIC": round(float(result.aic), 2),
    }


# ------------------------------------------------------------
# 5. 희소 셀 진단
# ------------------------------------------------------------

def make_stratified_diagnostic(
    model_df: pd.DataFrame,
    outcome: str,
    outcome_label: str,
    stratifier: str,
    stratifier_label: str,
) -> pd.DataFrame:
    """
    특정 범주(stratifier)별 결과변수 0/1 빈도를 계산.
    0 또는 1만 존재하는 범주는 완전/준완전 분리 가능성이 있습니다.
    """
    records = []

    for category, group in model_df.groupby(stratifier, dropna=False):
        zero_count = int((group[outcome] == 0).sum())
        one_count = int((group[outcome] == 1).sum())
        total_n = zero_count + one_count

        records.append(
            {
                "결과변수": outcome,
                "결과변수 라벨": outcome_label,
                "진단 변수": stratifier,
                "진단 변수 라벨": stratifier_label,
                "범주": str(category),
                "결과=0 N": zero_count,
                "결과=1 N": one_count,
                "전체 N": total_n,
                "결과=1 비율(%)": round(
                    one_count / total_n * 100,
                    2,
                ) if total_n > 0 else np.nan,
                "희소·분리 경고": (
                    "YES"
                    if zero_count == 0 or one_count == 0
                    else ""
                ),
            }
        )

    return pd.DataFrame(records)


def make_country_by_sector_diagnostic(
    model_df: pd.DataFrame,
    outcome: str,
    outcome_label: str,
) -> pd.DataFrame:
    """
    국가 × 공공/비공공부문별 결과 0/1 빈도를 산출.
    이 표에서 outcome=1 또는 0이 전혀 없는 셀이 많으면
    국가 고정효과 로지스틱 모형의 수렴 문제가 발생할 수 있습니다.
    """
    records = []

    for (country, public_sector), group in model_df.groupby(
        ["country_fe", "public_sector"],
        dropna=False,
    ):
        zero_count = int((group[outcome] == 0).sum())
        one_count = int((group[outcome] == 1).sum())
        total_n = zero_count + one_count

        records.append(
            {
                "결과변수": outcome,
                "결과변수 라벨": outcome_label,
                "국가": str(country),
                "부문": (
                    "공공부문"
                    if int(public_sector) == 1
                    else "비공공부문"
                ),
                "결과=0 N": zero_count,
                "결과=1 N": one_count,
                "전체 N": total_n,
                "결과=1 비율(%)": round(
                    one_count / total_n * 100,
                    2,
                ) if total_n > 0 else np.nan,
                "희소·분리 경고": (
                    "YES"
                    if zero_count == 0 or one_count == 0
                    else ""
                ),
            }
        )

    return pd.DataFrame(records)


# ------------------------------------------------------------
# 6. 데이터 불러오기
# ------------------------------------------------------------

if not INPUT_PATH.exists():
    raise FileNotFoundError(
        "전처리 파일을 찾을 수 없습니다.\n"
        f"확인 경로: {INPUT_PATH}"
    )

df = pd.read_parquet(INPUT_PATH)

missing_columns = sorted(set(REQUIRED_COLUMNS) - set(df.columns))

if missing_columns:
    raise KeyError(
        "clean_data.parquet에 필요한 변수가 없습니다:\n"
        + "\n".join(f"- {column}" for column in missing_columns)
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 72)
print("Model 3 동일표본 민감도 분석 및 수렴 진단 시작")
print(f"전처리 데이터 행 수: {len(df):,}")
print("=" * 72)


# ------------------------------------------------------------
# 7. 실행
# ------------------------------------------------------------

effect_rows = []
diagnostic_tables = []
country_sector_diagnostics = []
sample_rows = []

for outcome, outcome_label in OUTCOMES.items():
    print(f"\n분석 중: {outcome}")

    complete_case_df = prepare_model3_complete_case(
        dataframe=df,
        outcome=outcome,
    )

    if len(complete_case_df) < 100:
        raise ValueError(
            f"{outcome}: Model 3 완전사례 표본이 너무 적습니다 "
            f"(N={len(complete_case_df)})."
        )

    # 동일 표본에서 Restricted Model 2와 Model 3을 비교
    models_to_run = {
        "Restricted Model 2 (same N)": formula_model2(outcome),
        "Model 3": formula_model3(outcome),
    }

    for model_name, formula in models_to_run.items():
        result, covariance_type, converged, warning_text = fit_model(
            model_df=complete_case_df,
            formula=formula,
        )

        effect_rows.append(
            extract_effect(
                result=result,
                outcome=outcome,
                outcome_label=outcome_label,
                model_name=model_name,
                model_df=complete_case_df,
                covariance_type=covariance_type,
                converged=converged,
                warning_text=warning_text,
            )
        )

        sample_rows.append(
            {
                "결과변수": outcome,
                "결과변수 라벨": outcome_label,
                "모형": model_name,
                "표본 N": int(result.nobs),
                "공공부문 N": int(
                    (complete_case_df["public_sector"] == 1).sum()
                ),
                "비공공부문 N": int(
                    (complete_case_df["public_sector"] == 0).sum()
                ),
                "국가 수": int(
                    complete_case_df["country_fe"].nunique()
                ),
                "수렴 여부": converged,
                "경고 메시지": warning_text,
            }
        )

    # 수렴 진단용 표
    for stratifier, stratifier_label in [
        ("country_fe", "국가"),
        ("occupation_code", "직업 유형"),
        ("workplace_size", "조직 규모"),
        (
            "automated_monitoring_exposure",
            "자동화 감시 경험",
        ),
        (
            "automated_performance_management_exposure",
            "자동화 성과평가 경험",
        ),
    ]:
        diagnostic_tables.append(
            make_stratified_diagnostic(
                model_df=complete_case_df,
                outcome=outcome,
                outcome_label=outcome_label,
                stratifier=stratifier,
                stratifier_label=stratifier_label,
            )
        )

    country_sector_diagnostics.append(
        make_country_by_sector_diagnostic(
            model_df=complete_case_df,
            outcome=outcome,
            outcome_label=outcome_label,
        )
    )


# ------------------------------------------------------------
# 8. 결과 저장
# ------------------------------------------------------------

same_sample_effects = pd.DataFrame(effect_rows)
same_sample_details = pd.DataFrame(sample_rows)

diagnostic_table = pd.concat(
    diagnostic_tables,
    ignore_index=True,
)

country_sector_diagnostic_table = pd.concat(
    country_sector_diagnostics,
    ignore_index=True,
)

# 희소·분리 가능성이 있는 셀만 별도 추출
sparse_cells = pd.concat(
    [
        diagnostic_table.loc[
            diagnostic_table["희소·분리 경고"] == "YES"
        ],
        country_sector_diagnostic_table.loc[
            country_sector_diagnostic_table["희소·분리 경고"] == "YES"
        ],
    ],
    ignore_index=True,
    sort=False,
)

tables = {
    "01_동일표본_공공부문효과": same_sample_effects,
    "02_동일표본_모형정보": same_sample_details,
    "03_범주별수렴진단": diagnostic_table,
    "04_국가부문별수렴진단": country_sector_diagnostic_table,
    "05_희소셀_요약": sparse_cells,
}

with pd.ExcelWriter(
    OUTPUT_XLSX,
    engine="openpyxl",
) as writer:
    for sheet_name, table in tables.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

for sheet_name, table in tables.items():
    safe_name = (
        sheet_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )
    save_csv(table, f"{safe_name}.csv")

format_excel(OUTPUT_XLSX)


# ------------------------------------------------------------
# 9. 콘솔 요약
# ------------------------------------------------------------

summary_columns = [
    "결과변수 라벨",
    "모형",
    "표본 N",
    "수렴 여부",
    "오즈비(OR)",
    "OR 95% CI 하한",
    "OR 95% CI 상한",
    "p-value 표기",
    "유의성",
    "비공공부문 조정예측확률(%)",
    "공공부문 조정예측확률(%)",
    "조정확률 차이(%p, 공공-비공공)",
]

print("\n[동일 표본: Restricted Model 2 vs Model 3]")
print(
    same_sample_effects[
        summary_columns
    ].to_string(index=False)
)

print("\n[희소·분리 경고 셀 수]")
print(f"{len(sparse_cells):,}개")

print("\n" + "=" * 72)
print("Model 3 동일표본 민감도 분석 및 수렴 진단 완료")
print(f"엑셀 파일: {OUTPUT_XLSX}")
print(f"결과 폴더: {OUTPUT_DIR}")
print("=" * 72)
