from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# 0. 경로 설정
#
# ai-employees/
# ├─ codes/
# │  ├─ preprocess.py
# │  └─ 01_descriptive_statistics.py
# ├─ clean_data/
# │  └─ clean_data.parquet
# └─ results/
#    └─ 01_descriptive_statistics/
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "01_descriptive_statistics"
OUTPUT_XLSX = OUTPUT_DIR / "01_descriptive_statistics.xlsx"


# ============================================================
# 1. 변수 라벨
# ============================================================

VARIABLE_LABELS = {
    "public_sector": "공공부문 여부",
    "sector_code": "산업·부문",
    "detailed_explanation": "상세 설명 제공",
    "personal_data_access": "개인정보 접근권 제공",
    "automated_analysis_access": "자동화 분석 결과 접근권 제공",
    "not_informed": "기술 사용 사실 무고지",
    "basic_notification_only": "단순 고지",
    "automated_monitoring_exposure": "자동화 감시 경험",
    "automated_performance_management_exposure": "자동화 성과평가 경험",
    "age": "연령",
    "gender": "성별",
    "education_age": "전일제 교육 종료 연령",
    "occupation_code": "직업 유형",
    "workplace_size": "조직 규모",
    "digital_skill_job": "직무 수행 디지털 역량",
}

GENDER_LABELS = {
    1: "Man",
    2: "Woman",
    3: "None of the above / Non-binary / Prefer not to say",
}

OCCUPATION_LABELS = {
    10: "Employed professional",
    11: "General management / director / top management",
    12: "Middle management / other management",
    13: "Working mainly at a desk",
    14: "Not at a desk but travelling",
    15: "Service job",
    16: "Supervisor",
    17: "Skilled manual worker",
    18: "Other employed manual worker",
}

WORKPLACE_SIZE_LABELS = {
    1: "1",
    2: "2-9",
    3: "10-49",
    4: "50-250",
    5: "More than 250",
}

SECTOR_LABELS = {
    1: "Agriculture, forestry and fishing",
    2: "Manufacturing",
    3: "Logistics",
    4: "Service",
    5: "Public sector",
}


# ============================================================
# 2. 보조 함수
# ============================================================

def to_numeric_series(series: pd.Series) -> pd.Series:
    """숫자로 변환할 수 없는 값은 결측치로 처리합니다."""
    return pd.to_numeric(series, errors="coerce")


def pct_yes(series: pd.Series) -> float:
    """이항변수에서 값이 1인 비율을 계산합니다."""
    numeric_series = to_numeric_series(series).dropna()

    if numeric_series.empty:
        return np.nan

    return float((numeric_series == 1).mean() * 100)


def calculate_binary_summary(
    data: pd.DataFrame,
    variable: str,
    group_name: str,
) -> dict:
    """이항 변수의 빈도와 비율을 계산합니다."""
    series = to_numeric_series(data[variable]).dropna()

    valid_n = len(series)
    yes_n = int((series == 1).sum())
    yes_pct = (yes_n / valid_n * 100) if valid_n > 0 else np.nan

    return {
        "집단": group_name,
        "변수명": variable,
        "변수 라벨": VARIABLE_LABELS[variable],
        "유효 N": valid_n,
        "예(1) N": yes_n,
        "예(1) 비율(%)": round(yes_pct, 2),
    }


def calculate_continuous_summary(
    data: pd.DataFrame,
    variable: str,
    group_name: str,
) -> dict:
    """연속형 변수의 평균, 표준편차, 중앙값 등을 계산합니다."""
    series = to_numeric_series(data[variable]).dropna()

    if series.empty:
        return {
            "집단": group_name,
            "변수명": variable,
            "변수 라벨": VARIABLE_LABELS[variable],
            "유효 N": 0,
            "평균": np.nan,
            "표준편차": np.nan,
            "중앙값": np.nan,
            "최소값": np.nan,
            "최대값": np.nan,
        }

    return {
        "집단": group_name,
        "변수명": variable,
        "변수 라벨": VARIABLE_LABELS[variable],
        "유효 N": len(series),
        "평균": round(float(series.mean()), 2),
        "표준편차": round(float(series.std(ddof=1)), 2),
        "중앙값": round(float(series.median()), 2),
        "최소값": round(float(series.min()), 2),
        "최대값": round(float(series.max()), 2),
    }


def calculate_categorical_summary(
    data: pd.DataFrame,
    variable: str,
    labels: dict[int, str],
    group_name: str,
) -> list[dict]:
    """범주형 변수의 빈도와 비율을 계산합니다."""
    series = to_numeric_series(data[variable])
    valid_n = int(series.notna().sum())

    rows = []

    for code, category_name in labels.items():
        category_n = int((series == code).sum())
        category_pct = (
            category_n / valid_n * 100
            if valid_n > 0
            else np.nan
        )

        rows.append(
            {
                "집단": group_name,
                "변수명": variable,
                "변수 라벨": VARIABLE_LABELS[variable],
                "코드": code,
                "범주": category_name,
                "N": category_n,
                "비율(%)": round(category_pct, 2),
            }
        )

    return rows


def save_csv(dataframe: pd.DataFrame, filename: str) -> None:
    """한글이 깨지지 않도록 UTF-8-SIG 형식으로 CSV를 저장합니다."""
    dataframe.to_csv(
        OUTPUT_DIR / filename,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel_file(excel_path: Path) -> None:
    """생성된 엑셀 파일의 기본 서식을 적용합니다."""
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)

            max_length = 0

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                45,
            )

    workbook.save(excel_path)


# ============================================================
# 3. 데이터 불러오기
# ============================================================

if not INPUT_PATH.exists():
    raise FileNotFoundError(
        "전처리 파일을 찾을 수 없습니다.\n"
        f"확인 경로: {INPUT_PATH}"
    )

df = pd.read_parquet(INPUT_PATH)

required_columns = [
    "respondent_id",
    "country_fe",
    "public_sector",
    "sector_code",
    "detailed_explanation",
    "personal_data_access",
    "automated_analysis_access",
    "not_informed",
    "basic_notification_only",
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

missing_columns = sorted(
    set(required_columns) - set(df.columns)
)

if missing_columns:
    raise KeyError(
        "clean_data.parquet에 필요한 변수가 없습니다:\n"
        + "\n".join(f"- {column}" for column in missing_columns)
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 국가 식별값의 결측을 별도 범주로 처리
df["country_fe"] = (
    df["country_fe"]
    .astype("string")
    .fillna("Missing")
)

print("=" * 70)
print("기술통계 분석 시작")
print(f"분석 데이터 행 수: {len(df):,}")
print("=" * 70)


# ============================================================
# 4. 분석 집단 정의
# ============================================================

groups = {
    "전체": df,
    "비공공부문": df.loc[df["public_sector"] == 0].copy(),
    "공공부문": df.loc[df["public_sector"] == 1].copy(),
}


# ============================================================
# 5. 표본 구성
# ============================================================

sample_summary_rows = []

for group_name, group_df in groups.items():
    sample_summary_rows.append(
        {
            "집단": group_name,
            "N": len(group_df),
            "전체 대비 비율(%)": round(
                len(group_df) / len(df) * 100,
                2,
            ),
        }
    )

sample_summary = pd.DataFrame(sample_summary_rows)


# ============================================================
# 6. 결측치 현황
# ============================================================

analysis_variables = [
    "public_sector",
    "detailed_explanation",
    "personal_data_access",
    "automated_analysis_access",
    "not_informed",
    "basic_notification_only",
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

missingness_rows = []

for variable in analysis_variables:
    total_n = len(df)
    missing_n = int(df[variable].isna().sum())
    valid_n = int(df[variable].notna().sum())

    missingness_rows.append(
        {
            "변수명": variable,
            "변수 라벨": VARIABLE_LABELS[variable],
            "전체 N": total_n,
            "결측 N": missing_n,
            "결측 비율(%)": round(
                missing_n / total_n * 100,
                2,
            ),
            "유효 N": valid_n,
        }
    )

missingness_table = pd.DataFrame(missingness_rows)


# ============================================================
# 7. 핵심 결과변수: 전체·부문별 비율
# ============================================================

outcome_variables = [
    "detailed_explanation",
    "personal_data_access",
    "automated_analysis_access",
    "not_informed",
    "basic_notification_only",
]

outcome_rows = []

for group_name, group_df in groups.items():
    for variable in outcome_variables:
        outcome_rows.append(
            calculate_binary_summary(
                data=group_df,
                variable=variable,
                group_name=group_name,
            )
        )

outcome_table = pd.DataFrame(outcome_rows)


# ============================================================
# 8. 보조 변수: 자동화 감시·성과평가 경험
# ============================================================

exposure_variables = [
    "automated_monitoring_exposure",
    "automated_performance_management_exposure",
]

exposure_rows = []

for group_name, group_df in groups.items():
    for variable in exposure_variables:
        exposure_rows.append(
            calculate_binary_summary(
                data=group_df,
                variable=variable,
                group_name=group_name,
            )
        )

exposure_table = pd.DataFrame(exposure_rows)


# ============================================================
# 9. 연속형 통제변수 기술통계
# ============================================================

continuous_variables = [
    "age",
    "education_age",
    "digital_skill_job",
]

continuous_rows = []

for group_name, group_df in groups.items():
    for variable in continuous_variables:
        continuous_rows.append(
            calculate_continuous_summary(
                data=group_df,
                variable=variable,
                group_name=group_name,
            )
        )

continuous_table = pd.DataFrame(continuous_rows)


# ============================================================
# 10. 범주형 통제변수 분포
# ============================================================

categorical_definitions = {
    "gender": GENDER_LABELS,
    "occupation_code": OCCUPATION_LABELS,
    "workplace_size": WORKPLACE_SIZE_LABELS,
}

categorical_rows = []

for group_name, group_df in groups.items():
    for variable, labels in categorical_definitions.items():
        categorical_rows.extend(
            calculate_categorical_summary(
                data=group_df,
                variable=variable,
                labels=labels,
                group_name=group_name,
            )
        )

categorical_table = pd.DataFrame(categorical_rows)


# ============================================================
# 11. 전체 표본의 산업·부문 분포
# ============================================================

sector_distribution_rows = calculate_categorical_summary(
    data=df,
    variable="sector_code",
    labels=SECTOR_LABELS,
    group_name="전체",
)

sector_distribution_table = pd.DataFrame(
    sector_distribution_rows
)


# ============================================================
# 12. 국가별 표본 구성
# ============================================================

country_sample_table = (
    df.groupby("country_fe", dropna=False)
    .agg(
        total_n=("respondent_id", "size"),
        public_sector_n=("public_sector", "sum"),
    )
    .reset_index()
)

country_sample_table["non_public_sector_n"] = (
    country_sample_table["total_n"]
    - country_sample_table["public_sector_n"]
)

country_sample_table["public_sector_pct"] = (
    country_sample_table["public_sector_n"]
    / country_sample_table["total_n"]
    * 100
)

country_sample_table = (
    country_sample_table
    .rename(
        columns={
            "country_fe": "국가",
            "total_n": "전체 N",
            "public_sector_n": "공공부문 N",
            "non_public_sector_n": "비공공부문 N",
            "public_sector_pct": "공공부문 비율(%)",
        }
    )
    .round(2)
    .sort_values(by="국가")
    .reset_index(drop=True)
)


# ============================================================
# 13. 국가별 핵심 결과변수 비율
# ============================================================

country_outcome_table = (
    df.groupby("country_fe", dropna=False)
    .agg(
        n=("respondent_id", "size"),
        detailed_explanation_pct=(
            "detailed_explanation",
            pct_yes,
        ),
        personal_data_access_pct=(
            "personal_data_access",
            pct_yes,
        ),
        automated_analysis_access_pct=(
            "automated_analysis_access",
            pct_yes,
        ),
        not_informed_pct=(
            "not_informed",
            pct_yes,
        ),
    )
    .reset_index()
)

country_outcome_table = (
    country_outcome_table
    .rename(
        columns={
            "country_fe": "국가",
            "n": "N",
            "detailed_explanation_pct": "상세 설명 제공 비율(%)",
            "personal_data_access_pct": "개인정보 접근권 비율(%)",
            "automated_analysis_access_pct": (
                "자동화 분석 결과 접근권 비율(%)"
            ),
            "not_informed_pct": "무고지 비율(%)",
        }
    )
    .round(2)
    .sort_values(by="국가")
    .reset_index(drop=True)
)


# ============================================================
# 14. 엑셀 및 CSV 저장
# ============================================================

tables = {
    "00_표본요약": sample_summary,
    "01_결측치현황": missingness_table,
    "02_핵심결과변수_부문별": outcome_table,
    "03_자동화관리경험_부문별": exposure_table,
    "04_연속형통제변수": continuous_table,
    "05_범주형통제변수": categorical_table,
    "06_산업부문분포": sector_distribution_table,
    "07_국가별표본구성": country_sample_table,
    "08_국가별핵심결과": country_outcome_table,
}

with pd.ExcelWriter(
    OUTPUT_XLSX,
    engine="openpyxl",
) as writer:
    for sheet_name, table in tables.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

for sheet_name, table in tables.items():
    safe_filename = (
        sheet_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )

    save_csv(
        dataframe=table,
        filename=f"{safe_filename}.csv",
    )

format_excel_file(OUTPUT_XLSX)


# ============================================================
# 15. 콘솔 요약
# ============================================================

print("\n[표본 구성]")
print(sample_summary.to_string(index=False))

print("\n[핵심 결과변수: 부문별 비율]")
print(
    outcome_table[
        [
            "집단",
            "변수 라벨",
            "유효 N",
            "예(1) N",
            "예(1) 비율(%)",
        ]
    ].to_string(index=False)
)

print("\n[결측치 현황]")
print(
    missingness_table[
        [
            "변수명",
            "결측 N",
            "결측 비율(%)",
        ]
    ].to_string(index=False)
)

print("\n" + "=" * 70)
print("기술통계 분석 완료")
print(f"엑셀 파일: {OUTPUT_XLSX}")
print(f"결과 폴더: {OUTPUT_DIR}")
print("=" * 70)