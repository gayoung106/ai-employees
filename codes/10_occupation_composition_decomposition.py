from __future__ import annotations

from pathlib import Path
import math
import warnings

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pyreadstat
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 10_occupation_composition_decomposition.py
#
# 목적
#   공공·비공공부문 간 상세 설명 제공률의 차이를 직업 유형 기준으로
#   Shapley형 표준화 분해합니다.
#
# 핵심 질문
#   Model 1 -> Model 2 변화에서 직업 유형이 큰 역할을 한 이유가:
#   (1) 공공·비공공부문의 직업 구성 차이 때문인지
#   (2) 동일 직업 안에서의 공공·비공공부문 차이 때문인지
#   정량적으로 나눠 봅니다.
#
# 중요한 해석 한계
# - 인과적 Oaxaca/매개분석이 아닙니다.
# - '직업 구성 성분'과 '직업 내 부문 연관성 성분'은
#   상호작용 로짓모형에 기초한 표준화 진단값입니다.
# - 직업은 공공부문 고용의 원인·결과·구성요인일 수 있으므로,
#   어떤 성분도 인과효과라고 표현하지 않습니다.
# ============================================================


# ------------------------------------------------------------
# 0. Paths
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
RAW_SAV_PATH = PROJECT_ROOT / "datas" / "raw_data.sav"

OUTPUT_DIR = (
    PROJECT_ROOT
    / "results"
    / "10_occupation_composition_decomposition"
)
OUTPUT_XLSX = OUTPUT_DIR / "10_occupation_composition_decomposition.xlsx"
OUTPUT_PNG = (
    OUTPUT_DIR
    / "Figure_A2_OccupationDecomposition_Contributions.png"
)


# ------------------------------------------------------------
# 1. Analysis definition
# ------------------------------------------------------------

OUTCOME = "detailed_explanation"
OUTCOME_LABEL = "상세 설명 제공"

REQUIRED_COLUMNS = [
    OUTCOME,
    "public_sector",
    "country_fe",
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

INTERACTION_FORMULA = (
    f"{OUTCOME} ~ public_sector * C(occupation_code)"
    " + C(country_fe) + age + C(gender)"
    " + education_age + C(workplace_size)"
    " + digital_skill_job"
)


# ------------------------------------------------------------
# 2. Utilities
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 60))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

        for row_number in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 24

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. Sample / labels
# ------------------------------------------------------------

def prepare_complete_case_sample(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Existing Model 2와 동일한 complete-case sample을 재현합니다.
    """
    model_df = dataframe[REQUIRED_COLUMNS].copy()

    for column in [
        OUTCOME,
        "public_sector",
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    model_df = model_df[
        model_df[OUTCOME].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
    ].copy()

    model_df = model_df.dropna().copy()

    model_df[OUTCOME] = to_int64(model_df[OUTCOME])
    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in ["age", "education_age", "digital_skill_job"]:
        model_df[column] = to_float64(model_df[column])

    for column in ["gender", "occupation_code", "workplace_size"]:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


def load_occupation_labels() -> dict[int, str]:
    fallback = {
        code: f"Occupation code {code}"
        for code in range(10, 19)
    }

    if not RAW_SAV_PATH.exists():
        return fallback

    try:
        _, metadata = pyreadstat.read_sav(
            RAW_SAV_PATH,
            metadataonly=True,
        )

        raw_labels = metadata.variable_value_labels.get("d15a", {})

        if not raw_labels:
            return fallback

        labels = {}

        for code in range(10, 19):
            label = (
                raw_labels.get(float(code))
                or raw_labels.get(code)
                or fallback[code]
            )
            labels[code] = str(label)

        return labels

    except Exception:
        return fallback


# ------------------------------------------------------------
# 4. Interaction model and standardised occupation predictions
# ------------------------------------------------------------

def fit_interaction_model(model_df: pd.DataFrame):
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.glm(
            formula=INTERACTION_FORMULA,
            data=model_df,
            family=sm.families.Binomial(),
            missing="raise",
        )

        result = model.fit(
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    return result, " | ".join(warning_messages)


def occupation_standardised_predictions(
    result,
    model_df: pd.DataFrame,
    occupation_labels: dict[int, str],
) -> pd.DataFrame:
    """
    각 직업 내에서 실제 비부문 공변량 분포를 유지한 채,
    공공부문=0 / 공공부문=1의 표준화 예측확률을 계산합니다.

    즉, 같은 직업 유형 안에서 공공 여부만 반사실적으로 변경합니다.
    """
    rows = []

    total_public = int((model_df["public_sector"] == 1).sum())
    total_nonpublic = int((model_df["public_sector"] == 0).sum())
    total_n = len(model_df)

    for code in sorted(model_df["occupation_code"].unique()):
        occupation_df = model_df.loc[
            model_df["occupation_code"] == code
        ].copy()

        counterfactual_nonpublic = occupation_df.copy()
        counterfactual_public = occupation_df.copy()

        counterfactual_nonpublic["public_sector"] = 0
        counterfactual_public["public_sector"] = 1

        m_nonpublic = float(
            result.predict(counterfactual_nonpublic).mean()
        )
        m_public = float(
            result.predict(counterfactual_public).mean()
        )

        public_n = int(
            (occupation_df["public_sector"] == 1).sum()
        )
        nonpublic_n = int(
            (occupation_df["public_sector"] == 0).sum()
        )

        share_public = public_n / total_public
        share_nonpublic = nonpublic_n / total_nonpublic
        share_pooled = len(occupation_df) / total_n

        rows.append(
            {
                "Occupation code": int(code),
                "Occupation label": occupation_labels[int(code)],
                "Public N": public_n,
                "Non-public N": nonpublic_n,
                "Public occupation share": share_public,
                "Non-public occupation share": share_nonpublic,
                "Pooled occupation share": share_pooled,
                "m_public (standardised probability)": m_public,
                "m_non-public (standardised probability)": m_nonpublic,
                "Within-occupation difference (public - non-public)": (
                    m_public - m_nonpublic
                ),
            }
        )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 5. Shapley decomposition
# ------------------------------------------------------------

def shapley_decomposition(
    occupation_predictions: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    직업 구성 차이와 직업 내 공공부문 차이를 Shapley 대칭 방식으로 분해.

    표기:
    p_Pk: 공공부문 내 직업 k의 실제 비중
    p_Nk: 비공공부문 내 직업 k의 실제 비중
    m_Pk: 직업 k에서 public=1로 표준화한 예측확률
    m_Nk: 직업 k에서 public=0으로 표준화한 예측확률

    Total gap =
      Σ p_Pk * m_Pk - Σ p_Nk * m_Nk

    Composition component =
      1/2 Σ (p_Pk - p_Nk) * (m_Pk + m_Nk)

    Within-occupation sector-association component =
      1/2 Σ (p_Pk + p_Nk) * (m_Pk - m_Nk)

    두 성분의 합은 total gap과 일치합니다.
    """
    decomposition = occupation_predictions.copy()

    p_public = decomposition["Public occupation share"]
    p_nonpublic = decomposition["Non-public occupation share"]
    m_public = decomposition[
        "m_public (standardised probability)"
    ]
    m_nonpublic = decomposition[
        "m_non-public (standardised probability)"
    ]

    decomposition[
        "Composition contribution (probability)"
    ] = (
        0.5
        * (p_public - p_nonpublic)
        * (m_public + m_nonpublic)
    )

    decomposition[
        "Within-occupation sector-association contribution (probability)"
    ] = (
        0.5
        * (p_public + p_nonpublic)
        * (m_public - m_nonpublic)
    )

    decomposition[
        "Total occupation contribution (probability)"
    ] = (
        decomposition[
            "Composition contribution (probability)"
        ]
        + decomposition[
            "Within-occupation sector-association contribution (probability)"
        ]
    )

    decomposition[
        "Composition contribution (pp)"
    ] = (
        decomposition[
            "Composition contribution (probability)"
        ] * 100
    )

    decomposition[
        "Within-occupation sector-association contribution (pp)"
    ] = (
        decomposition[
            "Within-occupation sector-association contribution (probability)"
        ] * 100
    )

    decomposition[
        "Total occupation contribution (pp)"
    ] = (
        decomposition[
            "Total occupation contribution (probability)"
        ] * 100
    )

    overall_public = float((p_public * m_public).sum())
    overall_nonpublic = float((p_nonpublic * m_nonpublic).sum())
    total_gap = overall_public - overall_nonpublic

    composition_component = float(
        decomposition[
            "Composition contribution (probability)"
        ].sum()
    )

    within_occupation_component = float(
        decomposition[
            "Within-occupation sector-association contribution (probability)"
        ].sum()
    )

    check_difference = (
        total_gap
        - composition_component
        - within_occupation_component
    )

    summary = pd.DataFrame(
        [
            {
                "Metric": "Standardised public-sector predicted probability",
                "Probability": overall_public,
                "Percentage points": overall_public * 100,
            },
            {
                "Metric": "Standardised non-public-sector predicted probability",
                "Probability": overall_nonpublic,
                "Percentage points": overall_nonpublic * 100,
            },
            {
                "Metric": "Total standardised gap (public - non-public)",
                "Probability": total_gap,
                "Percentage points": total_gap * 100,
            },
            {
                "Metric": "Occupation-composition component",
                "Probability": composition_component,
                "Percentage points": composition_component * 100,
            },
            {
                "Metric": "Within-occupation sector-association component",
                "Probability": within_occupation_component,
                "Percentage points": (
                    within_occupation_component * 100
                ),
            },
            {
                "Metric": "Decomposition check (should be ~0)",
                "Probability": check_difference,
                "Percentage points": check_difference * 100,
            },
        ]
    )

    return decomposition, summary


# ------------------------------------------------------------
# 6. Figure
# ------------------------------------------------------------

def create_contribution_plot(
    decomposition: pd.DataFrame,
) -> None:
    """
    직업별 total contribution을 표시.
    양수 = 공공부문이 높은 방향에 기여,
    음수 = 공공부문이 낮은 방향에 기여.
    """
    plot_df = decomposition.sort_values(
        "Total occupation contribution (pp)"
    ).reset_index(drop=True)

    labels = [
        (
            f"{code}: {label}"
            if len(label) <= 42
            else f"{code}: {label[:39]}..."
        )
        for code, label in zip(
            plot_df["Occupation code"],
            plot_df["Occupation label"],
        )
    ]

    y_positions = np.arange(len(plot_df))
    values = plot_df[
        "Total occupation contribution (pp)"
    ].to_numpy(dtype=float)

    plt.figure(figsize=(11, 7))
    plt.barh(y_positions, values)
    plt.axvline(x=0, linewidth=1)
    plt.yticks(y_positions, labels)
    plt.xlabel(
        "Contribution to standardised public − non-public gap "
        "(percentage points)"
    )
    plt.ylabel("Occupation category")
    plt.title(
        "Occupation-Specific Contributions to the Detailed-Explanation Gap"
    )
    plt.tight_layout()
    plt.savefig(
        OUTPUT_PNG,
        dpi=300,
        bbox_inches="tight",
    )
    plt.close()


# ------------------------------------------------------------
# 7. Interpretation guide
# ------------------------------------------------------------

def build_interpretation_guide() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Item": "Total standardised gap",
                "Interpretation": (
                    "Difference between public and non-public predicted "
                    "probabilities using each sector's observed occupation "
                    "distribution and occupation-specific standardised predictions."
                ),
            },
            {
                "Item": "Occupation-composition component",
                "Interpretation": (
                    "Portion of the gap associated with public and non-public "
                    "workers being differently distributed across occupations. "
                    "This is not a causal effect of occupation."
                ),
            },
            {
                "Item": "Within-occupation sector-association component",
                "Interpretation": (
                    "Portion of the gap associated with different predicted "
                    "public/non-public probabilities within occupation categories."
                ),
            },
            {
                "Item": "Positive component",
                "Interpretation": (
                    "Contributes to a higher predicted detailed-explanation "
                    "probability in the public sector."
                ),
            },
            {
                "Item": "Negative component",
                "Interpretation": (
                    "Contributes to a lower predicted detailed-explanation "
                    "probability in the public sector."
                ),
            },
            {
                "Item": "Manuscript language",
                "Interpretation": (
                    "Use 'composition-related component' and "
                    "'within-occupation conditional association component'; "
                    "do not use 'mediated effect' or 'causal mechanism'."
                ),
            },
        ]
    )


# ------------------------------------------------------------
# 8. Run
# ------------------------------------------------------------

def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            "전처리 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {INPUT_PATH}"
        )

    raw_df = pd.read_parquet(INPUT_PATH)

    missing_columns = sorted(
        set(REQUIRED_COLUMNS) - set(raw_df.columns)
    )

    if missing_columns:
        raise KeyError(
            "clean_data.parquet에 필요한 변수가 없습니다:\n"
            + "\n".join(
                f"- {column}"
                for column in missing_columns
            )
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    model_df = prepare_complete_case_sample(raw_df)
    occupation_labels = load_occupation_labels()

    print("=" * 78)
    print("직업구성-직업내 부문연관성 Shapley 분해 시작")
    print(f"분석표본 N: {len(model_df):,}")
    print(f"국가 수: {model_df['country_fe'].nunique()}")
    print("=" * 78)

    print("직업유형 상호작용 로지스틱 모형 추정")
    interaction_result, warning_text = fit_interaction_model(
        model_df
    )

    if warning_text:
        print(f"경고: {warning_text}")

    print("직업별 표준화 예측확률 산출")
    occupation_predictions = occupation_standardised_predictions(
        result=interaction_result,
        model_df=model_df,
        occupation_labels=occupation_labels,
    )

    print("Shapley 분해 산출")
    decomposition_table, decomposition_summary = (
        shapley_decomposition(occupation_predictions)
    )

    create_contribution_plot(decomposition_table)

    guide_table = build_interpretation_guide()

    tables = {
        "01_OccupationPredictions": occupation_predictions,
        "02_DecompositionByOccupation": decomposition_table,
        "03_DecompositionSummary": decomposition_summary,
        "04_InterpretationGuide": guide_table,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    for sheet_name, dataframe in tables.items():
        save_csv(
            dataframe=dataframe,
            file_name=f"{sheet_name}.csv",
        )

    format_excel(OUTPUT_XLSX)

    print("\n[분해 요약]")
    print(
        decomposition_summary.to_string(
            index=False,
            formatters={
                "Probability": "{:.6f}".format,
                "Percentage points": "{:.3f}".format,
            },
        )
    )

    print("\n[직업별 기여: 절대값 큰 순]")
    display_columns = [
        "Occupation code",
        "Occupation label",
        "Public occupation share",
        "Non-public occupation share",
        "Within-occupation difference (public - non-public)",
        "Composition contribution (pp)",
        "Within-occupation sector-association contribution (pp)",
        "Total occupation contribution (pp)",
    ]

    display_table = decomposition_table.copy()
    display_table["Absolute contribution"] = display_table[
        "Total occupation contribution (pp)"
    ].abs()

    display_table = display_table.sort_values(
        "Absolute contribution",
        ascending=False,
    )

    print(
        display_table[display_columns].to_string(
            index=False,
            formatters={
                "Public occupation share": "{:.3f}".format,
                "Non-public occupation share": "{:.3f}".format,
                "Within-occupation difference (public - non-public)": (
                    "{:.4f}".format
                ),
                "Composition contribution (pp)": "{:.3f}".format,
                "Within-occupation sector-association contribution (pp)": (
                    "{:.3f}".format
                ),
                "Total occupation contribution (pp)": "{:.3f}".format,
            },
        )
    )

    print("\n" + "=" * 78)
    print("분석 완료")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"Figure: {OUTPUT_PNG}")
    print("=" * 78)


if __name__ == "__main__":
    main()
