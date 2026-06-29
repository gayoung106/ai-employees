from __future__ import annotations

from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# 05_loco_detailed_explanation.py
#
# Leave-one-country-out robustness check
# Primary outcome only: detailed_explanation
#
# This script deliberately does NOT run the rare outcomes
# (especially automated_analysis_access), which can produce
# singular-matrix / convergence issues in repeated country-drop models.
# ============================================================


# ------------------------------------------------------------
# 0. Paths
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "05_loco_detailed_explanation"
OUTPUT_XLSX = OUTPUT_DIR / "05_loco_detailed_explanation.xlsx"


# ------------------------------------------------------------
# 1. Analysis variables
# ------------------------------------------------------------

OUTCOME = "detailed_explanation"
OUTCOME_LABEL = "상세 설명 제공"

REQUIRED_COLUMNS = [
    OUTCOME,
    "public_sector",
    "country_fe",
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]


# ------------------------------------------------------------
# 2. Utilities
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            max_width = 12
            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                max_width = max(max_width, min(len(text) + 2, 45))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = max_width

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. Data preparation
# ------------------------------------------------------------

def prepare_model2_data(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Model 2 complete-case sample:
    detailed_explanation ~ public_sector + country FE + controls
    """
    columns = [
        OUTCOME,
        "public_sector",
        "country_fe",
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]

    model_df = dataframe[columns].copy()

    # Convert numeric/nullable columns safely
    for column in [
        OUTCOME,
        "public_sector",
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    # Valid response/category ranges
    model_df = model_df[
        model_df[OUTCOME].isin([0.0, 1.0])
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
    ].copy()

    model_df = model_df.dropna().copy()

    # Convert to native NumPy dtypes required by Patsy/statsmodels
    model_df[OUTCOME] = to_int64(model_df[OUTCOME])
    model_df["public_sector"] = to_int64(model_df["public_sector"])
    model_df["age"] = to_float64(model_df["age"])
    model_df["education_age"] = to_float64(model_df["education_age"])
    model_df["digital_skill_job"] = to_float64(
        model_df["digital_skill_job"]
    )

    for column in ["gender", "occupation_code", "workplace_size"]:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


def model2_formula() -> str:
    return (
        "detailed_explanation ~ public_sector + C(country_fe)"
        " + age + C(gender) + education_age"
        " + C(occupation_code) + C(workplace_size)"
        " + digital_skill_job"
    )


# ------------------------------------------------------------
# 4. Estimation
# ------------------------------------------------------------

def fit_model(model_df: pd.DataFrame, formula: str):
    """
    Fit Model 2 with country-clustered standard errors.
    If a specific leave-one-country-out run cannot be estimated,
    the caller records it as skipped rather than terminating the script.
    """
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.logit(
            formula=formula,
            data=model_df,
            missing="raise",
        )

        result = model.fit(
            disp=False,
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    converged = getattr(result, "mle_retvals", {}).get(
        "converged",
        np.nan,
    )

    return result, converged, " | ".join(warning_messages)


def extract_effect(
    result,
    excluded_country: str,
    model_df: pd.DataFrame,
    converged,
    warning_text: str,
) -> dict:
    parameter = "public_sector"

    coefficient = float(result.params[parameter])
    standard_error = float(result.bse[parameter])
    p_value = float(result.pvalues[parameter])

    confidence_interval = result.conf_int().loc[parameter]
    ci_low = float(confidence_interval.iloc[0])
    ci_high = float(confidence_interval.iloc[1])

    return {
        "제외 국가": excluded_country,
        "표본 N": int(result.nobs),
        "포함 국가 수": int(model_df["country_fe"].nunique()),
        "수렴 여부": converged,
        "경고 메시지": warning_text,
        "공공부문 계수(log-odds)": round(coefficient, 4),
        "표준오차": round(standard_error, 4),
        "오즈비(OR)": round(math.exp(coefficient), 3),
        "OR 95% CI 하한": round(math.exp(ci_low), 3),
        "OR 95% CI 상한": round(math.exp(ci_high), 3),
        "p-value": p_value,
        "p-value 표기": p_value_label(p_value),
        "유의성": significance_mark(p_value),
    }


# ------------------------------------------------------------
# 5. Load data
# ------------------------------------------------------------

if not INPUT_PATH.exists():
    raise FileNotFoundError(
        "전처리 파일을 찾을 수 없습니다.\n"
        f"확인 경로: {INPUT_PATH}"
    )

df = pd.read_parquet(INPUT_PATH)

missing_columns = sorted(set(REQUIRED_COLUMNS) - set(df.columns))

if missing_columns:
    raise KeyError(
        "clean_data.parquet에 필요한 변수가 없습니다:\n"
        + "\n".join(f"- {column}" for column in missing_columns)
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 72)
print("상세 설명 제공: Leave-one-country-out(Model 2) 강건성 분석 시작")
print(f"전처리 데이터 행 수: {len(df):,}")
print("=" * 72)


# ------------------------------------------------------------
# 6. Full model and leave-one-country-out estimation
# ------------------------------------------------------------

analysis_df = prepare_model2_data(df)
formula = model2_formula()
countries = sorted(analysis_df["country_fe"].unique().tolist())

result_rows = []
skipped_rows = []

# Full Model 2
try:
    full_result, full_converged, full_warning_text = fit_model(
        analysis_df,
        formula,
    )

    result_rows.append(
        extract_effect(
            result=full_result,
            excluded_country="None (full Model 2)",
            model_df=analysis_df,
            converged=full_converged,
            warning_text=full_warning_text,
        )
    )
except Exception as error:
    raise RuntimeError(
        "기준 Model 2 추정에 실패했습니다.\n"
        f"원인: {type(error).__name__}: {error}"
    ) from error

# Country-by-country omission
for country in countries:
    print(f"국가 제외 분석 중: {country}")

    leave_out_df = analysis_df.loc[
        analysis_df["country_fe"] != country
    ].copy()

    # Ensure both sector groups remain
    if leave_out_df["public_sector"].nunique() < 2:
        skipped_rows.append(
            {
                "제외 국가": country,
                "상태": "SKIPPED",
                "사유": "제외 후 공공/비공공부문 중 한 집단이 없어 추정 불가",
            }
        )
        continue

    try:
        result, converged, warning_text = fit_model(
            leave_out_df,
            formula,
        )

        result_rows.append(
            extract_effect(
                result=result,
                excluded_country=country,
                model_df=leave_out_df,
                converged=converged,
                warning_text=warning_text,
            )
        )
    except Exception as error:
        # A singular matrix on one omission does not invalidate the
        # full Model 2. Record the failed replicate and continue.
        skipped_rows.append(
            {
                "제외 국가": country,
                "상태": "SKIPPED",
                "사유": f"{type(error).__name__}: {error}",
            }
        )


# ------------------------------------------------------------
# 7. Summary
# ------------------------------------------------------------

loco_results = pd.DataFrame(result_rows)

if skipped_rows:
    skipped_table = pd.DataFrame(skipped_rows)
else:
    skipped_table = pd.DataFrame(
        columns=["제외 국가", "상태", "사유"]
    )

full_row = loco_results.loc[
    loco_results["제외 국가"] == "None (full Model 2)"
].iloc[0]

replicates = loco_results.loc[
    loco_results["제외 국가"] != "None (full Model 2)"
].copy()

summary_table = pd.DataFrame(
    [
        {
            "기준 Model 2 OR": full_row["오즈비(OR)"],
            "기준 Model 2 p-value": full_row["p-value"],
            "성공한 국가 제외 반복 수": len(replicates),
            "제외 분석 실패/건너뜀 수": len(skipped_table),
            "반복 OR 최솟값": round(
                float(replicates["오즈비(OR)"].min()),
                3,
            ),
            "반복 OR 최댓값": round(
                float(replicates["오즈비(OR)"].max()),
                3,
            ),
            "OR < 1 반복 수": int(
                (replicates["오즈비(OR)"] < 1).sum()
            ),
            "p < .05 반복 수": int(
                (replicates["p-value"] < 0.05).sum()
            ),
            "수렴 실패 반복 수": int(
                (replicates["수렴 여부"] != True).sum()
            ),
        }
    ]
)

country_sample_table = (
    analysis_df.groupby("country_fe")
    .agg(
        N=("public_sector", "size"),
        공공부문_N=("public_sector", lambda x: int((x == 1).sum())),
        비공공부문_N=("public_sector", lambda x: int((x == 0).sum())),
        상세설명_예_N=(OUTCOME, lambda x: int((x == 1).sum())),
    )
    .reset_index()
    .rename(columns={"country_fe": "국가"})
)

country_sample_table["상세설명_예_비율(%)"] = (
    country_sample_table["상세설명_예_N"]
    / country_sample_table["N"]
    * 100
).round(2)


# ------------------------------------------------------------
# 8. Save files
# ------------------------------------------------------------

tables = {
    "01_LOCO_전체결과": loco_results,
    "02_LOCO_요약": summary_table,
    "03_국가별표본": country_sample_table,
    "04_건너뜀및오류": skipped_table,
}

with pd.ExcelWriter(
    OUTPUT_XLSX,
    engine="openpyxl",
) as writer:
    for sheet_name, table in tables.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

for sheet_name, table in tables.items():
    safe_name = (
        sheet_name
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )
    save_csv(table, f"{safe_name}.csv")

format_excel(OUTPUT_XLSX)


# ------------------------------------------------------------
# 9. Console output
# ------------------------------------------------------------

print("\n[Leave-one-country-out 요약: 상세 설명 제공]")
print(summary_table.to_string(index=False))

print("\n" + "=" * 72)
print("상세 설명 제공: Leave-one-country-out(Model 2) 강건성 분석 완료")
print(f"엑셀 파일: {OUTPUT_XLSX}")
print(f"결과 폴더: {OUTPUT_DIR}")
print("=" * 72)
