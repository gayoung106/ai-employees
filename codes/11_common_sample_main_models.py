from __future__ import annotations

from pathlib import Path
import math
import warnings

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 11_common_sample_main_models.py
#
# 목적
# - RQ1~RQ4의 기준모형과 조정모형을 동일한 완전사례 표본에서 재추정
# - 메인 회귀표에서 결과변수별 표본 N이 섞이지 않도록 확정
#
# 핵심 원칙
# - 네 결과변수 모두 동일한 master complete-case sample을 사용
# - Baseline model: public_sector + country fixed effects
# - Adjusted model: baseline + age + gender + education + occupation
#                   + workplace size + digital skill
# - 국가 단위 cluster-robust SE 적용
# ============================================================


# ------------------------------------------------------------
# 0. Paths
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

INPUT_PATH = PROJECT_ROOT / "clean_data" / "clean_data.parquet"
OUTPUT_DIR = PROJECT_ROOT / "results" / "11_common_sample_main_models"
OUTPUT_XLSX = OUTPUT_DIR / "11_common_sample_main_models.xlsx"


# ------------------------------------------------------------
# 1. Definitions
# ------------------------------------------------------------

OUTCOMES = {
    "detailed_explanation": "상세 설명 제공 (RQ1)",
    "personal_data_access": "개인정보 접근 경험 (RQ2)",
    "automated_analysis_access": "자동화 분석 결과 접근 경험 (RQ3)",
    "not_informed": "기술 사용 사실 무고지 경험 (RQ4)",
}

CONTROL_COLUMNS = [
    "age",
    "gender",
    "education_age",
    "occupation_code",
    "workplace_size",
    "digital_skill_job",
]

CATEGORICAL_CONTROL_COLUMNS = [
    "gender",
    "occupation_code",
    "workplace_size",
]

REQUIRED_COLUMNS = [
    "public_sector",
    "country_fe",
    *OUTCOMES.keys(),
    *CONTROL_COLUMNS,
]


# ------------------------------------------------------------
# 2. Utilities
# ------------------------------------------------------------

def to_float64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return pd.Series(
        values.to_numpy(dtype="float64", na_value=np.nan),
        index=series.index,
        dtype="float64",
    )


def to_int64(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="raise")
    return pd.Series(
        values.to_numpy(dtype="int64"),
        index=series.index,
        dtype="int64",
    )


def to_object_string(series: pd.Series) -> pd.Series:
    values = series.fillna("Missing").astype(str).to_numpy(dtype=object)
    return pd.Series(values, index=series.index, dtype=object)


def p_value_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "< .001"
    return f"{p_value:.3f}"


def significance_mark(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def format_or_ci(or_value: float, ci_low: float, ci_high: float) -> str:
    return f"{or_value:.3f} [{ci_low:.3f}, {ci_high:.3f}]"


def save_csv(dataframe: pd.DataFrame, file_name: str) -> None:
    dataframe.to_csv(
        OUTPUT_DIR / file_name,
        index=False,
        encoding="utf-8-sig",
    )


def format_excel(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            width = 12

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(text) + 2, 55))

            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = width

        for row_number in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 24

    workbook.save(excel_path)


# ------------------------------------------------------------
# 3. Common complete-case sample
# ------------------------------------------------------------

def prepare_common_sample(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    RQ1-RQ4 모든 결과변수와 최종 통제변수가 유효한 공통 표본을 만듭니다.
    이 표본을 기준/조정모형 모두에 사용합니다.
    """
    model_df = dataframe[REQUIRED_COLUMNS].copy()

    numeric_columns = [
        "public_sector",
        *OUTCOMES.keys(),
        "age",
        "gender",
        "education_age",
        "occupation_code",
        "workplace_size",
        "digital_skill_job",
    ]

    for column in numeric_columns:
        model_df[column] = to_float64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    valid_outcome_mask = np.ones(len(model_df), dtype=bool)

    for outcome in OUTCOMES:
        valid_outcome_mask &= model_df[outcome].isin([0.0, 1.0])

    model_df = model_df[
        valid_outcome_mask
        & model_df["public_sector"].isin([0.0, 1.0])
        & model_df["gender"].isin([1.0, 2.0, 3.0])
        & model_df["occupation_code"].isin(
            [float(code) for code in range(10, 19)]
        )
        & model_df["workplace_size"].isin([1.0, 2.0, 3.0, 4.0, 5.0])
    ].copy()

    model_df = model_df.dropna().copy()

    for outcome in OUTCOMES:
        model_df[outcome] = to_int64(model_df[outcome])

    model_df["public_sector"] = to_int64(model_df["public_sector"])

    for column in ["age", "education_age", "digital_skill_job"]:
        model_df[column] = to_float64(model_df[column])

    for column in CATEGORICAL_CONTROL_COLUMNS:
        model_df[column] = to_int64(model_df[column])

    model_df["country_fe"] = to_object_string(model_df["country_fe"])

    return model_df


# ------------------------------------------------------------
# 4. Estimation
# ------------------------------------------------------------

def baseline_formula(outcome: str) -> str:
    return f"{outcome} ~ public_sector + C(country_fe)"


def adjusted_formula(outcome: str) -> str:
    return (
        f"{outcome} ~ public_sector + C(country_fe)"
        " + age + C(gender) + education_age"
        " + C(occupation_code) + C(workplace_size)"
        " + digital_skill_job"
    )


def fit_logit(
    formula: str,
    model_df: pd.DataFrame,
):
    warning_messages = []

    with warnings.catch_warnings(record=True) as warning_list:
        warnings.simplefilter("always")

        model = smf.logit(
            formula=formula,
            data=model_df,
            missing="raise",
        )

        result = model.fit(
            disp=False,
            maxiter=500,
            cov_type="cluster",
            cov_kwds={
                "groups": model_df["country_fe"].to_numpy(dtype=object),
            },
        )

        for warning_item in warning_list:
            warning_messages.append(str(warning_item.message))

    converged = getattr(result, "mle_retvals", {}).get(
        "converged",
        np.nan,
    )

    return result, converged, " | ".join(warning_messages)


def extract_public_sector_effect(
    result,
    outcome: str,
    outcome_label: str,
    specification: str,
    model_df: pd.DataFrame,
    converged,
    warning_text: str,
) -> dict:
    parameter = "public_sector"

    beta = float(result.params[parameter])
    se = float(result.bse[parameter])
    p_value = float(result.pvalues[parameter])

    ci = result.conf_int().loc[parameter]
    ci_low = float(ci.iloc[0])
    ci_high = float(ci.iloc[1])

    counterfactual_nonpublic = model_df.copy()
    counterfactual_public = model_df.copy()

    counterfactual_nonpublic["public_sector"] = 0
    counterfactual_public["public_sector"] = 1

    predicted_nonpublic = float(
        result.predict(counterfactual_nonpublic).mean()
    )
    predicted_public = float(
        result.predict(counterfactual_public).mean()
    )

    return {
        "Outcome": outcome,
        "Outcome label": outcome_label,
        "Specification": specification,
        "N": int(result.nobs),
        "Countries": int(model_df["country_fe"].nunique()),
        "Converged": converged,
        "Warning": warning_text,
        "Public-sector beta": round(beta, 5),
        "SE": round(se, 5),
        "OR": round(math.exp(beta), 3),
        "OR 95% CI lower": round(math.exp(ci_low), 3),
        "OR 95% CI upper": round(math.exp(ci_high), 3),
        "OR [95% CI]": format_or_ci(
            math.exp(beta),
            math.exp(ci_low),
            math.exp(ci_high),
        ),
        "p-value": p_value,
        "p-value label": p_value_label(p_value),
        "Significance": significance_mark(p_value),
        "Adjusted Pr(non-public, %)": round(
            predicted_nonpublic * 100,
            2,
        ),
        "Adjusted Pr(public, %)": round(
            predicted_public * 100,
            2,
        ),
        "AAPD (public - non-public, pp)": round(
            (predicted_public - predicted_nonpublic) * 100,
            2,
        ),
    }


def extract_all_coefficients(
    result,
    outcome: str,
    specification: str,
) -> pd.DataFrame:
    ci = result.conf_int()
    records = []

    for parameter in result.params.index:
        beta = float(result.params[parameter])
        p_value = float(result.pvalues[parameter])
        ci_low = float(ci.loc[parameter].iloc[0])
        ci_high = float(ci.loc[parameter].iloc[1])

        records.append(
            {
                "Outcome": outcome,
                "Specification": specification,
                "Parameter": parameter,
                "Beta": beta,
                "SE": float(result.bse[parameter]),
                "OR": math.exp(beta),
                "OR 95% CI lower": math.exp(ci_low),
                "OR 95% CI upper": math.exp(ci_high),
                "p-value": p_value,
                "p-value label": p_value_label(p_value),
                "Significance": significance_mark(p_value),
            }
        )

    return pd.DataFrame(records)


# ------------------------------------------------------------
# 5. Manuscript-ready wide table
# ------------------------------------------------------------

def build_main_table(long_results: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for outcome, outcome_label in OUTCOMES.items():
        subset = long_results.loc[
            long_results["Outcome"] == outcome
        ].copy()

        baseline = subset.loc[
            subset["Specification"] == "Baseline (common N)"
        ].iloc[0]

        adjusted = subset.loc[
            subset["Specification"] == "Adjusted (common N)"
        ].iloc[0]

        rows.append(
            {
                "Outcome": outcome_label,
                "Baseline OR [95% CI]": baseline["OR [95% CI]"],
                "Baseline p": (
                    f"{baseline['p-value label']}"
                    f"{baseline['Significance']}"
                ),
                "Baseline AAPD (pp)": (
                    f"{float(baseline['AAPD (public - non-public, pp)']):+.2f}"
                ),
                "Adjusted OR [95% CI]": adjusted["OR [95% CI]"],
                "Adjusted p": (
                    f"{adjusted['p-value label']}"
                    f"{adjusted['Significance']}"
                ),
                "Adjusted AAPD (pp)": (
                    f"{float(adjusted['AAPD (public - non-public, pp)']):+.2f}"
                ),
                "N": int(adjusted["N"]),
                "Country FE": "Yes",
                "Controls in adjusted model": (
                    "Age, gender, education, occupation, "
                    "workplace size, digital skill"
                ),
            }
        )

    return pd.DataFrame(rows)


# ------------------------------------------------------------
# 6. Run
# ------------------------------------------------------------

def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            "전처리 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {INPUT_PATH}"
        )

    raw_df = pd.read_parquet(INPUT_PATH)

    missing_columns = sorted(
        set(REQUIRED_COLUMNS) - set(raw_df.columns)
    )

    if missing_columns:
        raise KeyError(
            "clean_data.parquet에 필요한 변수가 없습니다:\n"
            + "\n".join(f"- {column}" for column in missing_columns)
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    model_df = prepare_common_sample(raw_df)

    print("=" * 76)
    print("RQ1-RQ4 공통 완전사례 표본 메인모형 재추정 시작")
    print(f"공통 분석표본 N: {len(model_df):,}")
    print(f"국가 수: {model_df['country_fe'].nunique()}")
    print("=" * 76)

    result_rows = []
    coefficient_tables = []

    for outcome, outcome_label in OUTCOMES.items():
        print(f"\n분석 중: {outcome}")

        for specification, formula in [
            ("Baseline (common N)", baseline_formula(outcome)),
            ("Adjusted (common N)", adjusted_formula(outcome)),
        ]:
            result, converged, warning_text = fit_logit(
                formula=formula,
                model_df=model_df,
            )

            result_rows.append(
                extract_public_sector_effect(
                    result=result,
                    outcome=outcome,
                    outcome_label=outcome_label,
                    specification=specification,
                    model_df=model_df,
                    converged=converged,
                    warning_text=warning_text,
                )
            )

            coefficient_tables.append(
                extract_all_coefficients(
                    result=result,
                    outcome=outcome,
                    specification=specification,
                )
            )

    long_results = pd.DataFrame(result_rows)
    wide_main_table = build_main_table(long_results)
    coefficient_table = pd.concat(
        coefficient_tables,
        ignore_index=True,
    )

    sample_table = pd.DataFrame(
        [
            {
                "Metric": "Common complete-case N",
                "Value": int(len(model_df)),
            },
            {
                "Metric": "Non-public sector N",
                "Value": int(
                    (model_df["public_sector"] == 0).sum()
                ),
            },
            {
                "Metric": "Public sector N",
                "Value": int(
                    (model_df["public_sector"] == 1).sum()
                ),
            },
            {
                "Metric": "Countries",
                "Value": int(model_df["country_fe"].nunique()),
            },
            {
                "Metric": "Note",
                "Value": (
                    "All RQ1-RQ4 baseline and adjusted models use "
                    "the same master complete-case sample."
                ),
            },
        ]
    )

    interpretation_table = pd.DataFrame(
        [
            {
                "Item": "Main-table rule",
                "Interpretation": (
                    "Use only the Baseline (common N) and Adjusted "
                    "(common N) estimates in the main manuscript table."
                ),
            },
            {
                "Item": "Why common N matters",
                "Interpretation": (
                    "It ensures that coefficient changes between baseline "
                    "and adjusted models cannot be attributed to changing "
                    "sample composition."
                ),
            },
            {
                "Item": "Old unrestricted baselines",
                "Interpretation": (
                    "Earlier country-FE-only models estimated on N=10,811 "
                    "may be retained only as supplementary or appendix checks, "
                    "not combined with common-N adjusted results in one table."
                ),
            },
        ]
    )

    tables = {
        "01_MainTable_CommonN": wide_main_table,
        "02_LongResults_CommonN": long_results,
        "03_AllCoefficients": coefficient_table,
        "04_CommonSample": sample_table,
        "05_InterpretationGuide": interpretation_table,
    }

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    for sheet_name, dataframe in tables.items():
        save_csv(
            dataframe=dataframe,
            file_name=f"{sheet_name}.csv",
        )

    format_excel(OUTPUT_XLSX)

    print("\n[Main regression table: common complete-case sample]")
    print(wide_main_table.to_string(index=False))

    print("\n" + "=" * 76)
    print("분석 완료")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"Results folder: {OUTPUT_DIR}")
    print("=" * 76)


if __name__ == "__main__":
    main()
